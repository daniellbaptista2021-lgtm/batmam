"""Gerador de planilhas XLSX."""
from __future__ import annotations
import json
from pathlib import Path
from .base import STATIC_DIR, ask_ai, unique_name, file_url

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate(prompt: str) -> dict:
    system = """Voce e um especialista em planilhas. Gere uma estrutura JSON para uma planilha Excel.
Formato EXATO do JSON (sem markdown, sem explicacao):
{
  "title": "Nome da Planilha",
  "sheets": [
    {
      "name": "Nome da Aba",
      "headers": ["Col1", "Col2", "Col3"],
      "rows": [
        ["valor1", "valor2", "valor3"],
        ["valor4", "valor5", "valor6"]
      ],
      "col_widths": [20, 15, 15]
    }
  ]
}
- Preencha com dados de exemplo realistas (minimo 5 linhas)
- Inclua formulas quando fizer sentido (ex: =SUM(B2:B10))
- Retorne APENAS o JSON valido"""

    raw = ask_ai(prompt, system=system, max_tokens=2048)

    # Limpa markdown
    text = raw.strip()
    if text.startswith("```"):
        text = "\n".join(text.split("\n")[1:])
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()

    data = json.loads(text)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for sheet_data in data.get("sheets", [data]):
        ws = wb.create_sheet(title=sheet_data.get("name", "Dados")[:31])
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])
        col_widths = sheet_data.get("col_widths", [])

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        if not col_widths:
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    filename = unique_name(data.get("title", prompt[:30]), ".xlsx")
    out_dir = STATIC_DIR / "files"
    out_dir.mkdir(parents=True, exist_ok=True)
    filepath = out_dir / filename
    wb.save(filepath)

    url = file_url(f"static/files/{filename}")
    size = filepath.stat().st_size

    return {
        "type": "xlsx",
        "name": filename,
        "url": url,
        "size": size,
        "path": str(filepath),
    }
