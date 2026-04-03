"""Clow Web App — FastAPI + WebSocket com UI estilo Claude Code.

Features:
  #18 — Web App (FastAPI + WebSocket com UI completa)
  #19 — Health Check + Monitoring Endpoint
  #24 — Dashboard de Metricas
  #26 — Autenticacao via API Key / Bearer Token
  #27 — Rate Limiting por IP (configurable)
  #28 — CORS configuravel
  #29 — HTTPS/TLS support (via uvicorn ssl)
"""

from __future__ import annotations
import json
import asyncio
import os
import time
import hashlib
import secrets
import logging
from collections import defaultdict
from typing import Any
from pathlib import Path

try:
    from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request, HTTPException, Depends, UploadFile, File, Form
    from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, RedirectResponse
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.staticfiles import StaticFiles
    HAS_FASTAPI = True
except ImportError:
    HAS_FASTAPI = False

import base64
import re as _re
import io

from . import __version__
from . import config

app = FastAPI(
    title="Clow",
    version=__version__,
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json",
) if HAS_FASTAPI else None


# ── Autenticacao ─────────────────────────────────────────────────

def _get_api_keys() -> list[str]:
    """Carrega API keys do settings ou env."""
    settings = config.load_settings()
    keys = settings.get("webapp", {}).get("api_keys", [])
    env_key = os.getenv("CLOW_API_KEY", "")
    if env_key and env_key not in keys:
        keys.append(env_key)
    return keys


def _generate_api_key() -> str:
    """Gera uma nova API key segura."""
    return f"clow_{secrets.token_urlsafe(32)}"


def _verify_api_key(key: str) -> bool:
    """Verifica se uma API key e valida."""
    valid_keys = _get_api_keys()
    if not valid_keys:
        return True  # Sem keys configuradas = sem autenticacao (dev mode)
    return key in valid_keys


async def _auth_dependency(request: Request) -> None:
    """FastAPI dependency para verificar autenticacao."""
    keys = _get_api_keys()
    if not keys:
        return  # Dev mode — sem autenticacao

    # Tenta Authorization header
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:]
        if _verify_api_key(token):
            return

    # Tenta query param
    api_key = request.query_params.get("api_key", "")
    if api_key and _verify_api_key(api_key):
        return

    raise HTTPException(status_code=401, detail="API key invalida ou ausente")


# ── Login / Sessao Web (SQLite backed) ──────────────────────────

from .database import (
    authenticate_user, create_user, get_user_by_email, get_user_by_id,
    list_users, update_user, log_usage, get_user_usage_today, check_limit,
    get_admin_stats, create_conversation, list_conversations, delete_conversation,
    save_message, get_messages, update_conversation_title, PLANS, ADMIN_EMAIL,
)

_web_sessions: dict[str, dict] = {}  # token -> {"email", "user_id", "is_admin", "created"}
_SESSION_TTL = 86400 * 7  # 7 dias


def _create_session(user: dict) -> str:
    token = secrets.token_urlsafe(48)
    _web_sessions[token] = {
        "email": user["email"],
        "user_id": user["id"],
        "is_admin": bool(user.get("is_admin")),
        "plan": user.get("plan", "free"),
        "created": time.time(),
    }
    return token


def _validate_session(token: str) -> dict | None:
    sess = _web_sessions.get(token)
    if not sess:
        return None
    if time.time() - sess["created"] > _SESSION_TTL:
        del _web_sessions[token]
        return None
    return sess


def _get_session_from_request(request: Request) -> str | None:
    """Returns email string for backwards compat."""
    token = request.cookies.get("clow_session", "")
    sess = _validate_session(token)
    return sess["email"] if sess else None


def _get_user_session(request: Request) -> dict | None:
    """Returns full session dict."""
    token = request.cookies.get("clow_session", "")
    return _validate_session(token)


# ── Rate Limiting ────────────────────────────────────────────────

class RateLimiter:
    """Rate limiter por IP com sliding window."""

    def __init__(self, max_requests: int = 60, window_seconds: int = 60):
        self.max_requests = max_requests
        self.window = window_seconds
        self._requests: dict[str, list[float]] = defaultdict(list)

    def is_allowed(self, ip: str) -> bool:
        now = time.time()
        window_start = now - self.window

        # Limpa requests antigos
        self._requests[ip] = [t for t in self._requests[ip] if t > window_start]

        if len(self._requests[ip]) >= self.max_requests:
            return False

        self._requests[ip].append(now)
        return True

    def remaining(self, ip: str) -> int:
        now = time.time()
        window_start = now - self.window
        recent = [t for t in self._requests[ip] if t > window_start]
        return max(0, self.max_requests - len(recent))


_rate_limiter = RateLimiter(max_requests=60, window_seconds=60)
_ws_rate_limiter = RateLimiter(max_requests=10, window_seconds=60)


async def _rate_limit_dependency(request: Request) -> None:
    """FastAPI dependency para rate limiting."""
    client_ip = request.client.host if request.client else "unknown"
    if not _rate_limiter.is_allowed(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Rate limit excedido. Tente novamente em alguns segundos.",
            headers={"Retry-After": "60"},
        )


# ── Setup CORS e Middleware ──────────────────────────────────────

def _setup_middleware():
    """Configura CORS e middlewares de seguranca."""
    if not HAS_FASTAPI or app is None:
        return

    settings = config.load_settings()
    webapp_cfg = settings.get("webapp", {})

    # CORS
    allowed_origins = webapp_cfg.get("cors_origins", ["http://localhost:*", "http://127.0.0.1:*"])
    app.add_middleware(
        CORSMiddleware,
        allow_origins=allowed_origins,
        allow_credentials=True,
        allow_methods=["GET", "POST", "OPTIONS"],
        allow_headers=["*"],
    )


if HAS_FASTAPI:
    _setup_middleware()


def get_app():
    if not HAS_FASTAPI:
        raise RuntimeError("FastAPI nao instalado. Instale com: pip install 'clow[web]'")
    return app


def start_with_tls(host: str = "0.0.0.0", port: int = 8080, certfile: str = "", keyfile: str = ""):
    """Inicia o servidor com TLS/HTTPS se certificados forem fornecidos."""
    import uvicorn

    kwargs: dict[str, Any] = {
        "app": "clow.webapp:app",
        "host": host,
        "port": port,
        "log_level": "info",
    }
    if certfile and keyfile:
        kwargs["ssl_certfile"] = certfile
        kwargs["ssl_keyfile"] = keyfile

    uvicorn.run(**kwargs)


# ── Ações executadas (Feature #24 — tracking) ──────────────────
_recent_actions: list[dict] = []
MAX_RECENT_ACTIONS = 50


def track_action(action: str, details: str = "", status: str = "ok") -> None:
    """Registra ação recente para o dashboard."""
    _recent_actions.append({
        "action": action,
        "details": details[:100],
        "status": status,
        "timestamp": time.time(),
    })
    if len(_recent_actions) > MAX_RECENT_ACTIONS:
        _recent_actions.pop(0)


# ── Feature #19: Health Check ──────────────────────────────────

def _get_health_data() -> dict:
    """Coleta status de todos os componentes."""
    from .memory import list_memories
    from .cron import get_cron_manager
    from .triggers import get_trigger_server
    from .tasks import get_task_manager

    memories = list_memories()
    cron = get_cron_manager()
    trigger = get_trigger_server()
    tasks = get_task_manager()

    mem_by_type: dict[str, int] = {}
    for m in memories:
        t = m.get("type", "general")
        mem_by_type[t] = mem_by_type.get(t, 0) + 1

    all_tasks = tasks.list_all()
    tasks_by_status: dict[str, int] = {}
    for t in all_tasks:
        s = t.status.value
        tasks_by_status[s] = tasks_by_status.get(s, 0) + 1

    cron_jobs = cron.list_all()
    active_crons = [j for j in cron_jobs if j.active]

    return {
        "status": "healthy",
        "version": __version__,
        "uptime_info": time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()),
        "components": {
            "memory": {
                "status": "ok",
                "total": len(memories),
                "by_type": mem_by_type,
            },
            "cron": {
                "status": "ok",
                "total_jobs": len(cron_jobs),
                "active_jobs": len(active_crons),
                "jobs": [
                    {
                        "id": j.id,
                        "prompt": j.prompt[:50],
                        "interval": cron.format_interval(j.interval_seconds),
                        "active": j.active,
                        "run_count": j.run_count,
                        "last_run": j.last_run,
                        "next_run": j.last_run + j.interval_seconds if j.last_run else j.created_at + j.interval_seconds,
                    }
                    for j in cron_jobs
                ],
            },
            "triggers": {
                "status": "ok" if trigger.running else "stopped",
                "running": trigger.running,
                "port": trigger.port,
                "results_count": len(trigger.list_results()),
            },
            "tasks": {
                "status": "ok",
                "total": len(all_tasks),
                "by_status": tasks_by_status,
            },
        },
        "recent_actions": _recent_actions[-10:],
    }


def _build_multimodal_message(text: str, file_data: dict) -> Any:
    """Monta mensagem multimodal (content blocks) para API Anthropic."""
    ftype = file_data.get("type", "")
    content_blocks = []

    if ftype == "image":
        content_blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": file_data.get("media_type", "image/jpeg"),
                "data": file_data.get("base64", ""),
            },
        })
        if text:
            content_blocks.append({"type": "text", "text": text})
        else:
            content_blocks.append({"type": "text", "text": "Analise esta imagem e descreva o que você vê."})

    elif ftype == "pdf":
        content_blocks.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": file_data.get("base64", ""),
            },
        })
        if text:
            content_blocks.append({"type": "text", "text": text})
        else:
            content_blocks.append({"type": "text", "text": "Analise este PDF e resuma o conteúdo."})

    elif ftype == "audio":
        transcription = file_data.get("transcription", "")
        if transcription and not transcription.startswith("[Erro"):
            prompt = f"[Audio transcrito do usuario]\nTranscricao: {transcription}"
            if text:
                prompt += f"\n\nMensagem adicional: {text}"
            return prompt
        else:
            return text or "[O usuario enviou um audio mas a transcricao nao esta disponivel]"

    elif ftype in ("spreadsheet", "document", "code"):
        extracted = file_data.get("extracted_text", "")
        fname = file_data.get("file_name", "arquivo")
        if ftype == "spreadsheet":
            prefix = f"[Planilha: {fname}]\n\nDados:\n{extracted}"
        elif ftype == "code":
            lang = file_data.get("language", "text")
            prefix = f"[Codigo: {fname}]\n\n```{lang}\n{extracted}\n```"
        else:
            prefix = f"[Documento: {fname}]\n\nConteudo:\n{extracted}"
        if text:
            prefix += f"\n\nPedido do usuario: {text}"
        return prefix

    else:
        return text or f"[O usuario enviou um arquivo: {file_data.get('file_name', 'arquivo')}]"

    return content_blocks


# ── Templates — carrega HTML de arquivos separados ─────────────

_TEMPLATE_DIR = Path(__file__).parent / "templates"
_STATIC_DIR = Path(__file__).parent / "static"


def _load_template(name: str) -> str:
    """Carrega template HTML do diretorio templates/."""
    path = _TEMPLATE_DIR / name
    try:
        return path.read_text(encoding="utf-8")
    except FileNotFoundError:
        logger.error("Template not found: %s", path)
        return f"<h1>Template {name} not found</h1>"


# Lazy-loaded templates (cached on first access)
_template_cache: dict[str, str] = {}


def _get_template(name: str) -> str:
    """Retorna template com cache."""
    if name not in _template_cache:
        _template_cache[name] = _load_template(name)
    return _template_cache[name]


# Template accessors
def _webapp_html() -> str:
    return _get_template("chat.html")

def _dashboard_html() -> str:
    return _get_template("dashboard.html")

def _login_html() -> str:
    return _get_template("login.html")

def _admin_html() -> str:
    return _get_template("admin.html")






if HAS_FASTAPI:
    # ── Login routes (sem auth) ──────────────────────────────────
    @app.get("/login", response_class=HTMLResponse)
    async def login_page(request: Request):
        if _get_session_from_request(request):
            return RedirectResponse("/", status_code=302)
        html = _login_html().replace("__ERROR_CLASS__", "").replace("__ERROR_MSG__", "")
        return HTMLResponse(html)

    @app.post("/login")
    async def login_submit(request: Request):
        form = await request.form()
        email = form.get("email", "").strip().lower()
        password = form.get("password", "")

        user = authenticate_user(email, password)
        if user:
            token = _create_session(user)
            resp = RedirectResponse("/", status_code=302)
            resp.set_cookie(
                "clow_session", token,
                max_age=_SESSION_TTL, httponly=True,
                samesite="lax", secure=False,
            )
            return resp

        html = _login_html().replace("__ERROR_CLASS__", "show").replace("__ERROR_MSG__", "Email ou senha incorretos")
        return HTMLResponse(html, status_code=401)

    @app.get("/logout")
    async def logout(request: Request):
        token = request.cookies.get("clow_session", "")
        if token in _web_sessions:
            del _web_sessions[token]
        resp = RedirectResponse("/login", status_code=302)
        resp.delete_cookie("clow_session")
        return resp

    # ── Mount static assets (CSS/JS extraídos) ─────────────────
    _css_dir = _STATIC_DIR / "css"
    _js_dir = _STATIC_DIR / "js"
    if _css_dir.is_dir():
        app.mount("/static/css", StaticFiles(directory=str(_css_dir)), name="css_static")
    if _js_dir.is_dir():
        app.mount("/static/js", StaticFiles(directory=str(_js_dir)), name="js_static")

    # ── Protected routes ─────────────────────────────────────────
    @app.get("/", response_class=HTMLResponse)
    async def index(request: Request):
        if not _get_session_from_request(request):
            return RedirectResponse("/login", status_code=302)
        return _webapp_html()

    @app.get("/dashboard", response_class=HTMLResponse)
    async def dashboard_page(request: Request):
        if not _get_session_from_request(request):
            return RedirectResponse("/login", status_code=302)
        return _dashboard_html()

    @app.get("/admin", response_class=HTMLResponse)
    async def admin_page(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return RedirectResponse("/login", status_code=302)
        return _admin_html()

    # ── PWA Routes (System Clow App) ──────────────────────────────
    @app.get("/pwa", response_class=HTMLResponse)
    async def pwa_index():
        """Página principal do PWA."""
        static_dir = Path(__file__).parent.parent / "static"
        if (static_dir / "index.html").exists():
            with open(static_dir / "index.html") as f:
                return f.read()
        return HTMLResponse("<h1>System Clow</h1><p>PWA app</p>")

    @app.get("/static/manifest.json")
    async def manifest():
        """Manifest do PWA."""
        static_dir = Path(__file__).parent.parent / "static"
        manifest_path = static_dir / "manifest.json"
        if manifest_path.exists():
            return FileResponse(manifest_path, media_type="application/manifest+json")
        return JSONResponse({"name": "System Clow", "short_name": "Clow"})

    @app.get("/static/service-worker.js")
    async def service_worker():
        """Service Worker para PWA."""
        static_dir = Path(__file__).parent.parent / "static"
        sw_path = static_dir / "service-worker.js"
        if sw_path.exists():
            return FileResponse(sw_path, media_type="application/javascript")
        return JSONResponse({"error": "Service Worker not found"}, status_code=404)

    @app.get("/static/{file_path:path}")
    async def static_files(file_path: str):
        """Serve arquivos estáticos (CSS, JS, imagens, etc)."""
        static_dir = Path(__file__).parent.parent / "static"
        full_path = (static_dir / file_path).resolve()
        
        # Security: previne path traversal
        if not str(full_path).startswith(str(static_dir)):
            return JSONResponse({"error": "Forbidden"}, status_code=403)
        
        if full_path.exists() and full_path.is_file():
            return FileResponse(full_path)
        return JSONResponse({"error": "Not found"}, status_code=404)

    # Feature #19: Health Check (publico — sem auth)
    @app.get("/health", dependencies=[Depends(_rate_limit_dependency)])
    async def health():
        return JSONResponse(_get_health_data())

    # API endpoints protegidos
    @app.get("/api/v1/status", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_status():
        """Status completo da API com metricas."""
        return JSONResponse({
            "version": __version__,
            "status": "ok",
            **_get_health_data(),
        })

    @app.post("/api/v1/generate-key", dependencies=[Depends(_auth_dependency)])
    async def generate_key():
        """Gera nova API key para autenticacao."""
        key = _generate_api_key()
        return JSONResponse({"api_key": key, "note": "Adicione esta key em settings.json > webapp > api_keys"})

    @app.get("/api/v1/metrics", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_metrics():
        """Retorna metricas coletadas (counters, gauges, histograms)."""
        from .logging import metrics
        return JSONResponse(metrics.snapshot())

    @app.get("/api/v1/sessions", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_sessions():
        """Lista sessoes salvas."""
        from .session import list_sessions
        return JSONResponse({"sessions": list_sessions()})

    @app.get("/api/v1/memory", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_memory():
        """Lista memorias persistidas."""
        from .memory import list_memories
        return JSONResponse({"memories": list_memories()})

    @app.get("/api/v1/tools", dependencies=[Depends(_auth_dependency)])
    async def api_tools():
        """Lista todas as ferramentas disponiveis."""
        from .tools.base import create_default_registry
        registry = create_default_registry()
        tools = [{"name": t.name, "description": t.description} for t in registry.all_tools()]
        return JSONResponse({"tools": tools, "count": len(tools)})

    # ── API: Conversations ──────────────────────────────────────────

    @app.get("/api/v1/conversations")
    async def api_conversations(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        convs = list_conversations(sess["user_id"])
        return JSONResponse({"conversations": convs})

    @app.post("/api/v1/conversations")
    async def api_create_conversation(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        cid = create_conversation(sess["user_id"])
        return JSONResponse({"id": cid})

    @app.get("/api/v1/conversations/{conv_id}/messages")
    async def api_get_messages(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        msgs = get_messages(conv_id)
        return JSONResponse({"messages": msgs})

    @app.delete("/api/v1/conversations/{conv_id}")
    async def api_delete_conversation(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        delete_conversation(sess["user_id"], conv_id)
        return JSONResponse({"ok": True})

    @app.post("/api/v1/conversations/{conv_id}/title")
    async def api_update_conv_title(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        body = await request.json()
        title = body.get("title", "")[:50]
        if title:
            update_conversation_title(conv_id, title)
        return JSONResponse({"ok": True})

    # ── API: Upload de Arquivos ─────────────────────────────────────

    _ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    _ALLOWED_DOC_EXT = {".pdf", ".docx", ".doc", ".txt", ".md"}
    _ALLOWED_SHEET_EXT = {".xlsx", ".xls", ".csv"}
    _ALLOWED_CODE_EXT = {".py", ".js", ".html", ".css", ".json", ".ts", ".jsx", ".tsx"}
    _ALLOWED_AUDIO_EXT = {".webm", ".mp3", ".ogg", ".wav", ".m4a"}
    _BLOCKED_EXT = {".exe", ".bat", ".sh", ".cmd", ".com", ".msi", ".scr", ".ps1"}
    _MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
    _UPLOAD_LIMITS = {"free": 5, "basic": 20, "pro": 100, "unlimited": 0}
    _upload_counts: dict[str, dict] = {}  # user_id -> {"date": "YYYY-MM-DD", "count": N}

    def _check_upload_limit(user_id: str, plan: str) -> bool:
        import datetime
        today = datetime.date.today().isoformat()
        rec = _upload_counts.get(user_id, {"date": "", "count": 0})
        if rec["date"] != today:
            rec = {"date": today, "count": 0}
        limit = _UPLOAD_LIMITS.get(plan, 5)
        if limit == 0:
            return True
        return rec["count"] < limit

    def _inc_upload_count(user_id: str):
        import datetime
        today = datetime.date.today().isoformat()
        rec = _upload_counts.get(user_id, {"date": "", "count": 0})
        if rec["date"] != today:
            rec = {"date": today, "count": 0}
        rec["count"] += 1
        _upload_counts[user_id] = rec

    def _sanitize_filename(name: str) -> str:
        name = _re.sub(r'[^\w\-_. ]', '', name)
        return name.strip()[:100] or "arquivo"

    def _format_size(size: int) -> str:
        if size > 1024 * 1024:
            return f"{size / (1024*1024):.1f} MB"
        if size > 1024:
            return f"{size / 1024:.1f} KB"
        return f"{size} bytes"

    def _extract_text_docx(data: bytes) -> str:
        try:
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[Erro ao ler DOCX: {e}]"

    def _extract_text_xlsx(data: bytes) -> tuple[str, int]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            total_rows = 0
            for sheet in wb.worksheets:
                lines.append(f"\n### Aba: {sheet.title}\n")
                rows = list(sheet.iter_rows(values_only=True))
                total_rows += len(rows)
                if rows:
                    # Header
                    header = [str(c) if c is not None else "" for c in rows[0]]
                    lines.append("| " + " | ".join(header) + " |")
                    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                    for row in rows[1:51]:
                        cells = [str(c) if c is not None else "" for c in row]
                        lines.append("| " + " | ".join(cells) + " |")
                    if len(rows) > 51:
                        lines.append(f"\n... e mais {len(rows) - 51} linhas")
            wb.close()
            return "\n".join(lines), total_rows
        except Exception as e:
            return f"[Erro ao ler XLSX: {e}]", 0

    def _extract_text_csv(data: bytes) -> tuple[str, int]:
        import csv
        try:
            text = data.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            lines = []
            if rows:
                header = rows[0]
                lines.append("| " + " | ".join(header) + " |")
                lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                for row in rows[1:51]:
                    lines.append("| " + " | ".join(row) + " |")
                if len(rows) > 51:
                    lines.append(f"\n... e mais {len(rows) - 51} linhas")
            return "\n".join(lines), len(rows)
        except Exception as e:
            return f"[Erro ao ler CSV: {e}]", 0

    def _resize_image(data: bytes, max_px: int = 2000) -> bytes:
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(data))
            if max(img.size) > max_px:
                img.thumbnail((max_px, max_px), Image.LANCZOS)
                buf = io.BytesIO()
                fmt = img.format or "JPEG"
                if fmt.upper() == "WEBP":
                    img.save(buf, format="WEBP", quality=85)
                else:
                    img.save(buf, format="JPEG", quality=85)
                return buf.getvalue()
        except Exception:
            pass
        return data

    async def _transcribe_audio(file_path: str) -> str:
        """Transcricao de audio — agora feita no frontend via Web Speech API.
        Backend nao usa mais OpenAI Whisper. Retorna vazio para fallback."""
        return ""

    @app.post("/api/v1/upload")
    async def api_upload(request: Request, file: UploadFile = File(...), message: str = Form("")):
        """Upload de arquivo com processamento automatico."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        user_id = sess["user_id"]
        user_plan = sess.get("plan", "free")

        if not _check_upload_limit(user_id, user_plan):
            limit = _UPLOAD_LIMITS.get(user_plan, 5)
            return JSONResponse({"error": f"Limite de {limit} uploads/dia atingido. Faca upgrade do plano."}, status_code=429)

        # Lê arquivo
        data = await file.read()
        if len(data) > _MAX_FILE_SIZE:
            return JSONResponse({"error": "Arquivo muito grande. Maximo: 20MB"}, status_code=413)

        original_name = file.filename or "arquivo"
        safe_name = _sanitize_filename(original_name)
        ext = Path(original_name).suffix.lower()

        if ext in _BLOCKED_EXT:
            return JSONResponse({"error": f"Tipo de arquivo não permitido: {ext}"}, status_code=400)

        # Salva arquivo
        upload_dir = Path(__file__).parent.parent / "static" / "uploads" / user_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        saved_name = f"{ts}_{safe_name}"
        saved_path = upload_dir / saved_name
        saved_path.write_bytes(data)

        _inc_upload_count(user_id)
        file_size = len(data)
        file_url = f"/static/uploads/{user_id}/{saved_name}"

        result: dict[str, Any] = {
            "ok": True,
            "file_name": safe_name,
            "file_url": file_url,
            "file_size": _format_size(file_size),
            "file_ext": ext,
            "type": "unknown",
        }

        # Processa por tipo
        if ext in _ALLOWED_IMAGE_EXT:
            resized = _resize_image(data)
            # Detecta media_type real pelos magic bytes (apos resize)
            if resized[:3] == b'\xff\xd8\xff':
                media_type = 'image/jpeg'
            elif resized[:4] == b'\x89PNG':
                media_type = 'image/png'
            elif resized[:4] == b'GIF8':
                media_type = 'image/gif'
            elif resized[:4] == b'RIFF' and len(resized) > 11 and resized[8:12] == b'WEBP':
                media_type = 'image/webp'
            else:
                media_type = 'image/jpeg'
            b64 = base64.b64encode(resized).decode("ascii")
            result["type"] = "image"
            result["media_type"] = media_type
            result["base64"] = b64

        elif ext == ".pdf":
            b64 = base64.b64encode(data).decode("ascii")
            # Conta paginas
            pages = 0
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(data))
                pages = len(reader.pages)
            except Exception:
                pass
            if pages > 50:
                result["warning"] = f"PDF com {pages} paginas. Apenas as primeiras 50 serao analisadas."
            result["type"] = "pdf"
            result["media_type"] = "application/pdf"
            result["base64"] = b64
            result["pages"] = pages

        elif ext in _ALLOWED_SHEET_EXT:
            if ext == ".csv":
                text, rows = _extract_text_csv(data)
            else:
                text, rows = _extract_text_xlsx(data)
            result["type"] = "spreadsheet"
            result["extracted_text"] = text
            result["rows"] = rows

        elif ext in _ALLOWED_DOC_EXT:
            if ext == ".docx":
                text = _extract_text_docx(data)
            else:
                text = data.decode("utf-8", errors="replace")
            words = len(text.split())
            if words > 5000:
                text = " ".join(text.split()[:5000]) + f"\n\n... (truncado, {words} palavras total)"
            result["type"] = "document"
            result["extracted_text"] = text
            result["words"] = words

        elif ext in _ALLOWED_CODE_EXT:
            text = data.decode("utf-8", errors="replace")
            lang_map = {".py": "python", ".js": "javascript", ".ts": "typescript",
                        ".html": "html", ".css": "css", ".json": "json", ".jsx": "jsx", ".tsx": "tsx"}
            result["type"] = "code"
            result["extracted_text"] = text
            result["language"] = lang_map.get(ext, "text")

        elif ext in _ALLOWED_AUDIO_EXT:
            # Transcreve audio
            loop = asyncio.get_event_loop()
            transcription = await _transcribe_audio(str(saved_path))
            result["type"] = "audio"
            result["transcription"] = transcription
            result["has_transcription"] = bool(transcription and not transcription.startswith("[Erro"))

        else:
            result["type"] = "file"

        track_action("file_upload", f"{ext}: {safe_name}", "ok")
        return JSONResponse(result)

    # ── API: Usage ───────────────────────────────────────────────────

    @app.get("/api/v1/usage")
    async def api_usage(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        usage = get_user_usage_today(sess["user_id"])
        plan = PLANS.get(sess.get("plan", "free"), PLANS["free"])
        return JSONResponse({
            "usage": usage,
            "plan": sess.get("plan", "free"),
            "plan_label": plan["label"],
            "daily_limit": plan["daily_tokens"],
        })

    @app.get("/api/v1/me")
    async def api_me(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        plan = sess.get("plan", "free")
        is_admin = sess.get("is_admin", False)
        if is_admin:
            models = ["claude-code"]
        else:
            models = ["haiku"]
            if plan in ("pro", "unlimited"):
                models.append("sonnet")
        return JSONResponse({
            "email": sess["email"],
            "user_id": sess["user_id"],
            "is_admin": is_admin,
            "plan": plan,
            "available_models": models,
        })

    # ── API: Admin ───────────────────────────────────────────────────

    @app.get("/api/v1/admin/stats")
    async def api_admin_stats(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        return JSONResponse(get_admin_stats())

    @app.get("/api/v1/admin/users")
    async def api_admin_users(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        return JSONResponse({"users": list_users()})

    @app.post("/api/v1/admin/users/{user_id}")
    async def api_admin_update_user(user_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        body = await request.json()
        update_user(user_id, **body)
        return JSONResponse({"ok": True})

    @app.post("/api/v1/admin/create-user")
    async def api_admin_create_user(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        body = await request.json()
        email = body.get("email", "").strip().lower()
        password = body.get("password", "")
        name = body.get("name", "")
        plan = body.get("plan", "free")
        if not email or len(password) < 6:
            return JSONResponse({"error": "Email e senha (min 6 chars) obrigatorios"}, status_code=400)
        user = create_user(email, password, name)
        if not user:
            return JSONResponse({"error": "Email ja cadastrado"}, status_code=400)
        if plan != "free":
            update_user(user["id"], plan=plan)
        return JSONResponse({"ok": True, "user": user})

    # ── API: Missions ────────────────────────────────────────────────
    _mission_progress: dict[str, list] = {}  # mission_id -> [events]

    @app.post("/api/v1/missions/plan")
    async def api_mission_plan(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        body = await request.json()
        description = body.get("description", "").strip()
        if not description:
            return JSONResponse({"error": "Descricao vazia"}, status_code=400)

        from .agents.mission_engine import plan_mission
        loop = asyncio.get_event_loop()
        try:
            plan = await loop.run_in_executor(None, plan_mission, description)
            return JSONResponse({"plan": plan})
        except Exception as e:
            return JSONResponse({"error": str(e)[:300]}, status_code=500)

    @app.post("/api/v1/missions/start")
    async def api_mission_start(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        body = await request.json()
        description = body.get("description", "")
        plan_data = body.get("plan", {})
        title = plan_data.get("title", description[:60])
        steps = plan_data.get("steps", [])

        if not steps:
            return JSONResponse({"error": "Plano sem etapas"}, status_code=400)

        from .database import create_mission
        mission_id = create_mission(sess["user_id"], title, description, steps)
        _mission_progress[mission_id] = []

        async def on_progress(mid, event_type, data):
            _mission_progress.setdefault(mid, []).append({
                "type": event_type, "data": data, "time": time.time()
            })

        from .agents.mission_engine import execute_mission
        asyncio.create_task(execute_mission(mission_id, sess["user_id"], on_progress))

        return JSONResponse({"mission_id": mission_id, "title": title, "total_steps": len(steps)})

    @app.get("/api/v1/missions/{mission_id}/progress")
    async def api_mission_progress(mission_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        after = float(request.query_params.get("after", "0"))
        events = _mission_progress.get(mission_id, [])
        new_events = [e for e in events if e["time"] > after]

        from .database import get_mission
        mission = get_mission(mission_id)
        status = mission["status"] if mission else "unknown"

        return JSONResponse({"status": status, "events": new_events})

    @app.get("/api/v1/missions")
    async def api_list_missions(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        from .database import list_missions
        return JSONResponse({"missions": list_missions(sess["user_id"])})

    @app.get("/api/v1/missions/{mission_id}")
    async def api_get_mission(mission_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        from .database import get_mission
        m = get_mission(mission_id)
        if not m:
            return JSONResponse({"error": "Missão não encontrada"}, status_code=404)
        return JSONResponse({"mission": m})

    # ── HTTP Chat Fallback (para mobile sem WebSocket) ────────────
    _http_sessions: dict[str, Any] = {}

    @app.post("/api/v1/chat", dependencies=[Depends(_rate_limit_dependency)])
    async def api_chat(request: Request):
        """Chat HTTP com deteccao automatica de geracao de arquivos."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        # Checa limite de uso
        allowed, pct = check_limit(sess["user_id"])
        if not allowed:
            return JSONResponse({
                "session_id": "",
                "response": "Voce atingiu seu limite diario de tokens. Volte amanha ou faca upgrade do seu plano.\n\nUse `/plan` para ver seu plano atual.",
                "tools": [], "file": None,
            })

        from .agent import Agent
        import uuid

        body = await request.json()
        content = body.get("content", "").strip()
        conv_id = body.get("conversation_id", "")
        session_id = body.get("session_id", "")
        chosen_model = body.get("model", "haiku")  # haiku ou sonnet
        file_data = body.get("file_data")

        if not content and not file_data:
            return JSONResponse({"error": "content vazio"}, status_code=400)

        user_email = sess["email"]
        user_id = sess["user_id"]
        user_plan = sess.get("plan", "free")
        is_admin = sess.get("is_admin", False)

        # Admin SEMPRE usa Claude Code CLI (conta Max, gratis)
        if is_admin:
            from .claude_code_bridge import ask_claude_code, log_claude_code_usage
            track_action("user_message_claude_code", content[:60])

            if conv_id:
                save_message(conv_id, "user", content)

            loop = asyncio.get_event_loop()
            response_text, elapsed = await loop.run_in_executor(None, lambda: ask_claude_code(content, "/root/clow/workspace", conv_id))

            log_claude_code_usage(user_id, content, elapsed)
            track_action("claude_code_response", response_text[:60] if response_text else "")

            if conv_id:
                save_message(conv_id, "assistant", response_text)

            return JSONResponse({
                "session_id": session_id or str(uuid.uuid4())[:8],
                "response": response_text,
                "tools": [], "file": None,
            })

        # Nao-admin: valida modelo pelo plano
        from .generators.base import MODELS as AI_MODELS
        allowed_models = ["haiku"]
        if user_plan in ("pro", "unlimited"):
            allowed_models.append("sonnet")
        if chosen_model not in allowed_models:
            chosen_model = "haiku"
        model_id = AI_MODELS.get(chosen_model, AI_MODELS["haiku"])
        track_action("user_message_http", content[:60])

        # Salva mensagem do usuario no historico
        if conv_id:
            save_message(conv_id, "user", content)

        # ── Comandos internos ──
        if content.startswith("/"):
            cmd_lower = content.lower().strip()
            cmd_resp = None

            if cmd_lower.startswith("/skills"):
                from .skills.loader import format_skills_list
                cat = content[7:].strip()
                cmd_resp = format_skills_list(cat)
            elif cmd_lower == "/memories":
                from .memory_web import format_memories_list
                cmd_resp = format_memories_list(user_id)
            elif cmd_lower.startswith("/forget"):
                from .memory_web import forget_memory
                kw = content[7:].strip()
                cmd_resp = forget_memory(user_id, kw) if kw else "Use: `/forget palavra-chave`"
            elif cmd_lower.startswith("/mission"):
                mission_desc = content[8:].strip() if len(content) > 8 else ""
                if not mission_desc:
                    cmd_resp = (
                        "## Missões Autônomas\n\n"
                        "Descreva uma missao complexa e o Clow executa sozinho:\n\n"
                        "**Exemplos:**\n"
                        "- `/mission Cria um site completo para uma pizzaria com cardapio e contato`\n"
                        "- `/mission Campanha de tráfego para seguro de vida com landing page e copies`\n"
                        "- `/mission Setup digital completo para barbearia`\n\n"
                        "O Clow vai planejar, mostrar as etapas, e executar tudo automaticamente."
                    )
                else:
                    # Planeja e inicia missao
                    from .agents.mission_engine import plan_mission
                    loop = asyncio.get_event_loop()
                    try:
                        plan_data = await loop.run_in_executor(None, plan_mission, mission_desc)
                        steps = plan_data.get("steps", [])
                        title = plan_data.get("title", mission_desc[:60])
                        est = plan_data.get("estimated_minutes", 5)

                        # Cria e inicia missao
                        from .database import create_mission as db_create_mission
                        mid = db_create_mission(user_id, title, mission_desc, steps)
                        _mission_progress[mid] = []

                        async def on_progress(m_id, evt, data):
                            _mission_progress.setdefault(m_id, []).append({"type": evt, "data": data, "time": time.time()})

                        from .agents.mission_engine import execute_mission
                        asyncio.create_task(execute_mission(mid, user_id, on_progress))

                        # Mostra plano e inicia
                        steps_text = "\n".join(f"{i+1}. {s.get('title', '?')}" for i, s in enumerate(steps))
                        cmd_resp = (
                            f"## Missão Iniciada\n\n"
                            f"**{title}**\n\n"
                            f"### Plano ({len(steps)} etapas, ~{est} min):\n{steps_text}\n\n"
                            f"Executando em background... Acompanhe o progresso abaixo."
                        )

                        # Retorna com mission_id pra frontend fazer polling
                        if conv_id:
                            save_message(conv_id, "assistant", cmd_resp)
                        return JSONResponse({
                            "session_id": session_id or str(uuid.uuid4())[:8],
                            "response": cmd_resp,
                            "tools": [], "file": None,
                            "mission": {"id": mid, "title": title, "total_steps": len(steps)},
                        })
                    except Exception as e:
                        cmd_resp = f"Erro ao planejar missao: {str(e)[:200]}"

            elif cmd_lower == "/help":
                cmd_resp = (
                    "## Comandos Disponiveis\n\n"
                    "| Comando | Descricao |\n|---------|----------|\n"
                    "| `/mission X` | Iniciar missao autonoma |\n"
                    "| `/skills` | Listar skills disponiveis |\n"
                    "| `/memories` | Ver memorias salvas |\n"
                    "| `/forget X` | Esquecer memoria |\n"
                    "| `/connect` | Conectar servico externo |\n"
                    "| `/connections` | Ver conexões ativas |\n"
                    "| `/disconnect X` | Desconectar servico |\n"
                    "| `/usage` | Ver consumo de tokens hoje |\n"
                    "| `/plan` | Ver plano atual e limites |\n"
                    "| `/help` | Esta lista de comandos |\n\n"
                    "**Missões:** `/mission cria um site completo para pizzaria`\n\n"
                    "**Geracao de arquivos:** peca naturalmente (ex: 'cria uma planilha de vendas')\n\n"
                    "**Integracoes:** pergunte direto (ex: 'mostra minhas campanhas meta ads')"
                )
            elif cmd_lower == "/usage":
                usage = get_user_usage_today(user_id)
                plan_info = PLANS.get(sess.get("plan", "free"), PLANS["free"])
                limit = plan_info["daily_tokens"]
                used = usage["total_tokens"]
                pct_str = f"{(used/limit*100):.0f}%" if limit > 0 else "ilimitado"
                cmd_resp = (
                    f"## Seu Consumo Hoje\n\n"
                    f"- Tokens usados: **{used:,}**\n"
                    f"- Limite diario: **{limit:,}** ({pct_str})\n"
                    f"- Requests: **{usage['requests']}**\n"
                    f"- Custo estimado: **${usage['total_cost']:.4f}**"
                )
            elif cmd_lower == "/plan":
                plan_info = PLANS.get(sess.get("plan", "free"), PLANS["free"])
                cmd_resp = (
                    f"## Seu Plano: {plan_info['label']}\n\n"
                    f"- Limite diario: **{plan_info['daily_tokens']:,} tokens**\n\n"
                    "**Planos disponiveis:**\n"
                    "| Plano | Tokens/dia |\n|-------|------------|\n"
                    "| Free | 50.000 |\n| Basic | 200.000 |\n| Pro | 1.000.000 |\n| Unlimited | Sem limite |"
                )

            if cmd_resp:
                if conv_id:
                    save_message(conv_id, "assistant", cmd_resp)
                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": cmd_resp, "tools": [], "file": None,
                })

            # /connect, /disconnect, /connections
            from .integrations.command_handler import handle_command
            cmd_result = handle_command(content, user_email)
            if cmd_result:
                if conv_id:
                    save_message(conv_id, "assistant", cmd_result["response"])
                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": cmd_result["response"],
                    "tools": [], "file": None,
                })

        # ── Detecta pedidos de integracao (meta ads, supabase, etc) ──
        from .integrations.command_handler import detect_integration_request
        int_result = detect_integration_request(content, user_email)
        if int_result:
            return JSONResponse({
                "session_id": session_id or str(uuid.uuid4())[:8],
                "response": int_result["response"],
                "tools": [],
                "file": None,
            })

        # ── Detecta geracao de arquivo ──
        from .generators.dispatcher import detect, run_generator
        gen_module, gen_type = detect(content)

        if gen_module:
            loop = asyncio.get_event_loop()
            try:
                result = await loop.run_in_executor(None, run_generator, gen_module, content, model_id, user_id)
                track_action("file_generated", f"{gen_type}: {result.get('name', '')}", "ok")

                if result.get("type") == "text":
                    return JSONResponse({
                        "session_id": session_id or str(uuid.uuid4())[:8],
                        "response": result["content"],
                        "tools": [],
                        "file": None,
                    })

                # Formata tamanho
                size_raw = result.get("size", 0)
                if isinstance(size_raw, str):
                    size_str = size_raw
                elif isinstance(size_raw, (int, float)) and size_raw > 1024 * 1024:
                    size_str = f"{size_raw / (1024*1024):.1f} MB"
                elif isinstance(size_raw, (int, float)) and size_raw > 1024:
                    size_str = f"{size_raw / 1024:.1f} KB"
                else:
                    size_str = f"{size_raw} bytes"

                type_labels = {
                    "landing_page": "Landing Page",
                    "app": "Web App",
                    "xlsx": "Planilha Excel",
                    "docx": "Documento Word",
                    "pptx": "Apresentação PowerPoint",
                }
                type_label = type_labels.get(result["type"], result["type"])
                msg = f"Pronto! Aqui esta seu arquivo:\n\n**{type_label}** — {result['name']} ({size_str})"

                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": msg,
                    "tools": [],
                    "file": {
                        "type": result["type"],
                        "name": result["name"],
                        "url": result["url"],
                        "size": size_str,
                    },
                })
            except Exception as e:
                track_action("file_gen_error", str(e)[:60], "error")
                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": f"Erro ao gerar arquivo: {str(e)}",
                    "tools": [],
                    "file": None,
                }, status_code=500)

        # ── Injetar skills no prompt ──
        from .skills.loader import build_skill_prompt
        skill_context = build_skill_prompt(content)
        if skill_context:
            content = f"[CONTEXTO DE SKILLS ATIVAS - siga estas instrucoes]\n{skill_context}\n[FIM DO CONTEXTO]\n\nPedido do usuario: {content}"

        # ── Chat normal via Agent ──
        session_key = f"{session_id}_{chosen_model}"
        if session_id and session_key in _http_sessions:
            agent = _http_sessions[session_key]["agent"]
        else:
            session_id = str(uuid.uuid4())[:8]
            session_key = f"{session_id}_{chosen_model}"
            agent = Agent(cwd=os.getcwd(), model=model_id, auto_approve=True)
            _http_sessions[session_key] = {"agent": agent, "last_used": time.time()}

        _http_sessions[session_key]["last_used"] = time.time()

        now = time.time()
        stale = [k for k, v in _http_sessions.items() if now - v["last_used"] > 1800]
        for k in stale:
            del _http_sessions[k]

        loop = asyncio.get_event_loop()
        collected_text: list[str] = []
        tools_used: list[dict] = []

        def on_text_delta(delta: str):
            collected_text.append(delta)

        def on_tool_call(name: str, args: dict):
            tools_used.append({"name": name, "args": args, "status": "running", "output": ""})
            track_action("tool_call", name, "running")

        def on_tool_result(name: str, status: str, output: str):
            for t in tools_used:
                if t["name"] == name and t["status"] == "running":
                    t["status"] = status
                    t["output"] = output[:500]
                    break
            track_action("tool_result", f"{name}: {status}", status)

        agent.on_text_delta = on_text_delta
        agent.on_text_done = lambda t: None
        agent.on_tool_call = on_tool_call
        agent.on_tool_result = on_tool_result

        # Monta mensagem multimodal se tem arquivo
        if file_data:
            user_msg = _build_multimodal_message(content, file_data)
        else:
            user_msg = content

        try:
            result = await loop.run_in_executor(None, agent.run_turn, user_msg)
        except Exception as e:
            return JSONResponse({
                "session_id": session_id,
                "error": str(e),
            }, status_code=500)

        response_text = "".join(collected_text) if collected_text else (result or "")
        track_action("agent_response_http", response_text[:60] if response_text else "")

        return JSONResponse({
            "session_id": session_id,
            "response": response_text,
            "tools": tools_used,
            "file": None,
        })

    # ── Helpers para geração de imagens ────────────────────────────

    def _should_generate_image(content: str) -> bool:
        """Detecta se o pedido é para gerar imagem."""
        keywords = [
            "gera imagem", "cria imagem", "faz uma imagem", "gerar imagem",
            "criar imagem", "desenha", "ilustração", "arte", "visual",
            "criativo visual", "banner", "thumbnail", "foto", "picture", "image",
            "criativo pra", "criativo de", "criativo para"
        ]
        content_lower = content.lower()
        return any(kw in content_lower for kw in keywords)

    async def _process_image_request(content: str, agent) -> tuple[str | None, str | None, str]:
        """
        Processa pedido de imagem:
        1. Otimiza prompt em inglês via Claude
        2. Gera imagem via Pollinations
        3. Retorna (filepath, filename, resposta_formatada)
        """
        from .generators.image_gen import optimize_prompt_for_image, generate_image
        
        # Step 1: Otimiza prompt
        try:
            optimized_prompt = optimize_prompt_for_image(content, agent._anthropic)
            if not optimized_prompt:
                optimized_prompt = content
        except Exception as e:
            optimized_prompt = content
            logging.error(f"Erro ao otimizar prompt: {e}")

        # Step 2: Gera imagem
        try:
            filepath, filename = await asyncio.get_event_loop().run_in_executor(
                None, lambda: generate_image(optimized_prompt, 1024, 1024)
            )
            
            if filepath and filename:
                html = f'''<div style="margin: 12px 0;">
  <img src="/static/files/{filename}" style="max-width:400px;border-radius:12px;cursor:pointer;border:1px solid #ddd;" onclick="window.open(this.src)">
  <div style="margin-top:12px;display:flex;gap:8px;align-items:center;">
    <a href="/static/files/{filename}" download style="padding:8px 16px;background:#5b5fc7;color:#fff;border-radius:6px;text-decoration:none;font-size:12px;cursor:pointer;">⬇️ Baixar</a>
    <span style="font-size:11px;color:#999;">Prompt: {optimized_prompt[:60]}...</span>
  </div>
</div>'''
                response = f"✨ Pronto! Aqui está sua imagem.\n\n{html}"
                return filepath, filename, response
            else:
                return None, None, "⏳ A geração de imagem demorou mais que o esperado. Tente novamente ou peça um briefing visual que eu monto pra você usar no Canva."
        except Exception as e:
            logging.error(f"Erro ao gerar imagem: {e}")
            return None, None, f"❌ Erro ao gerar imagem: {str(e)}"

    @app.websocket("/ws")
    async def websocket_endpoint(websocket: WebSocket):
        # Verificacao de sessao via cookie para WebSocket
        ws_cookie = websocket.cookies.get("clow_session", "")
        ws_sess = _validate_session(ws_cookie)
        if not ws_sess:
            # Fallback: API key via query param
            api_key = websocket.query_params.get("api_key", "")
            keys = _get_api_keys()
            if keys and not _verify_api_key(api_key):
                await websocket.close(code=4001, reason="Nao autenticado")
                return
            elif not keys and not ws_cookie:
                await websocket.close(code=4001, reason="Nao autenticado")
                return

        ws_is_admin = ws_sess.get("is_admin", False) if ws_sess else False
        ws_user_id = ws_sess.get("user_id", "") if ws_sess else ""

        # Rate limit para WebSocket
        client_ip = websocket.client.host if websocket.client else "unknown"
        if not _ws_rate_limiter.is_allowed(client_ip):
            await websocket.close(code=4029, reason="Rate limit excedido")
            return

        await websocket.accept()

        from .agent import Agent

        loop = asyncio.get_event_loop()

        # Cria agente com callbacks que enviam via WebSocket
        send_queue: asyncio.Queue = asyncio.Queue()

        def on_text_delta(delta: str):
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "text_delta", "content": delta}),
                loop,
            )

        def on_text_done(text: str):
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "text_done"}),
                loop,
            )

        def on_tool_call(name: str, args: dict):
            track_action("tool_call", f"{name}", "running")
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "tool_call", "name": name, "args": args}),
                loop,
            )

        def on_tool_result(name: str, status: str, output: str):
            track_action("tool_result", f"{name}: {status}", status)
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "tool_result", "name": name, "status": status, "output": output[:2000]}),
                loop,
            )

        agent = Agent(
            cwd=os.getcwd(),
            on_text_delta=on_text_delta,
            on_text_done=on_text_done,
            on_tool_call=on_tool_call,
            on_tool_result=on_tool_result,
            auto_approve=True,
        )

        # Task para enviar mensagens da fila para o WebSocket
        async def send_loop():
            try:
                while True:
                    msg = await send_queue.get()
                    await websocket.send_json(msg)
            except Exception:
                pass

        sender_task = asyncio.create_task(send_loop())

        try:
            while True:
                data = await websocket.receive_json()
                if data.get("type") == "message":
                    content = data.get("content", "")
                    file_data = data.get("file_data")
                    ws_model = data.get("model", "haiku")

                    if not content and not file_data:
                        continue

                    track_action("user_message", content[:60])

                    # Envia thinking
                    await websocket.send_json({"type": "thinking_start"})

                    # ── Admin SEMPRE usa Claude Code CLI com streaming (Opus via Max) ──
                    if ws_is_admin:
                        await websocket.send_json({"type": "thinking_end"})

                        from .claude_code_bridge import ask_claude_code_stream, log_claude_code_usage
                        track_action("user_message_claude_code", content[:60])

                        def _on_delta(delta: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "text_delta", "content": delta}),
                                loop,
                            )

                        def _on_done(full: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "text_done"}),
                                loop,
                            )

                        def _on_error(err: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "error", "content": err}),
                                loop,
                            )

                        def _on_tool_call(name: str, args: dict):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "tool_call", "name": name, "args": args}),
                                loop,
                            )

                        def _on_tool_result(name: str, status: str, output: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "tool_result", "name": name, "status": status, "output": output}),
                                loop,
                            )

                        ws_conv_id = data.get("conversation_id", "")
                        try:
                            elapsed = await loop.run_in_executor(
                                None, lambda: ask_claude_code_stream(
                                    content, _on_delta, _on_done, _on_error,
                                    "/root/clow/workspace", _on_tool_call, _on_tool_result,
                                    ws_conv_id
                                )
                            )
                            log_claude_code_usage(ws_user_id, content, elapsed)
                            track_action("claude_code_response_stream", f"{elapsed:.1f}s")
                        except Exception as e:
                            await websocket.send_json({"type": "error", "content": str(e)})
                            track_action("claude_code_error", str(e)[:60], "error")

                        await websocket.send_json({"type": "turn_complete"})
                        continue

                    # ── Detecta e processa pedido de imagem ──
                    if _should_generate_image(content) and not file_data:
                        await websocket.send_json({"type": "thinking_end"})
                        
                        # Gera imagem
                        try:
                            filepath, filename, response_html = await _process_image_request(content, agent)
                            await websocket.send_json({"type": "text_delta", "content": response_html})
                            await websocket.send_json({"type": "text_done"})
                            track_action("image_generated", filename or "failed")
                        except Exception as e:
                            await websocket.send_json({"type": "error", "content": f"Erro ao gerar imagem: {str(e)}"})
                            track_action("image_error", str(e)[:60], "error")
                        
                        # Finaliza turno
                        await websocket.send_json({"type": "turn_complete"})
                        continue

                    # Monta mensagem multimodal se tem arquivo
                    if file_data:
                        user_msg = _build_multimodal_message(content, file_data)
                    else:
                        user_msg = content

                    # Executa agente em thread separada (chat normal)
                    try:
                        result = await loop.run_in_executor(
                            None, agent.run_turn, user_msg
                        )
                        track_action("agent_response", result[:60] if result else "")
                    except Exception as e:
                        await websocket.send_json({"type": "thinking_end"})
                        await websocket.send_json({"type": "error", "content": str(e)})
                        track_action("agent_error", str(e)[:60], "error")

                    # Finaliza turno
                    await websocket.send_json({"type": "turn_complete"})

        except WebSocketDisconnect:
            pass
        except Exception:
            pass
        finally:
            sender_task.cancel()


# ── Mount de arquivos estáticos ──────────────────────────────────
if app and HAS_FASTAPI:
    try:
        # Garante que diretório existe
        os.makedirs("/root/clow/static/files", exist_ok=True)
        # Mount da pasta /static
        app.mount("/static", StaticFiles(directory="/root/clow/static"), name="static")
    except Exception as e:
        logging.error(f"Erro ao montar /static: {e}")
