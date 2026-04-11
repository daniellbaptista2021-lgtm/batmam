"""Spreadsheet Tool — cria e manipula planilhas Excel/CSV."""

from __future__ import annotations
import csv
import json
import os
from typing import Any
from .base import BaseTool


class SpreadsheetTool(BaseTool):
    name = "spreadsheet"
    description = (
        "Cria e manipula planilhas Excel/CSV. "
        "IMPORTANTE: ao criar planilha, passe TODAS as linhas de dados no campo 'rows' dentro da acao 'create'. "
        "NAO use add_row repetidamente — passe tudo de uma vez no create. "
        "Acoes: create (com headers + rows), read, add_row, add_sheet, to_csv, from_csv, info."
    )
    requires_confirmation = True

    def get_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["create", "read", "add_row", "add_sheet", "to_csv", "from_csv", "info"],
                    "description": "Ação a executar",
                },
                "file_path": {"type": "string", "description": "Caminho do arquivo"},
                "sheet_name": {"type": "string", "description": "Nome da aba (padrão: Sheet1)"},
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Cabeçalhos da planilha (para create)",
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array", "items": {"type": "string"}},
                    "description": "TODAS as linhas de dados de uma vez. Ex: [[\"val1\",\"val2\"],[\"val3\",\"val4\"]]",
                },
                "sheets": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "headers": {"type": "array", "items": {"type": "string"}},
                            "rows": {"type": "array", "items": {"type": "array", "items": {"type": "string"}}},
                        },
                    },
                    "description": "Multiplas abas de uma vez. Cada item: {name, headers, rows}",
                },
                "row": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Uma linha para adicionar",
                },
            },
            "required": ["action", "file_path"],
        }

    def execute(self, **kwargs: Any) -> str:
        action = kwargs.get("action", "")
        file_path = kwargs.get("file_path", "")
        sheet_name = kwargs.get("sheet_name", "Sheet1")

        if not file_path:
            return "Erro: file_path é obrigatório."

        if action == "create":
            sheets = kwargs.get("sheets")
            if sheets and isinstance(sheets, list):
                return self._create_multi(file_path, sheets)
            return self._create(file_path, sheet_name, kwargs.get("headers", []), kwargs.get("rows", []))
        elif action == "read":
            return self._read(file_path, sheet_name)
        elif action == "add_row":
            return self._add_row(file_path, sheet_name, kwargs.get("row", []))
        elif action == "add_sheet":
            return self._add_sheet(file_path, sheet_name)
        elif action == "to_csv":
            return self._to_csv(file_path, sheet_name)
        elif action == "from_csv":
            return self._from_csv(file_path, kwargs.get("file_path", ""))
        elif action == "info":
            return self._info(file_path)
        return f"Ação '{action}' não reconhecida."

    def _create_multi(self, path: str, sheets: list) -> str:
        """Cria planilha com multiplas abas de uma vez."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()

            for i, sheet_data in enumerate(sheets):
                name = sheet_data.get("name", f"Sheet{i+1}")
                headers = sheet_data.get("headers", [])
                rows = sheet_data.get("rows", [])

                if i == 0:
                    ws = wb.active
                    ws.title = name
                else:
                    ws = wb.create_sheet(title=name)

                if headers:
                    ws.append(headers)
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                for row in rows:
                    ws.append(row)

                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            static_dir = "/root/clow/static/files"
            os.makedirs(static_dir, exist_ok=True)
            fname = os.path.basename(path)
            if not fname.endswith(".xlsx"):
                fname = fname.rsplit(".", 1)[0] + ".xlsx" if "." in fname else fname + ".xlsx"
            fpath = os.path.join(static_dir, fname)
            wb.save(fpath)
            try:
                wb.save(path)
            except Exception:
                pass

            total_rows = sum(len(s.get("rows", [])) for s in sheets)
            domain = os.getenv("CLOW_DOMAIN", "clow.pvcorretor01.com.br")
            url = f"https://{domain}/static/files/{fname}"
            return f"Planilha criada com {len(sheets)} abas e {total_rows} linhas!\n\nBaixar: {url}"

        except ImportError:
            return "Erro: openpyxl nao instalado."
        except Exception as e:
            return f"Erro: {e}"

    def _create(self, path: str, sheet: str, headers: list, rows: list) -> str:
        if path.endswith(".csv"):
            return self._create_csv(path, headers, rows)

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet

            if headers:
                ws.append(headers)
                # Estiliza header
                from openpyxl.styles import Font, PatternFill
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for row in rows:
                ws.append(row)

            # Auto-ajusta largura das colunas
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            # Save to static/files for web download
            import time as _t
            static_dir = "/root/clow/static/files"
            os.makedirs(static_dir, exist_ok=True)
            fname = os.path.basename(path)
            if not fname.endswith(".xlsx"):
                fname = fname.rsplit(".", 1)[0] + ".xlsx" if "." in fname else fname + ".xlsx"
            fpath = os.path.join(static_dir, fname)
            wb.save(fpath)
            # Also save to requested path
            try:
                wb.save(path)
            except Exception:
                pass
            domain = os.getenv("CLOW_DOMAIN", "clow.pvcorretor01.com.br")
            url = f"https://{domain}/static/files/{fname}"
            return f"Planilha criada com {len(rows)} linhas!\n\nBaixar: {url}\n\nAbrir no Google Sheets: copie o link, va em sheets.google.com > Arquivo > Importar"

        except ImportError:
            return self._create_csv(path.replace(".xlsx", ".csv"), headers, rows)

    def _create_csv(self, path: str, headers: list, rows: list) -> str:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if headers:
                writer.writerow(headers)
            writer.writerows(rows)
        # Also save to static/files for web
        import time as _t
        static_dir = "/root/clow/static/files"
        os.makedirs(static_dir, exist_ok=True)
        fname = os.path.basename(path)
        web_path = os.path.join(static_dir, fname)
        try:
            import shutil
            shutil.copy2(path, web_path)
        except Exception:
            with open(web_path, "w", newline="", encoding="utf-8") as wf:
                writer = csv.writer(wf)
                if headers:
                    writer.writerow(headers)
                writer.writerows(rows)
        url = f"https://clow.pvcorretor01.com.br/static/files/{fname}"
        return f"Planilha criada com {len(rows)} linhas!\n\nBaixar: {url}"

    def _read(self, path: str, sheet: str) -> str:
        if path.endswith(".csv"):
            return self._read_csv(path)

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active

            rows = []
            for row in ws.iter_rows(max_row=100, values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            wb.close()

            if not rows:
                return "(planilha vazia)"

            # Formata como tabela markdown
            lines = ["| " + " | ".join(rows[0]) + " |"]
            lines.append("| " + " | ".join("---" for _ in rows[0]) + " |")
            for row in rows[1:]:
                lines.append("| " + " | ".join(c[:40] for c in row) + " |")

            total = ws.max_row or 0
            result = "\n".join(lines)
            if total > 100:
                result += f"\n\n... +{total - 100} linhas"
            return result

        except ImportError:
            return self._read_csv(path)
        except Exception as e:
            return f"Erro: {e}"

    def _read_csv(self, path: str) -> str:
        try:
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                return "(vazio)"

            lines = ["| " + " | ".join(rows[0]) + " |"]
            lines.append("| " + " | ".join("---" for _ in rows[0]) + " |")
            for row in rows[1:101]:
                lines.append("| " + " | ".join(c[:40] for c in row) + " |")

            if len(rows) > 101:
                lines.append(f"\n... +{len(rows) - 101} linhas")
            return "\n".join(lines)
        except Exception as e:
            return f"Erro CSV: {e}"

    def _add_row(self, path: str, sheet: str, row: list) -> str:
        if not row:
            return "Erro: row é obrigatório."

        if path.endswith(".csv"):
            with open(path, "a", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(row)
            return f"Linha adicionada ao CSV: {path}"

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            ws.append(row)
            wb.save(path)
            return f"Linha adicionada à aba '{ws.title}' de {path}"
        except ImportError:
            return "Erro: instale openpyxl."
        except Exception as e:
            return f"Erro: {e}"

    def _add_sheet(self, path: str, sheet: str) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            wb.create_sheet(sheet)
            wb.save(path)
            return f"Aba '{sheet}' criada em {path}"
        except ImportError:
            return "Erro: instale openpyxl."
        except Exception as e:
            return f"Erro: {e}"

    def _to_csv(self, path: str, sheet: str) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active

            csv_path = path.rsplit(".", 1)[0] + ".csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([str(c) if c is not None else "" for c in row])
            wb.close()
            return f"Exportado para CSV: {csv_path}"
        except ImportError:
            return "Erro: instale openpyxl."
        except Exception as e:
            return f"Erro: {e}"

    def _from_csv(self, csv_path: str, xlsx_path: str) -> str:
        xlsx_path = xlsx_path or csv_path.rsplit(".", 1)[0] + ".xlsx"
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active

            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)

            wb.save(xlsx_path)
            return f"CSV convertido para Excel: {xlsx_path}"
        except ImportError:
            return "Erro: instale openpyxl."
        except Exception as e:
            return f"Erro: {e}"

    def _info(self, path: str) -> str:
        if path.endswith(".csv"):
            with open(path, "r", encoding="utf-8") as f:
                lines = sum(1 for _ in f)
            size = os.path.getsize(path)
            return f"CSV: {path}\nLinhas: {lines}\nTamanho: {size} bytes"

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            info = [f"Excel: {path}", f"Abas: {', '.join(sheets)}"]
            for name in sheets:
                ws = wb[name]
                info.append(f"  {name}: {ws.max_row} linhas × {ws.max_column} colunas")
            wb.close()
            return "\n".join(info)
        except ImportError:
            return "Erro: instale openpyxl."
        except Exception as e:
            return f"Erro: {e}"
