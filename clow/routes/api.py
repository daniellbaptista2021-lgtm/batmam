"""API routes: health, status, metrics, sessions, memory, tools,
conversations CRUD, upload handling, usage, /me."""

from __future__ import annotations
import asyncio
import base64
import io
import time
import re as _re
import logging
from pathlib import Path
from typing import Any

from fastapi import FastAPI, Request, Depends, UploadFile, File, Form
from fastapi.responses import JSONResponse

from .. import __version__
from .auth import (
    _auth_dependency, _rate_limit_dependency, _generate_api_key,
    _get_user_session, _create_session, _validate_session,
)
from ..webapp import track_action, _get_health_data
from ..database import (
    authenticate_user,
    log_usage, get_user_usage_today, check_limit,
    count_user_messages_today, count_user_messages_week,
    create_conversation, list_conversations, delete_conversation,
    save_message, get_messages, update_conversation_title, PLANS,
)


def register_api_routes(app: FastAPI) -> None:
    """Register health, status, conversations, upload, usage, /me routes."""

    # ── Auth API (JSON, for Chrome Extension) ──────────────────
    @app.post("/api/v1/auth/login")
    async def api_auth_login(request: Request):
        """Login via JSON — returns session token for Chrome Extension."""
        body = await request.json()
        email = body.get("email", "").strip().lower()
        password = body.get("password", "")
        if not email or not password:
            return JSONResponse({"error": "Email e senha obrigatorios"}, status_code=400)
        user = authenticate_user(email, password)
        if not user:
            return JSONResponse({"error": "Email ou senha incorretos"}, status_code=401)
        token = _create_session(user)
        return JSONResponse({
            "token": token,
            "email": user["email"],
            "plan": user.get("plan", "lite"),
            "is_admin": bool(user.get("is_admin")),
        })

    @app.get("/api/v1/auth/verify")
    async def api_auth_verify(request: Request):
        """Verify if a token is still valid."""
        auth = request.headers.get("Authorization", "")
        token = auth.replace("Bearer ", "") if auth.startswith("Bearer ") else ""
        if not token:
            return JSONResponse({"valid": False}, status_code=401)
        sess = _validate_session(token)
        if not sess:
            return JSONResponse({"valid": False}, status_code=401)
        return JSONResponse({"valid": True, "email": sess["email"], "plan": sess.get("plan", "lite")})

    # Feature #19: Health Check (publico — sem auth)
    @app.get("/health", dependencies=[Depends(_rate_limit_dependency)])
    async def health():
        return JSONResponse(_get_health_data())

    # API endpoints protegidos
    @app.get("/api/v1/status", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_status():
        """Status completo da API com metricas."""
        return JSONResponse({
            "version": __version__,
            "status": "ok",
            **_get_health_data(),
        })

    @app.post("/api/v1/generate-key", dependencies=[Depends(_auth_dependency)])
    async def generate_key():
        """Gera nova API key para autenticacao."""
        key = _generate_api_key()
        return JSONResponse({"api_key": key, "note": "Adicione esta key em settings.json > webapp > api_keys"})

    @app.get("/api/v1/metrics", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_metrics():
        """Retorna metricas coletadas (counters, gauges, histograms)."""
        from ..logging import metrics
        return JSONResponse(metrics.snapshot())

    @app.get("/api/v1/sessions", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)], include_in_schema=False)
    async def api_sessions(request: Request):
        """Lista sessoes salvas (admin only)."""
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Admin only"}, status_code=403)
        from ..session import list_sessions
        return JSONResponse({"sessions": list_sessions()})

    @app.get("/api/v1/memory", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)], include_in_schema=False)
    async def api_memory(request: Request):
        """Lista memorias persistidas (admin only)."""
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Admin only"}, status_code=403)
        from ..memory import list_memories
        return JSONResponse({"memories": list_memories()})

    @app.get("/api/v1/tools", dependencies=[Depends(_auth_dependency)], include_in_schema=False)
    async def api_tools(request: Request):
        """Lista ferramentas disponiveis (nomes apenas)."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        from ..tools.base import create_default_registry
        registry = create_default_registry()
        # Retorna apenas nomes, sem schemas/descricoes detalhadas
        tools = [t.name for t in registry.all_tools()]
        return JSONResponse({"tools": tools, "count": len(tools)})

    # ── API: Conversations ──────────────────────────────────────────

    @app.get("/api/v1/conversations")
    async def api_conversations(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        convs = list_conversations(sess["user_id"])
        return JSONResponse({"conversations": convs})

    @app.post("/api/v1/conversations")
    async def api_create_conversation(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        cid = create_conversation(sess["user_id"])
        return JSONResponse({"id": cid})

    @app.get("/api/v1/conversations/{conv_id}/messages")
    async def api_get_messages(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        # Valida ownership: conversa pertence ao usuario
        user_convs = list_conversations(sess["user_id"])
        if not any(c["id"] == conv_id for c in user_convs):
            return JSONResponse({"error": "Conversa nao encontrada"}, status_code=404)
        msgs = get_messages(conv_id)
        return JSONResponse({"messages": msgs})

    @app.delete("/api/v1/conversations/{conv_id}")
    async def api_delete_conversation(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        delete_conversation(sess["user_id"], conv_id)
        return JSONResponse({"ok": True})

    @app.post("/api/v1/conversations/{conv_id}/title")
    async def api_update_conv_title(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        # Valida ownership
        user_convs = list_conversations(sess["user_id"])
        if not any(c["id"] == conv_id for c in user_convs):
            return JSONResponse({"error": "Conversa nao encontrada"}, status_code=404)
        body = await request.json()
        title = body.get("title", "")[:50]
        if title:
            update_conversation_title(conv_id, title)
        return JSONResponse({"ok": True})

    # ── API: Upload de Arquivos ─────────────────────────────────────

    _ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    _ALLOWED_DOC_EXT = {".pdf", ".docx", ".doc", ".txt", ".md"}
    _ALLOWED_SHEET_EXT = {".xlsx", ".xls", ".csv"}
    _ALLOWED_CODE_EXT = {".py", ".js", ".html", ".css", ".json", ".ts", ".jsx", ".tsx"}
    _ALLOWED_AUDIO_EXT = {".webm", ".mp3", ".ogg", ".wav", ".m4a"}
    _BLOCKED_EXT = {".exe", ".bat", ".sh", ".cmd", ".com", ".msi", ".scr", ".ps1"}
    _MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
    _UPLOAD_LIMITS = {"lite": 50, "starter": 100, "pro": 200, "business": 500, "unlimited": 0}
    _upload_counts: dict[str, dict] = {}  # user_id -> {"date": "YYYY-MM-DD", "count": N}

    def _check_upload_limit(user_id: str, plan: str) -> bool:
        import datetime
        today = datetime.date.today().isoformat()
        rec = _upload_counts.get(user_id, {"date": "", "count": 0})
        if rec["date"] != today:
            rec = {"date": today, "count": 0}
        limit = _UPLOAD_LIMITS.get(plan, 5)
        if limit == 0:
            return True
        return rec["count"] < limit

    def _inc_upload_count(user_id: str):
        import datetime
        today = datetime.date.today().isoformat()
        rec = _upload_counts.get(user_id, {"date": "", "count": 0})
        if rec["date"] != today:
            rec = {"date": today, "count": 0}
        rec["count"] += 1
        _upload_counts[user_id] = rec

    def _sanitize_filename(name: str) -> str:
        name = _re.sub(r'[^\w\-_. ]', '', name)
        return name.strip()[:100] or "arquivo"

    def _format_size(size: int) -> str:
        if size > 1024 * 1024:
            return f"{size / (1024*1024):.1f} MB"
        if size > 1024:
            return f"{size / 1024:.1f} KB"
        return f"{size} bytes"

    def _extract_text_docx(data: bytes) -> str:
        try:
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[Erro ao ler DOCX: {e}]"

    def _extract_text_xlsx(data: bytes) -> tuple[str, int]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            total_rows = 0
            for sheet in wb.worksheets:
                lines.append(f"\n### Aba: {sheet.title}\n")
                rows = list(sheet.iter_rows(values_only=True))
                total_rows += len(rows)
                if rows:
                    # Header
                    header = [str(c) if c is not None else "" for c in rows[0]]
                    lines.append("| " + " | ".join(header) + " |")
                    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                    for row in rows[1:51]:
                        cells = [str(c) if c is not None else "" for c in row]
                        lines.append("| " + " | ".join(cells) + " |")
                    if len(rows) > 51:
                        lines.append(f"\n... e mais {len(rows) - 51} linhas")
            wb.close()
            return "\n".join(lines), total_rows
        except Exception as e:
            return f"[Erro ao ler XLSX: {e}]", 0

    def _extract_text_csv(data: bytes) -> tuple[str, int]:
        import csv
        try:
            text = data.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            lines = []
            if rows:
                header = rows[0]
                lines.append("| " + " | ".join(header) + " |")
                lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                for row in rows[1:51]:
                    lines.append("| " + " | ".join(row) + " |")
                if len(rows) > 51:
                    lines.append(f"\n... e mais {len(rows) - 51} linhas")
            return "\n".join(lines), len(rows)
        except Exception as e:
            return f"[Erro ao ler CSV: {e}]", 0

    def _resize_image(data: bytes, max_px: int = 2000) -> bytes:
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(data))
            if max(img.size) > max_px:
                img.thumbnail((max_px, max_px), Image.LANCZOS)
                buf = io.BytesIO()
                fmt = img.format or "JPEG"
                if fmt.upper() == "WEBP":
                    img.save(buf, format="WEBP", quality=85)
                else:
                    img.save(buf, format="JPEG", quality=85)
                return buf.getvalue()
        except Exception:
            pass
        return data

    async def _transcribe_audio(file_path: str) -> str:
        """Transcricao de audio — agora feita no frontend via Web Speech API.
        Backend nao usa mais OpenAI Whisper. Retorna vazio para fallback."""
        return ""

    @app.post("/api/v1/upload")
    async def api_upload(request: Request, file: UploadFile = File(...), message: str = Form("")):
        """Upload de arquivo com processamento automatico."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        user_id = sess["user_id"]
        user_plan = sess.get("plan", "lite")

        if not _check_upload_limit(user_id, user_plan):
            limit = _UPLOAD_LIMITS.get(user_plan, 5)
            return JSONResponse({"error": f"Limite de {limit} uploads/dia atingido. Faca upgrade do plano."}, status_code=429)

        # Le arquivo
        data = await file.read()
        if len(data) > _MAX_FILE_SIZE:
            return JSONResponse({"error": "Arquivo muito grande. Maximo: 20MB"}, status_code=413)

        original_name = file.filename or "arquivo"
        safe_name = _sanitize_filename(original_name)
        ext = Path(original_name).suffix.lower()

        if ext in _BLOCKED_EXT:
            return JSONResponse({"error": f"Tipo de arquivo nao permitido: {ext}"}, status_code=400)

        # Salva arquivo
        upload_dir = Path(__file__).parent.parent.parent / "static" / "uploads" / user_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        saved_name = f"{ts}_{safe_name}"
        saved_path = upload_dir / saved_name
        saved_path.write_bytes(data)

        _inc_upload_count(user_id)
        file_size = len(data)
        file_url = f"/static/uploads/{user_id}/{saved_name}"

        result: dict[str, Any] = {
            "ok": True,
            "file_name": safe_name,
            "file_url": file_url,
            "file_size": _format_size(file_size),
            "file_ext": ext,
            "type": "unknown",
        }

        # Processa por tipo
        if ext in _ALLOWED_IMAGE_EXT:
            resized = _resize_image(data)
            # Detecta media_type real pelos magic bytes (apos resize)
            if resized[:3] == b'\xff\xd8\xff':
                media_type = 'image/jpeg'
            elif resized[:4] == b'\x89PNG':
                media_type = 'image/png'
            elif resized[:4] == b'GIF8':
                media_type = 'image/gif'
            elif resized[:4] == b'RIFF' and len(resized) > 11 and resized[8:12] == b'WEBP':
                media_type = 'image/webp'
            else:
                media_type = 'image/jpeg'
            b64 = base64.b64encode(resized).decode("ascii")
            result["type"] = "image"
            result["media_type"] = media_type
            result["base64"] = b64

        elif ext == ".pdf":
            b64 = base64.b64encode(data).decode("ascii")
            # Conta paginas
            pages = 0
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(data))
                pages = len(reader.pages)
            except Exception:
                pass
            if pages > 50:
                result["warning"] = f"PDF com {pages} paginas. Apenas as primeiras 50 serao analisadas."
            result["type"] = "pdf"
            result["media_type"] = "application/pdf"
            result["base64"] = b64
            result["pages"] = pages

        elif ext in _ALLOWED_SHEET_EXT:
            if ext == ".csv":
                text, rows = _extract_text_csv(data)
            else:
                text, rows = _extract_text_xlsx(data)
            result["type"] = "spreadsheet"
            result["extracted_text"] = text
            result["rows"] = rows

        elif ext in _ALLOWED_DOC_EXT:
            if ext == ".docx":
                text = _extract_text_docx(data)
            else:
                text = data.decode("utf-8", errors="replace")
            words = len(text.split())
            if words > 5000:
                text = " ".join(text.split()[:5000]) + f"\n\n... (truncado, {words} palavras total)"
            result["type"] = "document"
            result["extracted_text"] = text
            result["words"] = words

        elif ext in _ALLOWED_CODE_EXT:
            text = data.decode("utf-8", errors="replace")
            lang_map = {".py": "python", ".js": "javascript", ".ts": "typescript",
                        ".html": "html", ".css": "css", ".json": "json", ".jsx": "jsx", ".tsx": "tsx"}
            result["type"] = "code"
            result["extracted_text"] = text
            result["language"] = lang_map.get(ext, "text")

        elif ext in _ALLOWED_AUDIO_EXT:
            # Transcreve audio
            transcription = await _transcribe_audio(str(saved_path))
            result["type"] = "audio"
            result["transcription"] = transcription
            result["has_transcription"] = bool(transcription and not transcription.startswith("[Erro"))

        else:
            result["type"] = "file"

        track_action("file_upload", f"{ext}: {safe_name}", "ok")
        return JSONResponse(result)

    # ── API: Usage ───────────────────────────────────────────────────

    @app.get("/api/v1/usage")
    async def api_usage(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        from .. import config
        user_id = sess["user_id"]
        usage = get_user_usage_today(user_id)
        plan = PLANS.get(sess.get("plan", "lite"), PLANS["lite"])
        messages_today = count_user_messages_today(user_id)
        messages_week = count_user_messages_week(user_id)
        return JSONResponse({
            "usage": usage,
            "plan": sess.get("plan", "lite"),
            "plan_label": plan["label"],
            "daily_token_limit": plan["daily_tokens"],
            "messages_today": messages_today,
            "messages_week": messages_week,
            "daily_message_limit": config.CLOW_DAILY_LIMIT,
            "weekly_message_limit": config.CLOW_WEEKLY_LIMIT,
        })

    @app.get("/api/v1/me")
    async def api_me(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        plan = sess.get("plan", "lite")
        is_admin = sess.get("is_admin", False)
        if is_admin:
            models = ["claude-code"]
        else:
            models = ["haiku"]
            if plan in ("pro", "unlimited"):
                models.append("sonnet")
        # Flags de primeiro login pra tour guiado
        first_login = 0
        try:
            from ..database import get_db
            with get_db() as db:
                row = db.execute(
                    "SELECT first_login FROM users WHERE id=?",
                    (sess["user_id"],),
                ).fetchone()
                if row:
                    first_login = int(row["first_login"] or 0)
        except Exception:
            pass
        return JSONResponse({
            "email": sess["email"],
            "user_id": sess["user_id"],
            "is_admin": is_admin,
            "plan": plan,
            "available_models": models,
            "first_login": first_login,
        })

    @app.post("/api/v1/user/mark-tour-done", tags=["user"])
    async def api_mark_tour_done(request: Request):
        """Marca tour guiado como concluido (limpa first_login)."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        try:
            from ..database import get_db
            with get_db() as db:
                db.execute("UPDATE users SET first_login=0 WHERE id=?", (sess["user_id"],))
            return JSONResponse({"success": True})
        except Exception as e:
            return JSONResponse({"error": str(e)[:200]}, status_code=500)

    @app.get("/api/v1/user/setup-state", tags=["user"])
    async def api_setup_state(request: Request):
        """Estado rapido pro chat.html: onboarding, WhatsApp, tour.
        Independente do /api/v1/onboarding/status que depende de funcoes
        que nao existem em database.py.
        """
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        uid = sess["user_id"]
        try:
            from ..database import get_db
            with get_db() as db:
                u = db.execute(
                    "SELECT onboarding_completed, first_login, plan FROM users WHERE id=?",
                    (uid,),
                ).fetchone()
                wa = db.execute(
                    "SELECT status FROM whatsapp_credentials WHERE user_id=? "
                    "ORDER BY id DESC LIMIT 1",
                    (uid,),
                ).fetchone()
            onboarding_done = bool(u and u["onboarding_completed"])
            first_login = bool(u and u["first_login"])
            plan = u["plan"] if u else "free"
            is_paid = plan in ("lite", "starter", "pro", "business")
            wa_connected = bool(wa and wa["status"] == "connected")
            return JSONResponse({
                "onboarding_completed": onboarding_done,
                "first_login": first_login,
                "whatsapp_connected": wa_connected,
                "plan": plan,
                "is_paid_plan": is_paid,
            })
        except Exception as e:
            return JSONResponse({"error": str(e)[:200]}, status_code=500)
