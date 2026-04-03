"""Clow Web App — FastAPI + WebSocket com UI estilo Claude Code.

Features:
  #18 — Web App (FastAPI + WebSocket com UI completa)
  #19 — Health Check + Monitoring Endpoint
  #24 — Dashboard de Metricas
  #26 — Autenticacao via API Key / Bearer Token
  #27 — Rate Limiting por IP (configurable)
  #28 — CORS configuravel
  #29 — HTTPS/TLS support (via uvicorn ssl)
"""

from __future__ import annotations
import json
import asyncio
import os
import time
import hashlib
import secrets
import logging
from collections import defaultdict
from typing import Any
from pathlib import Path

try:
    from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request, HTTPException, Depends, UploadFile, File, Form
    from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, RedirectResponse
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.staticfiles import StaticFiles
    HAS_FASTAPI = True
except ImportError:
    HAS_FASTAPI = False

import base64
import re as _re
import io

from . import __version__
from . import config

app = FastAPI(
    title="Clow",
    version=__version__,
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json",
) if HAS_FASTAPI else None


# ── Autenticacao ─────────────────────────────────────────────────

def _get_api_keys() -> list[str]:
    """Carrega API keys do settings ou env."""
    settings = config.load_settings()
    keys = settings.get("webapp", {}).get("api_keys", [])
    env_key = os.getenv("CLOW_API_KEY", "")
    if env_key and env_key not in keys:
        keys.append(env_key)
    return keys


def _generate_api_key() -> str:
    """Gera uma nova API key segura."""
    return f"clow_{secrets.token_urlsafe(32)}"


def _verify_api_key(key: str) -> bool:
    """Verifica se uma API key e valida."""
    valid_keys = _get_api_keys()
    if not valid_keys:
        return True  # Sem keys configuradas = sem autenticacao (dev mode)
    return key in valid_keys


async def _auth_dependency(request: Request) -> None:
    """FastAPI dependency para verificar autenticacao."""
    keys = _get_api_keys()
    if not keys:
        return  # Dev mode — sem autenticacao

    # Tenta Authorization header
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:]
        if _verify_api_key(token):
            return

    # Tenta query param
    api_key = request.query_params.get("api_key", "")
    if api_key and _verify_api_key(api_key):
        return

    raise HTTPException(status_code=401, detail="API key invalida ou ausente")


# ── Login / Sessao Web (SQLite backed) ──────────────────────────

from .database import (
    authenticate_user, create_user, get_user_by_email, get_user_by_id,
    list_users, update_user, log_usage, get_user_usage_today, check_limit,
    get_admin_stats, create_conversation, list_conversations, delete_conversation,
    save_message, get_messages, update_conversation_title, PLANS, ADMIN_EMAIL,
)

_web_sessions: dict[str, dict] = {}  # token -> {"email", "user_id", "is_admin", "created"}
_SESSION_TTL = 86400 * 7  # 7 dias


def _create_session(user: dict) -> str:
    token = secrets.token_urlsafe(48)
    _web_sessions[token] = {
        "email": user["email"],
        "user_id": user["id"],
        "is_admin": bool(user.get("is_admin")),
        "plan": user.get("plan", "free"),
        "created": time.time(),
    }
    return token


def _validate_session(token: str) -> dict | None:
    sess = _web_sessions.get(token)
    if not sess:
        return None
    if time.time() - sess["created"] > _SESSION_TTL:
        del _web_sessions[token]
        return None
    return sess


def _get_session_from_request(request: Request) -> str | None:
    """Returns email string for backwards compat."""
    token = request.cookies.get("clow_session", "")
    sess = _validate_session(token)
    return sess["email"] if sess else None


def _get_user_session(request: Request) -> dict | None:
    """Returns full session dict."""
    token = request.cookies.get("clow_session", "")
    return _validate_session(token)


# ── Rate Limiting ────────────────────────────────────────────────

class RateLimiter:
    """Rate limiter por IP com sliding window."""

    def __init__(self, max_requests: int = 60, window_seconds: int = 60):
        self.max_requests = max_requests
        self.window = window_seconds
        self._requests: dict[str, list[float]] = defaultdict(list)

    def is_allowed(self, ip: str) -> bool:
        now = time.time()
        window_start = now - self.window

        # Limpa requests antigos
        self._requests[ip] = [t for t in self._requests[ip] if t > window_start]

        if len(self._requests[ip]) >= self.max_requests:
            return False

        self._requests[ip].append(now)
        return True

    def remaining(self, ip: str) -> int:
        now = time.time()
        window_start = now - self.window
        recent = [t for t in self._requests[ip] if t > window_start]
        return max(0, self.max_requests - len(recent))


_rate_limiter = RateLimiter(max_requests=60, window_seconds=60)
_ws_rate_limiter = RateLimiter(max_requests=10, window_seconds=60)


async def _rate_limit_dependency(request: Request) -> None:
    """FastAPI dependency para rate limiting."""
    client_ip = request.client.host if request.client else "unknown"
    if not _rate_limiter.is_allowed(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Rate limit excedido. Tente novamente em alguns segundos.",
            headers={"Retry-After": "60"},
        )


# ── Setup CORS e Middleware ──────────────────────────────────────

def _setup_middleware():
    """Configura CORS e middlewares de seguranca."""
    if not HAS_FASTAPI or app is None:
        return

    settings = config.load_settings()
    webapp_cfg = settings.get("webapp", {})

    # CORS
    allowed_origins = webapp_cfg.get("cors_origins", ["http://localhost:*", "http://127.0.0.1:*"])
    app.add_middleware(
        CORSMiddleware,
        allow_origins=allowed_origins,
        allow_credentials=True,
        allow_methods=["GET", "POST", "OPTIONS"],
        allow_headers=["*"],
    )


if HAS_FASTAPI:
    _setup_middleware()


def get_app():
    if not HAS_FASTAPI:
        raise RuntimeError("FastAPI nao instalado. Instale com: pip install 'clow[web]'")
    return app


def start_with_tls(host: str = "0.0.0.0", port: int = 8080, certfile: str = "", keyfile: str = ""):
    """Inicia o servidor com TLS/HTTPS se certificados forem fornecidos."""
    import uvicorn

    kwargs: dict[str, Any] = {
        "app": "clow.webapp:app",
        "host": host,
        "port": port,
        "log_level": "info",
    }
    if certfile and keyfile:
        kwargs["ssl_certfile"] = certfile
        kwargs["ssl_keyfile"] = keyfile

    uvicorn.run(**kwargs)


# ── Ações executadas (Feature #24 — tracking) ──────────────────
_recent_actions: list[dict] = []
MAX_RECENT_ACTIONS = 50


def track_action(action: str, details: str = "", status: str = "ok") -> None:
    """Registra ação recente para o dashboard."""
    _recent_actions.append({
        "action": action,
        "details": details[:100],
        "status": status,
        "timestamp": time.time(),
    })
    if len(_recent_actions) > MAX_RECENT_ACTIONS:
        _recent_actions.pop(0)


# ── Feature #19: Health Check ──────────────────────────────────

def _get_health_data() -> dict:
    """Coleta status de todos os componentes."""
    from .memory import list_memories
    from .cron import get_cron_manager
    from .triggers import get_trigger_server
    from .tasks import get_task_manager

    memories = list_memories()
    cron = get_cron_manager()
    trigger = get_trigger_server()
    tasks = get_task_manager()

    mem_by_type: dict[str, int] = {}
    for m in memories:
        t = m.get("type", "general")
        mem_by_type[t] = mem_by_type.get(t, 0) + 1

    all_tasks = tasks.list_all()
    tasks_by_status: dict[str, int] = {}
    for t in all_tasks:
        s = t.status.value
        tasks_by_status[s] = tasks_by_status.get(s, 0) + 1

    cron_jobs = cron.list_all()
    active_crons = [j for j in cron_jobs if j.active]

    return {
        "status": "healthy",
        "version": __version__,
        "uptime_info": time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()),
        "components": {
            "memory": {
                "status": "ok",
                "total": len(memories),
                "by_type": mem_by_type,
            },
            "cron": {
                "status": "ok",
                "total_jobs": len(cron_jobs),
                "active_jobs": len(active_crons),
                "jobs": [
                    {
                        "id": j.id,
                        "prompt": j.prompt[:50],
                        "interval": cron.format_interval(j.interval_seconds),
                        "active": j.active,
                        "run_count": j.run_count,
                        "last_run": j.last_run,
                        "next_run": j.last_run + j.interval_seconds if j.last_run else j.created_at + j.interval_seconds,
                    }
                    for j in cron_jobs
                ],
            },
            "triggers": {
                "status": "ok" if trigger.running else "stopped",
                "running": trigger.running,
                "port": trigger.port,
                "results_count": len(trigger.list_results()),
            },
            "tasks": {
                "status": "ok",
                "total": len(all_tasks),
                "by_status": tasks_by_status,
            },
        },
        "recent_actions": _recent_actions[-10:],
    }


def _build_multimodal_message(text: str, file_data: dict) -> Any:
    """Monta mensagem multimodal (content blocks) para API Anthropic."""
    ftype = file_data.get("type", "")
    content_blocks = []

    if ftype == "image":
        content_blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": file_data.get("media_type", "image/jpeg"),
                "data": file_data.get("base64", ""),
            },
        })
        if text:
            content_blocks.append({"type": "text", "text": text})
        else:
            content_blocks.append({"type": "text", "text": "Analise esta imagem e descreva o que você vê."})

    elif ftype == "pdf":
        content_blocks.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": file_data.get("base64", ""),
            },
        })
        if text:
            content_blocks.append({"type": "text", "text": text})
        else:
            content_blocks.append({"type": "text", "text": "Analise este PDF e resuma o conteúdo."})

    elif ftype == "audio":
        transcription = file_data.get("transcription", "")
        if transcription and not transcription.startswith("[Erro"):
            prompt = f"[Audio transcrito do usuario]\nTranscricao: {transcription}"
            if text:
                prompt += f"\n\nMensagem adicional: {text}"
            return prompt
        else:
            return text or "[O usuario enviou um audio mas a transcricao nao esta disponivel]"

    elif ftype in ("spreadsheet", "document", "code"):
        extracted = file_data.get("extracted_text", "")
        fname = file_data.get("file_name", "arquivo")
        if ftype == "spreadsheet":
            prefix = f"[Planilha: {fname}]\n\nDados:\n{extracted}"
        elif ftype == "code":
            lang = file_data.get("language", "text")
            prefix = f"[Codigo: {fname}]\n\n```{lang}\n{extracted}\n```"
        else:
            prefix = f"[Documento: {fname}]\n\nConteudo:\n{extracted}"
        if text:
            prefix += f"\n\nPedido do usuario: {text}"
        return prefix

    else:
        return text or f"[O usuario enviou um arquivo: {file_data.get('file_name', 'arquivo')}]"

    return content_blocks


# ── HTML/CSS/JS — SPA Completo ────────────────────────────────



WEBAPP_HTML = r'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,minimum-scale=1.0,maximum-scale=3.0,user-scalable=yes,viewport-fit=cover">
<meta name="theme-color" content="#050510">
<meta name="description" content="Clow - AI Code Agent | Inteligência Infinita • Possibilidades Premium. Crie landing pages, planilhas, documentos, apps e muito mais com inteligência artificial.">
<meta name="keywords" content="AI, inteligência artificial, code agent, landing page, automação, marketing digital">
<meta name="author" content="Clow AI">
<meta name="robots" content="index, follow">
<meta property="og:title" content="Clow — Inteligência Infinita">
<meta property="og:description" content="AI Code Agent premium. Crie landing pages, planilhas, apps e muito mais.">
<meta property="og:image" content="/static/brand/logo.png">
<meta property="og:url" content="https://clow.pvcorretor01.com.br">
<meta property="og:type" content="website">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="Clow — Inteligência Infinita">
<meta name="twitter:description" content="AI Code Agent premium.">
<meta name="twitter:image" content="/static/brand/logo.png">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<link rel="manifest" href="/static/manifest.json">
<link rel="icon" type="image/png" href="/static/brand/favicon.png">
<link rel="apple-touch-icon" href="/static/brand/icon.png">
<title>Clow — Inteligência Infinita</title>
<script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root{
  --bg-0:#050510;--bg-1:#0A0A1A;--bg-2:#0F0F24;--bg-3:#14142E;--bg-h:#1A1A3A;
  --bd:rgba(100,100,180,.12);--bdh:rgba(100,100,180,.25);--bdf:rgba(124,92,252,.5);
  --p:#9B59FC;--pg:rgba(155,89,252,.15);--pgs:rgba(155,89,252,.3);--ph:#8B49EC;
  --bl:#4A9EFF;--blg:rgba(74,158,255,.15);
  --gp:linear-gradient(135deg,#9B59FC 0%,#4A9EFF 100%);--gph:linear-gradient(135deg,#8B49EC 0%,#3A8EEF 100%);
  --gg:linear-gradient(135deg,rgba(155,89,252,.15) 0%,rgba(74,158,255,.15) 100%);
  --g:#4ADE80;--am:#FBBF24;--r:#F87171;--cy:#22D3EE;
  --t1:#E8E8F0;--t2:#9898B8;--tm:#585878;
  --sans:'DM Sans',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
  --mono:'JetBrains Mono','Fira Code','Cascadia Code','Consolas',monospace;
  --st:env(safe-area-inset-top,0px);--sb:env(safe-area-inset-bottom,0px);--sw:260px;
}
*{margin:0;padding:0;box-sizing:border-box;-webkit-tap-highlight-color:transparent}
html{height:100%;overflow:hidden}
body{background:var(--bg-0);color:var(--t1);font-family:var(--sans);font-size:14px;font-weight:400;line-height:1.6;letter-spacing:-0.01em;height:100dvh;display:flex;overflow:hidden;-webkit-font-smoothing:antialiased;-moz-osx-font-smoothing:grayscale;text-rendering:optimizeLegibility;position:relative}
body::before{content:'';position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none;
  background:
    repeating-linear-gradient(-45deg,transparent,transparent 8px,rgba(80,70,140,.04) 8px,rgba(80,70,140,.04) 9px),
    radial-gradient(ellipse at 50% 40%,rgba(155,89,252,.06) 0%,rgba(74,158,255,.03) 35%,transparent 65%)}
body::after{content:'';position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none;
  background:radial-gradient(circle at 30% 20%,rgba(155,89,252,.04) 0%,transparent 50%),
  radial-gradient(circle at 70% 80%,rgba(74,158,255,.03) 0%,transparent 50%)}
.particle{position:fixed;border-radius:50%;pointer-events:none;z-index:1;filter:blur(1px)}
@keyframes floatParticle{0%,100%{transform:translateY(0) translateX(0);opacity:.3}25%{transform:translateY(-30px) translateX(10px);opacity:.6}50%{transform:translateY(-15px) translateX(-10px);opacity:.4}75%{transform:translateY(-40px) translateX(5px);opacity:.5}}
::-webkit-scrollbar{width:5px}::-webkit-scrollbar-track{background:transparent}::-webkit-scrollbar-thumb{background:linear-gradient(180deg,var(--p),var(--bl));border-radius:3px}::-webkit-scrollbar-thumb:hover{box-shadow:0 0 6px var(--pg)}

/* SIDEBAR */
.sb{width:var(--sw);height:100%;background:var(--bg-1);border-right:1px solid var(--bd);display:flex;flex-direction:column;flex-shrink:0;z-index:30;transition:transform .25s ease;box-shadow:1px 0 20px rgba(155,89,252,.03)}
.sb-head{padding:20px 16px 16px;padding-top:calc(16px + var(--st));display:flex;align-items:center;gap:10px}
.sb-logo-img{height:36px;object-fit:contain;filter:drop-shadow(0 0 12px rgba(155,89,252,.2))}
.sb-body{flex:1;overflow-y:auto;padding:4px 0;display:flex;flex-direction:column}
.sb-sep{height:1px;background:var(--bd);margin:8px 12px}
.sb-sec{padding:0 12px;margin-bottom:2px}
.sb-lbl{font-family:var(--sans);font-size:10px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:var(--tm);padding:12px 4px 6px;cursor:pointer;display:flex;align-items:center;justify-content:space-between;user-select:none}
.sb-lbl:hover{color:var(--t2)}.sb-lbl .ar{transition:transform .2s;font-size:8px;opacity:.5}.sb-lbl.open .ar{transform:rotate(90deg)}
.sb-ct{display:none;padding-bottom:4px}.sb-lbl.open+.sb-ct{display:block}
.sb-btn{display:flex;align-items:center;gap:8px;padding:9px 14px;border-radius:6px;cursor:pointer;font-size:13px;font-weight:500;color:var(--t2);transition:all .15s;border:none;background:none;width:100%;text-align:left;font-family:var(--sans);letter-spacing:.2px}
.sb-btn:hover,.sb-btn:active{background:rgba(155,89,252,.08);color:var(--t1);padding-left:16px}
.sb-btn:hover::before{content:'';width:2px;height:14px;background:var(--gp);border-radius:1px;flex-shrink:0}
.sb-btn.act{background:rgba(155,89,252,.1);color:var(--p);border-left:2px solid;border-image:var(--gp) 1;padding-left:10px}
.sb-btn .ic{font-size:15px;width:20px;text-align:center;flex-shrink:0;color:var(--tm)}.sb-btn:hover .ic,.sb-btn.act .ic{background:var(--gp);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.sb-new{margin:8px 12px 12px;padding:10px;border-radius:8px;border:none;background:transparent;color:#fff;font-family:var(--sans);font-size:13px;font-weight:600;cursor:pointer;text-align:center;transition:all .25s ease;letter-spacing:.3px;position:relative;overflow:hidden}
.sb-new::before{content:'';position:absolute;inset:0;border-radius:8px;padding:1px;background:var(--gp);mask:linear-gradient(#fff 0 0) content-box,linear-gradient(#fff 0 0);-webkit-mask:linear-gradient(#fff 0 0) content-box,linear-gradient(#fff 0 0);mask-composite:exclude;-webkit-mask-composite:xor}
.sb-new:hover{background:var(--gp);box-shadow:0 0 20px var(--pg)}
.sb-convs{flex:1;overflow-y:auto;padding:0 8px}
.sb-conv-search{padding:0 8px 8px;position:relative}
.sb-conv-search input{width:100%;background:var(--bg-2);border:1px solid var(--bd);border-radius:6px;padding:6px 10px 6px 30px;color:var(--t1);font-size:12px;font-family:var(--sans);outline:none;transition:border-color .15s}
.sb-conv-search input:focus{border-color:var(--p)}
.sb-conv-search input::placeholder{color:var(--tm)}
.sb-conv-search .search-icon{position:absolute;left:18px;top:50%;transform:translateY(-50%);color:var(--tm);font-size:12px;pointer-events:none;padding-bottom:8px}
.sb-grp-label{padding:6px 8px 4px;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:var(--tm);user-select:none}
.sb-conv-item{display:flex;align-items:center;padding:8px 10px;border-radius:6px;cursor:pointer;transition:background .15s ease;position:relative;gap:8px;margin:1px 0}
.sb-conv-item:hover{background:rgba(155,89,252,.06)}
.sb-conv-item.act{background:rgba(155,89,252,.1);border-left:2px solid;border-image:var(--gp) 1;padding-left:8px}
.sb-conv-item .conv-icon{font-size:13px;flex-shrink:0;color:var(--tm);width:16px;text-align:center}
.sb-conv-item.act .conv-icon{color:var(--p)}
.sb-conv-item .conv-title{flex:1;font-size:13px;color:var(--t2);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-family:var(--sans);line-height:1.3}
.sb-conv-item.act .conv-title{color:var(--t1);font-weight:500}
.sb-conv-item .conv-pin-static{font-size:10px;color:var(--tm);flex-shrink:0;margin-right:2px}
.sb-conv-item .conv-actions{display:flex;align-items:center;gap:2px;opacity:0;transition:opacity .15s ease;flex-shrink:0}
.sb-conv-item:hover .conv-actions{opacity:1}
.sb-conv-item .conv-actions .ca-btn{background:none;border:none;cursor:pointer;padding:3px 4px;border-radius:4px;font-size:12px;color:var(--tm);transition:all .15s;display:flex;align-items:center;justify-content:center;line-height:1}
.sb-conv-item .conv-actions .ca-btn:hover{background:var(--bg-3);color:var(--t1)}
/* Context menu */
.conv-ctx-menu{position:fixed;background:var(--bg-3);border:1px solid rgba(100,100,180,.15);border-radius:8px;padding:4px;min-width:180px;z-index:999;box-shadow:0 10px 30px rgba(0,0,0,.4);transform-origin:top left;animation:ctxIn .15s ease}
@keyframes ctxIn{from{opacity:0;transform:scale(.95)}to{opacity:1;transform:scale(1)}}
.conv-ctx-menu .ctx-item{display:flex;align-items:center;gap:8px;padding:8px 12px;border-radius:6px;cursor:pointer;font-size:13px;color:var(--t2);font-family:var(--sans);transition:background .12s;border:none;background:none;width:100%;text-align:left}
.conv-ctx-menu .ctx-item:hover{background:rgba(155,89,252,.08);color:var(--t1)}
.conv-ctx-menu .ctx-item.danger{color:var(--r)}
.conv-ctx-menu .ctx-item.danger:hover{background:rgba(248,113,113,.1)}
.conv-ctx-menu .ctx-item .ctx-icon{font-size:14px;width:18px;text-align:center;flex-shrink:0}
.conv-ctx-sep{height:1px;background:var(--bd);margin:4px 0}
/* Delete confirm */
.conv-del-confirm{display:flex;align-items:center;gap:6px;padding:6px 12px}
.conv-del-confirm span{font-size:12px;color:var(--t2);font-family:var(--sans)}
.conv-del-confirm button{padding:4px 12px;border-radius:4px;border:none;font-size:11px;font-weight:600;cursor:pointer;font-family:var(--sans);transition:all .15s}
.conv-del-confirm .del-yes{background:var(--r);color:#fff}
.conv-del-confirm .del-yes:hover{background:#ef4444}
.conv-del-confirm .del-no{background:var(--bg-3);color:var(--t2)}
.conv-del-confirm .del-no:hover{background:var(--bg-h);color:var(--t1)}
/* Rename inline */
.sb-conv-item .conv-rename-input{flex:1;background:var(--bg-2);border:1px solid var(--p);border-radius:4px;padding:2px 6px;color:var(--t1);font-size:12px;font-family:var(--sans);outline:none}
/* Animations */
@keyframes convSlideIn{from{opacity:0;transform:translateX(-12px)}to{opacity:1;transform:translateX(0)}}
@keyframes convFadeOut{to{opacity:0;transform:translateX(-12px)}}
.sb-conv-item.entering{animation:convSlideIn .2s ease forwards}
.sb-conv-item.leaving{animation:convFadeOut .2s ease forwards}
@keyframes pinBounce{0%,100%{transform:scale(1)}50%{transform:scale(1.3)}}
.pin-bounce{animation:pinBounce .3s ease}
/* Ver anteriores btn */
.sb-conv-more{display:flex;align-items:center;justify-content:center;padding:8px;font-size:11px;color:var(--tm);cursor:pointer;border:none;background:none;width:100%;font-family:var(--sans);border-radius:6px;transition:all .15s;margin-top:2px}
.sb-conv-more:hover{background:var(--bg-h);color:var(--t2)}
.sb-foot{padding:12px 16px;border-top:1px solid var(--bd);background:var(--bg-2)}
.sb-user{display:flex;align-items:center;gap:10px;cursor:pointer;padding:4px;border-radius:8px;transition:background .15s}
.sb-user:hover{background:var(--bg-h)}
.sb-av{width:28px;height:28px;border-radius:50%;background:var(--gp);display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:#fff;flex-shrink:0;position:relative}
.sb-av::before{content:'';position:absolute;inset:-2px;border-radius:50%;background:var(--gp);z-index:-1;opacity:.4}
.sb-uname{font-size:12px;color:var(--t2);flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.sb-uplan{font-size:10px;font-weight:600;padding:2px 6px;border-radius:4px;background:var(--gp);color:#fff;letter-spacing:.3px;text-transform:uppercase}

/* MAIN */
.main{flex:1;display:flex;flex-direction:column;overflow:hidden;min-width:0;background:var(--bg-0);position:relative;z-index:2}
.hdr{display:flex;align-items:center;padding:0 16px;padding-top:var(--st);height:calc(48px + var(--st));background:var(--bg-1);border-bottom:1px solid var(--bd);flex-shrink:0;gap:10px}
.ham{display:none;background:none;border:none;color:var(--t2);font-size:20px;cursor:pointer;padding:4px}
.hdr-t{flex:1;font-size:14px;font-weight:500;color:var(--t1);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-family:var(--sans)}
.mod-pill{height:24px;padding:0 10px;border-radius:12px;font-family:var(--mono);font-size:11px;font-weight:600;cursor:pointer;outline:none;appearance:none;-webkit-appearance:none;border:1px solid;background-repeat:no-repeat;background-position:right 6px center;padding-right:18px;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='8' height='8'%3E%3Cpath fill='%23606078' d='M0 2l4 4 4-4z'/%3E%3C/svg%3E")}
.mod-pill option{background:var(--bg-2);color:var(--t1)}
.mod-pill.haiku{background-color:rgba(74,222,128,.08);color:var(--g);border-color:rgba(74,222,128,.2)}
.mod-pill.sonnet{background-color:rgba(155,89,252,.08);color:var(--p);border-color:rgba(155,89,252,.2)}
.mod-pill.claude-code{background-color:rgba(251,191,36,.08);color:#FBBF24;border-color:rgba(251,191,36,.2)}
.mod-pill:disabled{opacity:.4;cursor:not-allowed}
.on-badge{display:flex;align-items:center;gap:5px;font-size:10px;font-weight:500;color:var(--g);letter-spacing:.3px}
.on-dot{width:6px;height:6px;border-radius:50%;background:var(--g);animation:pls 2s ease-in-out infinite}
@keyframes pls{0%,100%{box-shadow:0 0 0 0 rgba(74,222,128,.4)}50%{box-shadow:0 0 6px 3px rgba(74,222,128,.3)}}
.hdr-menu{position:relative}
.hdr-mbtn{background:var(--bg-2);border:1px solid var(--bd);color:var(--t2);width:32px;height:32px;border-radius:8px;cursor:pointer;font-size:16px;display:flex;align-items:center;justify-content:center;transition:all .15s}
.hdr-mbtn:hover{color:var(--p);border-color:var(--bdf)}
.hdr-drop{position:absolute;right:0;top:38px;background:var(--bg-2);border:1px solid var(--bd);border-radius:10px;padding:6px;min-width:170px;display:none;z-index:50;box-shadow:0 12px 32px rgba(0,0,0,.5)}
.hdr-drop.show{display:block}
.hdr-drop a,.hdr-drop button{display:flex;align-items:center;gap:8px;padding:8px 12px;border-radius:6px;font-size:12px;color:var(--t2);text-decoration:none;cursor:pointer;background:none;border:none;font-family:var(--sans);width:100%;text-align:left;transition:all .12s}
.hdr-drop a:hover,.hdr-drop button:hover{background:var(--bg-h);color:var(--t1)}

/* TERMINAL */
.term{flex:1;overflow-y:auto;padding:16px;-webkit-overflow-scrolling:touch;background:var(--bg-0);position:relative;z-index:1}.term::before{content:"E";position:fixed;top:50%;left:calc(50% + var(--sw)/2);transform:translate(-50%,-50%);font-size:320px;font-weight:100;background:linear-gradient(135deg,rgba(155,89,252,.07),rgba(74,158,255,.05));-webkit-background-clip:text;-webkit-text-fill-color:transparent;pointer-events:none;z-index:0;user-select:none;line-height:1;font-family:var(--sans);filter:blur(1px)}@media(max-width:768px){.term::before{left:50%;font-size:200px}}

/* WELCOME */
.welc{text-align:center;padding:32px 16px;animation:fadeUp .5s ease-out}
@keyframes fadeUp{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
.welc-logo{height:120px;margin:0 auto 12px;display:block;filter:drop-shadow(0 0 20px rgba(155,89,252,.2)) drop-shadow(0 0 20px rgba(74,158,255,.15));animation:float 4s ease-in-out infinite}
@keyframes float{0%,100%{transform:translateY(0)}50%{transform:translateY(-6px)}}
.welc-tagline{font-size:10px;letter-spacing:3px;color:var(--tm);text-transform:uppercase;margin-bottom:16px}
.welc-dot{color:var(--p)}
.welc h1{font-family:var(--sans);font-size:28px;font-weight:300;color:var(--t1);margin-bottom:8px;letter-spacing:-0.5px}
.welc h1 span{font-weight:700;background:var(--gp);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.welc p{font-size:14px;color:var(--tm);max-width:300px;margin:0 auto;font-weight:300}
.qgrid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;max-width:480px;margin:28px auto 0}
@media(max-width:600px){.qgrid{grid-template-columns:repeat(2,1fr)}}
@media(max-width:380px){.qgrid{grid-template-columns:1fr}}
.qc{background:var(--bg-2);border:1px solid var(--bd);border-left:2px solid var(--p);border-radius:8px;padding:16px 18px;cursor:pointer;transition:all .25s ease;text-align:left}
.qc:hover{border-color:rgba(155,89,252,.3);border-left-color:var(--bl);background:var(--bg-3);box-shadow:0 4px 20px rgba(0,0,0,.3),0 0 20px rgba(155,89,252,.04);transform:translateY(-2px)}
.qc .qi{display:none}
.qc .qt{font-family:var(--sans);font-size:14px;font-weight:600;color:var(--t1);letter-spacing:.2px}
.qc .qd{font-size:12px;color:var(--t2);margin-top:3px;font-weight:400}

/* MESSAGES */
.ml{margin-bottom:2px;animation:mIn .15s ease-out;display:flex;flex-direction:column}
@keyframes mIn{from{opacity:0}to{opacity:1}}
.ml.user{align-items:flex-end}
.ml.assistant{align-items:flex-start;width:100%}
.mh{display:flex;align-items:center;gap:8px;margin-bottom:4px}
.ml.user .mh{flex-direction:row-reverse}
.mav{width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.ml.user .mav{background:var(--gp);font-size:10px;font-weight:700;color:#fff}
.ml.assistant .mav{display:none}
.mn{font-size:11px;font-weight:600;letter-spacing:.2px}.ml.user .mn{color:var(--t2)}.ml.assistant .mn{color:var(--p)}
.mt{font-family:var(--mono);font-size:10px;font-weight:400;color:var(--tm);opacity:0;transition:opacity .2s}.ml:hover .mt{opacity:1}
.mb-wrap{max-width:100%;position:relative;width:100%}
.ml.assistant .mb-wrap{max-width:100%;width:100%}
.ml.user .mb-wrap{background:var(--bg-3);border:1px solid var(--bd);border-radius:12px 12px 4px 12px;padding:10px 14px;max-width:75%;box-shadow:0 2px 8px rgba(0,0,0,.2)}
.ml.user .mb-wrap:hover{border-color:var(--bdh)}
.ml.assistant .mb-wrap{padding:0}
.mb{white-space:pre-wrap;word-wrap:break-word;font-size:14px;line-height:1.65;color:var(--t1);font-weight:400;font-family:var(--mono)}
.ml.user .mb{color:var(--t2);font-size:13px;font-family:var(--sans)}
.mb h1,.mb h2,.mb h3{font-family:var(--sans);color:var(--t1);margin:14px 0 6px}.mb h1{font-size:18px;font-weight:700}.mb h2{font-size:16px;font-weight:700}.mb h3{font-size:14px;font-weight:700;color:var(--p)}
.mb ul,.mb ol{margin:4px 0;padding-left:20px}.mb li{margin-bottom:3px;color:var(--t1);font-family:var(--mono);font-size:14px}.mb li::marker{color:var(--p)}.mb p{margin:3px 0 8px}
.mb a{color:var(--bl);text-decoration:none}.mb a:hover{text-decoration:underline}
.mb strong{font-weight:700;color:#FFFFFF}.mb em{font-style:italic;color:var(--t2)}
.mb code{background:rgba(155,89,252,.08);padding:1px 5px;border-radius:3px;font-family:var(--mono);font-size:13px;font-weight:500;color:#C4B5FD}
.mb pre{background:rgba(0,0,0,.5);border:none;border-left:2px solid var(--p);border-radius:4px;padding:12px 16px;margin:6px 0;overflow-x:auto;font-size:13px;line-height:1.5;tab-size:2}
.mb pre code{background:none;padding:0;color:var(--t1);font-size:13px;font-weight:400}
.mb table{border-collapse:collapse;margin:8px 0;font-size:13px;font-family:var(--mono)}.mb th,.mb td{border:1px solid rgba(255,255,255,.06);padding:6px 10px;text-align:left}.mb th{background:rgba(155,89,252,.06);color:var(--p);font-weight:600}
.mb blockquote{border-left:2px solid var(--p);padding:4px 12px;margin:6px 0;color:var(--t2);font-size:13px}
.mb hr{border:none;border-top:1px solid rgba(155,89,252,.15);margin:12px 0}
/* THINKING */
.think{margin-bottom:8px;animation:mIn .15s ease-out}
.think-in{display:flex;align-items:center;gap:10px;padding:8px 0}
.think-logo{width:8px;height:8px;border-radius:50%;background:var(--p);animation:thinkDot 1.5s ease-in-out infinite;flex-shrink:0;box-shadow:0 0 12px rgba(155,89,252,.5)}
@keyframes thinkDot{0%,100%{opacity:.4;box-shadow:0 0 8px rgba(155,89,252,.3)}50%{opacity:1;box-shadow:0 0 16px rgba(155,89,252,.6)}}
.think-t{font-size:13px;color:var(--p);font-weight:500;font-family:var(--mono)}
.dots{display:none}
.scur{display:inline-block;width:2px;height:14px;background:var(--p);animation:cbk .6s step-end infinite;vertical-align:text-bottom;margin-left:1px;border-radius:1px}
@keyframes cbk{0%,50%{opacity:1}51%,100%{opacity:0}}
/* TOOLS - Claude Code terminal style */
.tblk{margin:4px 0;border:none;border-radius:0;overflow:hidden;font-size:13px;background:transparent;border-left:2px solid transparent;padding-left:0;transition:all .2s}
.thd{display:flex;align-items:center;gap:8px;padding:4px 0;cursor:pointer;font-family:var(--mono)}
.thd:hover{background:rgba(155,89,252,.03)}
.tdot{width:8px;height:8px;border-radius:50%;flex-shrink:0;transition:all .3s}
.tlb{color:var(--t1);font-weight:500;font-family:var(--mono);flex:1;font-size:13px}
.tlb .tn{color:var(--p);font-weight:700}
.tlb .ta{color:var(--t2);font-weight:400}
.tdr{color:var(--tm);font-size:11px;font-family:var(--mono)}
.tout{padding:0 0 0 18px;max-height:160px;overflow-y:auto;color:var(--t2);font-size:12px;font-family:var(--mono);display:none;border:none;background:transparent}
.tout pre{margin:0;background:transparent;border:none;padding:0;font-size:12px;color:var(--t2);white-space:pre-wrap}
.tout .tline{color:var(--tm);display:flex;gap:4px}
.tout .tline::before{content:'└';color:var(--tm)}
.tblk.act{border-left-color:var(--p)}
.tblk.act .tdot{background:var(--p);box-shadow:0 0 10px rgba(155,89,252,.5);animation:dotPulse 1.2s ease-in-out infinite}
.tblk.act .tlb .tn{color:var(--p)}
.tblk.done .tdot{background:var(--g);box-shadow:0 0 6px rgba(74,222,128,.3)}
.tblk.err .tdot{background:var(--r);box-shadow:0 0 6px rgba(248,113,113,.3)}
@keyframes dotPulse{0%,100%{opacity:.6;transform:scale(1)}50%{opacity:1;transform:scale(1.3)}}
.tblk.open .tout,.tblk.act .tout{display:block}
.tmore{color:var(--tm);font-size:11px;cursor:pointer;padding:2px 0 2px 18px;font-family:var(--mono)}.tmore:hover{color:var(--p)}
.fcard{display:flex;align-items:center;gap:12px;padding:14px;margin:8px 0;background:var(--bg-2);border:1px solid var(--bd);border-radius:12px;transition:border-color .2s}.fcard:hover{border-color:var(--bdf)}
.ficon{width:42px;height:42px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;background:var(--pg);border:1px solid rgba(124,92,252,.15);flex-shrink:0}
.finfo{flex:1;min-width:0}.fname{font-size:13px;font-weight:600;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.fmeta{font-size:10px;color:var(--tm);margin-top:2px}
.fbtn{padding:7px 14px;border-radius:8px;font-family:var(--sans);font-size:11px;font-weight:600;cursor:pointer;text-decoration:none;border:none;transition:all .15s}
.fbtn:active{transform:scale(.95)}.fbtn.pr{background:var(--p);color:#fff}.fbtn.pr:hover{background:var(--ph)}.fbtn.sc{background:var(--bg-h);color:var(--p);border:1px solid var(--bd)}
.eline{color:var(--r);background:rgba(248,113,113,.08);border:1px solid rgba(248,113,113,.15);border-radius:8px;padding:8px 12px;margin:8px 0;font-size:12px}
.mission-card{background:var(--bg-2);border:1px solid var(--pgs);border-radius:12px;padding:14px;margin:10px 0}
.mission-card .mctitle{font-family:var(--sans);font-size:14px;font-weight:600;color:var(--p);margin-bottom:8px;display:flex;align-items:center;gap:8px}
.mission-bar{height:4px;background:var(--bg-h);border-radius:2px;overflow:hidden;margin:8px 0}.mission-fill{height:100%;background:var(--p);border-radius:2px;transition:width .5s ease}
.mission-step{display:flex;align-items:center;gap:8px;padding:3px 0;font-size:12px;color:var(--t2)}.mission-step .ms-icon{width:16px;text-align:center;flex-shrink:0}
.mission-step.done{color:var(--g)}.mission-step.running{color:var(--p)}.mission-step.failed{color:var(--r)}.mission-step.pending{color:var(--tm)}
.mission-done{background:rgba(74,222,128,.05);border:1px solid rgba(74,222,128,.2);border-radius:12px;padding:14px;margin:10px 0}.mission-done .mctitle{color:var(--g)}

/* INPUT */
.input-area{flex-shrink:0;padding:12px 16px;padding-bottom:calc(12px + var(--sb));background:var(--bg-1);border-top:1px solid var(--bd);z-index:2;position:relative}
.file-preview{padding:8px 12px;display:flex;align-items:center;gap:10px;background:var(--bg-2);border:1px solid var(--bd);border-radius:10px;margin-bottom:8px;animation:mIn .2s ease-out}
.file-preview .fp-thumb{width:48px;height:48px;border-radius:8px;object-fit:cover;background:var(--bg-3);display:flex;align-items:center;justify-content:center;font-size:24px;flex-shrink:0;overflow:hidden}
.file-preview .fp-thumb img{width:100%;height:100%;object-fit:cover;border-radius:8px}
.file-preview .fp-info{flex:1;min-width:0}
.file-preview .fp-name{font-size:13px;font-weight:500;color:var(--t1);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.file-preview .fp-meta{font-size:11px;color:var(--tm);margin-top:2px}
.file-preview .fp-close{width:28px;height:28px;border-radius:50%;border:none;background:var(--bg-h);color:var(--t2);cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .15s;font-size:14px;flex-shrink:0}
.file-preview .fp-close:hover{background:var(--r);color:#fff}
.audio-preview{padding:10px 12px;display:flex;align-items:center;gap:10px;background:var(--bg-2);border:1px solid var(--bd);border-radius:10px;margin-bottom:8px;animation:mIn .2s ease-out}
.audio-preview .ap-play{width:36px;height:36px;border-radius:50%;border:none;background:var(--p);color:#fff;cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0;transition:all .15s}
.audio-preview .ap-play:hover{background:var(--ph)}
.audio-preview .ap-bar{flex:1;height:4px;background:var(--bg-h);border-radius:2px;overflow:hidden}
.audio-preview .ap-fill{height:100%;background:var(--p);border-radius:2px;width:0%;transition:width .1s linear}
.audio-preview .ap-dur{font-size:11px;color:var(--tm);font-family:var(--mono);flex-shrink:0;min-width:35px;text-align:right}
.audio-preview .ap-del{width:28px;height:28px;border-radius:50%;border:none;background:var(--bg-h);color:var(--t2);cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0;transition:all .15s}
.audio-preview .ap-del:hover{background:var(--r);color:#fff}
.ibox{position:relative;display:flex;align-items:flex-end;gap:6px}
.attach-btn{width:40px;height:40px;border-radius:50%;border:none;background:transparent;color:var(--tm);cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .15s;font-size:18px;flex-shrink:0}
.attach-btn:hover{color:var(--bl);background:var(--bg-h)}
.ibox textarea{flex:1;background:var(--bg-2);border:1px solid var(--bd);border-radius:14px;padding:14px 16px;color:var(--t1);font-family:var(--sans);font-size:14px;line-height:1.5;resize:none;outline:none;max-height:120px;min-height:20px;transition:border-color .2s,box-shadow .2s}
.ibox textarea::placeholder{color:var(--tm);font-weight:300}
.ibox textarea:focus{border-color:var(--bdf);box-shadow:0 0 20px rgba(155,89,252,.08)}
.ibox-right{display:flex;align-items:center;flex-shrink:0;position:relative;width:40px;height:40px}
.sbtn,.mic-btn{position:absolute;width:40px;height:40px;border-radius:50%;border:none;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .15s ease}
.sbtn{background:var(--gp);color:#fff;opacity:0;transform:scale(.7);pointer-events:none}
.sbtn.vis{opacity:1;transform:scale(1);pointer-events:auto}
.sbtn:hover{background:var(--gph);transform:scale(1.05);box-shadow:0 0 20px var(--pg)}.sbtn:active{transform:scale(.95)}.sbtn:disabled{opacity:.15;cursor:not-allowed;transform:none}
.sbtn svg{width:16px;height:16px}
.mic-btn{background:transparent;color:var(--tm);opacity:1;transform:scale(1)}
.mic-btn.vis{opacity:1;transform:scale(1);pointer-events:auto}
.mic-btn.hid{opacity:0;transform:scale(.7);pointer-events:none}
.mic-btn:hover{color:var(--p);background:var(--bg-h)}
.mic-btn.recording{background:var(--r);color:#fff;animation:micPulse 1.5s ease-in-out infinite}
@keyframes micPulse{0%,100%{box-shadow:0 0 0 0 rgba(248,113,113,.4)}50%{box-shadow:0 0 0 8px rgba(248,113,113,0)}}
.mic-btn svg{width:18px;height:18px}
.rec-timer{position:absolute;right:50px;bottom:12px;font-size:11px;font-family:var(--mono);color:var(--r);opacity:0;transition:opacity .2s}
.rec-timer.vis{opacity:1}
/* Attach Menu */
.attach-menu{position:absolute;bottom:calc(100% + 8px);left:0;background:var(--bg-2);border:1px solid var(--bd);border-radius:12px;padding:6px;min-width:200px;z-index:60;box-shadow:0 -10px 30px rgba(0,0,0,.3);transform-origin:bottom left;animation:amIn .2s ease}
@keyframes amIn{from{opacity:0;transform:translateY(8px) scale(.95)}to{opacity:1;transform:translateY(0) scale(1)}}
.attach-menu .am-item{display:flex;align-items:center;gap:10px;padding:10px 14px;border-radius:8px;cursor:pointer;font-size:13px;color:var(--t2);font-family:var(--sans);transition:background .12s;border:none;background:none;width:100%;text-align:left}
.attach-menu .am-item:hover{background:var(--bg-h);color:var(--t1)}
.attach-menu .am-item .am-icon{font-size:16px;width:22px;text-align:center;flex-shrink:0}
/* Drag overlay */
.drag-overlay{position:absolute;inset:0;background:rgba(9,9,15,.85);z-index:80;display:none;align-items:center;justify-content:center;flex-direction:column;gap:12px;border:2px dashed var(--p);border-radius:16px;margin:8px;pointer-events:none}
.drag-overlay.show{display:flex}
.drag-overlay .do-icon{font-size:48px;animation:float 2s ease-in-out infinite}
.drag-overlay .do-text{font-size:16px;font-weight:500;color:var(--p)}
/* Chat file cards */
.chat-img{max-width:300px;border-radius:12px;cursor:pointer;transition:transform .2s;margin:6px 0;display:block}
.chat-img:hover{transform:scale(1.02)}
.chat-file-card{display:flex;align-items:center;gap:10px;padding:10px 14px;background:var(--bg-2);border:1px solid var(--bd);border-radius:8px;margin:6px 0;max-width:280px}
.chat-file-card .cfc-icon{font-size:24px;flex-shrink:0}
.chat-file-card .cfc-info{flex:1;min-width:0}
.chat-file-card .cfc-name{font-size:12px;font-weight:500;color:var(--t1);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.chat-file-card .cfc-meta{font-size:10px;color:var(--tm);margin-top:2px}
.chat-file-card .cfc-dl{font-size:12px;color:var(--p);text-decoration:none;flex-shrink:0}
.chat-file-card .cfc-dl:hover{text-decoration:underline}
/* Audio player inline */
.chat-audio{display:flex;align-items:center;gap:8px;padding:8px 14px;background:var(--bg-2);border:1px solid var(--bd);border-radius:20px;margin:6px 0;max-width:280px}
.chat-audio .ca-play{width:30px;height:30px;border-radius:50%;border:none;background:var(--p);color:#fff;cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:12px;flex-shrink:0;transition:all .15s}
.chat-audio .ca-play:hover{background:var(--ph)}
.chat-audio .ca-bar{flex:1;height:3px;background:var(--bg-h);border-radius:2px;overflow:hidden;cursor:pointer}
.chat-audio .ca-fill{height:100%;background:var(--p);border-radius:2px;width:0%;transition:width .1s linear}
.chat-audio .ca-dur{font-size:10px;color:var(--tm);font-family:var(--mono);flex-shrink:0}
.chat-transcription{font-style:italic;color:var(--tm);font-size:12px;margin:4px 0 6px;max-width:280px}
.ap-transcript{font-style:italic;color:var(--tm);font-size:12px;margin:4px 0 2px;padding:0 4px}
/* Lightbox */
.lightbox{position:fixed;inset:0;background:rgba(0,0,0,.9);z-index:200;display:none;align-items:center;justify-content:center;cursor:zoom-out;padding:20px}
.lightbox.show{display:flex}
.lightbox img{max-width:95vw;max-height:95vh;object-fit:contain;border-radius:8px}
.lightbox .lb-close{position:absolute;top:20px;right:20px;width:40px;height:40px;border-radius:50%;border:none;background:rgba(255,255,255,.15);color:#fff;cursor:pointer;font-size:20px;display:flex;align-items:center;justify-content:center}
/* Upload progress */
.upload-progress{height:3px;background:var(--bg-h);border-radius:2px;overflow:hidden;margin:4px 0}
.upload-progress .up-fill{height:100%;background:var(--p);border-radius:2px;transition:width .3s ease;width:0%}
/* Toast */
.toast{position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:var(--bg-2);border:1px solid var(--bd);color:var(--t1);padding:10px 20px;border-radius:10px;font-size:13px;z-index:300;box-shadow:0 8px 24px rgba(0,0,0,.5);animation:toastIn .3s ease}
.toast.error{border-color:var(--r);color:var(--r)}
@keyframes toastIn{from{opacity:0;transform:translate(-50%,10px)}to{opacity:1;transform:translate(-50%,0)}}

/* MODAL */
.modal-bg{position:fixed;inset:0;background:rgba(5,5,16,.7);backdrop-filter:blur(12px);z-index:100;display:none;align-items:center;justify-content:center;padding:20px}
.modal-bg.show{display:flex}
.modal{background:var(--bg-2);border:1px solid var(--bd);border-radius:20px;padding:28px;max-width:420px;width:100%;max-height:80vh;overflow-y:auto;box-shadow:0 30px 60px rgba(0,0,0,.5),0 0 60px rgba(155,89,252,.05)}
.modal h3{font-family:var(--sans);font-size:18px;font-weight:600;color:var(--t1);margin-bottom:16px}
.modal label{display:block;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:var(--tm);margin:12px 0 4px}
.modal input,.modal select{width:100%;padding:10px 14px;background:var(--bg-3);border:1px solid var(--bd);border-radius:8px;color:var(--t1);font-family:var(--sans);font-size:13px;outline:none;transition:border-color .2s}
.modal input:focus{border-color:var(--bdf);box-shadow:0 0 0 3px var(--pg)}
.modal .mbtn{margin-top:16px;padding:12px;width:100%;border-radius:10px;border:none;background:var(--gp);color:#fff;font-family:var(--sans);font-size:14px;font-weight:600;cursor:pointer;transition:all .2s}.modal .mbtn:hover{box-shadow:0 0 30px var(--pg);filter:brightness(1.1)}
.adm-cards{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;padding:4px 0}
.adm-card{background:var(--bg-3);border:1px solid var(--bd);border-radius:8px;padding:10px 12px}
.adm-card .al{font-size:9px;font-weight:600;letter-spacing:.5px;text-transform:uppercase;color:var(--tm)}.adm-card .av{font-family:var(--mono);font-size:22px;font-weight:600;color:var(--p);margin-top:4px}
.adm-tbl{width:100%;font-size:11px;border-collapse:collapse;margin-top:8px}
.adm-tbl th{text-align:left;color:var(--tm);font-weight:600;padding:4px 6px;border-bottom:1px solid var(--bd);font-size:9px;text-transform:uppercase}
.adm-tbl td{padding:5px 6px;border-bottom:1px solid rgba(42,42,69,.5);color:var(--t2)}
.adm-tbl select,.adm-tbl button{background:var(--bg-h);border:1px solid var(--bd);color:var(--t1);font-family:var(--sans);font-size:10px;padding:3px 6px;border-radius:4px;cursor:pointer}
.adm-tbl button:hover{color:var(--p);border-color:var(--bdf)}

/* WATERMARK — logo pulse */
@keyframes logoPulse{
  0%,100%{opacity:.025;filter:drop-shadow(0 0 15px rgba(155,89,252,.05)) drop-shadow(0 0 15px rgba(74,158,255,.05))}
  50%{opacity:.045;filter:drop-shadow(0 0 40px rgba(155,89,252,.1)) drop-shadow(0 0 40px rgba(74,158,255,.1))}
}
@keyframes glowPulse{
  0%,100%{opacity:.6;transform:translate(-50%,-50%) scale(1)}
  50%{opacity:1;transform:translate(-50%,-50%) scale(1.05)}
}
.watermark{position:fixed;top:50%;left:calc(50% + var(--sw)/2);transform:translate(-50%,-50%);pointer-events:none;z-index:0;user-select:none;opacity:.05;font-size:320px;font-weight:100;background:linear-gradient(135deg,rgba(155,89,252,.9),rgba(74,158,255,.7));-webkit-background-clip:text;-webkit-text-fill-color:transparent;line-height:1;font-family:var(--sans);filter:blur(1px)}

/* glow now handled by body::after */

/* RESPONSIVE */
@media(max-width:768px){
  .sb{position:fixed;left:0;top:0;height:100%;transform:translateX(-100%);box-shadow:4px 0 24px rgba(0,0,0,.6)}
  .sb.open{transform:translateX(0)}
  .ham{display:block}
  .sb-ov{position:fixed;inset:0;background:rgba(5,5,16,.85);backdrop-filter:blur(12px);z-index:25;display:none}.sb-ov.show{display:block}
  .qgrid{grid-template-columns:repeat(2,1fr)}
  .mb-wrap{max-width:85%!important}
  .watermark{left:50%;font-size:180px}
}
@media(min-width:769px){.ham{display:none}}
/* SR-ONLY */
.sr-only{position:absolute;width:1px;height:1px;padding:0;margin:-1px;overflow:hidden;clip:rect(0,0,0,0);white-space:nowrap;border:0}
/* STOP BUTTON */
.stop-btn{position:absolute;width:40px;height:40px;border-radius:50%;border:none;cursor:pointer;display:none;align-items:center;justify-content:center;background:var(--r);color:#fff;transition:all .2s;animation:fadeInStop .2s ease}
.stop-btn.vis{display:flex}
.stop-btn:hover{background:#ef4444;box-shadow:0 0 15px rgba(248,113,113,.3)}
.stop-btn svg{width:14px;height:14px}
@keyframes fadeInStop{from{opacity:0;transform:scale(.7)}to{opacity:1;transform:scale(1)}}
</style>
</head>
<body>
<svg style="position:absolute;width:0;height:0"><defs>
<linearGradient id="sbGrad" x1="0%" y1="0%" x2="100%" y2="0%"><stop offset="0%" stop-color="#9B59FC"/><stop offset="100%" stop-color="#4A9EFF"/></linearGradient>
<linearGradient id="wGrad" x1="0%" y1="0%" x2="100%" y2="0%"><stop offset="0%" stop-color="#9B59FC"/><stop offset="100%" stop-color="#4A9EFF"/></linearGradient>
<linearGradient id="wmGrad" x1="0%" y1="0%" x2="100%" y2="0%"><stop offset="0%" stop-color="#9B59FC"/><stop offset="100%" stop-color="#4A9EFF"/></linearGradient>
</defs></svg>
<div class="sb-ov" id="sbOv" onclick="toggleSB()"></div>
<aside class="sb" id="sb" role="navigation" aria-label="Menu lateral">
  <div class="sb-head">
    <a href="/"><img src="/static/brand/logo-sidebar.png" alt="Clow - Inteligência Infinita" class="sb-logo-img"></a>
  </div>
  <div class="sb-body">
    <div class="sb-sec"><button class="sb-new" onclick="newConv()" aria-label="Iniciar nova conversa">+ Nova Conversa</button></div>
    <div class="sb-sec" style="flex:1;display:flex;flex-direction:column;overflow:hidden">
      <div class="sb-lbl open" onclick="this.classList.toggle('open')">Conversas <span class="ar">&#9654;</span></div>
      <div class="sb-ct" style="flex:1;display:flex;flex-direction:column;overflow:hidden">
        <div class="sb-conv-search" id="convSearchWrap" style="display:none">
          <label for="convSearchInp" class="sr-only">Buscar conversas</label>
          <span class="search-icon" style="font-size:12px;color:var(--tm)">&#x25CB;</span>
          <input type="text" id="convSearchInp" placeholder="Buscar conversas..." oninput="filterConvs(this.value)" onkeydown="if(event.key==='Escape'){closeConvSearch()}">
        </div>
        <div class="sb-convs" id="convList"></div>
      </div>
    </div>
    <div class="sb-sep"></div>
    <div class="sb-sec">
      <div class="sb-lbl open" onclick="this.classList.toggle('open')">Criar <span class="ar">&#9654;</span></div>
      <div class="sb-ct">
        <button class="sb-btn" onclick="qa('Cria uma landing page de ')">Landing Page</button>
        <button class="sb-btn" onclick="qa('Gera uma planilha de ')">Planilha</button>
        <button class="sb-btn" onclick="qa('Cria uma apresentação sobre ')">Apresentação</button>
        <button class="sb-btn" onclick="qa('Faz um documento de ')">Documento</button>
        <button class="sb-btn" onclick="qa('Me faz um app de ')">Web App</button>
      </div>
    </div>
    <div class="sb-sec">
      <div class="sb-lbl" onclick="this.classList.toggle('open')">Marketing <span class="ar">&#9654;</span></div>
      <div class="sb-ct">
        <button class="sb-btn" onclick="qa('Analise minha campanha de tráfego pago ')">Tráfego Pago</button>
        <button class="sb-btn" onclick="qa('Gera copy para anúncio de ')">Copy</button>
        <button class="sb-btn" onclick="qa('Cria conteúdo para instagram sobre ')">Redes Sociais</button>
        <button class="sb-btn" onclick="qa('Cria sequência de emails para ')">Email Marketing</button>
        <button class="sb-btn" onclick="qa('Faz auditoria SEO de ')">SEO</button>
        <button class="sb-btn" onclick="qa('Define estratégia de precificação para ')">Pricing</button>
      </div>
    </div>
    <div class="sb-sec">
      <div class="sb-lbl" onclick="this.classList.toggle('open')">Negócios <span class="ar">&#9654;</span></div>
      <div class="sb-ct">
        <button class="sb-btn" onclick="qa('Cria proposta comercial para ')">Proposta</button>
        <button class="sb-btn" onclick="qa('Analise métricas SaaS: ')">Métricas SaaS</button>
        <button class="sb-btn" onclick="qa('Cria contrato de ')">Contrato</button>
      </div>
    </div>
    <div class="sb-sec">
      <div class="sb-lbl" onclick="this.classList.toggle('open')">Missões <span class="ar">&#9654;</span></div>
      <div class="sb-ct">
        <button class="sb-btn" onclick="qa('/mission Cria um site completo para ')">Site Completo</button>
        <button class="sb-btn" onclick="qa('/mission Campanha de tráfego para ')">Campanha</button>
        <button class="sb-btn" onclick="qa('/mission Setup digital completo para ')">Setup Digital</button>
      </div>
    </div>
    <div class="sb-sep"></div>
    <div class="sb-sec">
      <div class="sb-lbl" onclick="this.classList.toggle('open')">Conexões <span class="ar">&#9654;</span></div>
      <div class="sb-ct">
        <button class="sb-btn" onclick="sendCmd('/connections')">Ver conexões</button>
        <button class="sb-btn" onclick="sendCmd('/connect')">Conectar</button>
      </div>
    </div>
    <div class="sb-sec" id="admSec" style="display:none">
      <div class="sb-lbl" onclick="this.classList.toggle('open')">Admin <span class="ar">&#9654;</span></div>
      <div class="sb-ct">
        <button class="sb-btn" onclick="showAdmUsers()">Usuários</button>
        <button class="sb-btn" onclick="showAdmStats()">Consumo</button>
        <button class="sb-btn" onclick="showCreateUsr()">Cadastrar</button>
      </div>
    </div>
  </div>
  <div class="sb-foot"><div class="sb-user" id="sbUsr"><div class="sb-av" id="sbAv">?</div><span class="sb-uname" id="sbEm">...</span><span class="sb-uplan" id="sbPl">...</span></div></div>
</aside>
<div class="main" role="main">
  <div class="hdr">
    <button class="ham" onclick="toggleSB()" aria-label="Abrir menu lateral">&#9776;</button>
    <div class="hdr-t" id="hdrT">Nova conversa</div>
    <label for="modSel" class="sr-only">Modelo AI</label>
    <select class="mod-pill haiku" id="modSel" onchange="onMod()" title="Modelo AI" aria-label="Selecionar modelo AI"><option value="haiku">Haiku</option><option value="sonnet">Sonnet</option></select>
    <div class="on-badge" id="onBdg" aria-live="polite"><span class="on-dot"></span><span id="onLbl">online</span></div>
    <div class="hdr-menu">
      <button class="hdr-mbtn" onclick="togDrop()" aria-label="Menu de opções">&#x22EE;</button>
      <div class="hdr-drop" id="hdrDrop" role="menu">
        <button onclick="sendCmd('/usage');clsDrop()">&#x1F4CA; Meu consumo</button>
        <button onclick="sendCmd('/plan');clsDrop()">&#x1F4E6; Meu plano</button>
        <button onclick="sendCmd('/memories');clsDrop()">Memórias</button>
        <button onclick="sendCmd('/help');clsDrop()">Ajuda</button>
        <a href="/logout">Sair</a>
      </div>
    </div>
  </div>
  <div class="watermark empty" id="wmark">&#8734;</div>
  <div class="term" id="term">
    <div class="welc" id="welc">
      <img src="/static/brand/logo.png" alt="Clow - Inteligência Infinita" class="welc-logo">
      <div class="welc-tagline">INTELIG&Ecirc;NCIA INFINITA <span class="welc-dot">&bull;</span> POSSIBILIDADES PREMIUM</div>
      <h1>O que vamos <span>criar</span>?</h1>
      <p>Escolha abaixo ou descreva o que precisa</p>
      <div class="qgrid">
        <div class="qc" role="button" tabindex="0" onclick="qa('Cria uma landing page de ')"><div class="qt">Landing Page</div><div class="qd">Site completo e responsivo</div></div>
        <div class="qc" role="button" tabindex="0" onclick="qa('Gera uma planilha de ')"><div class="qt">Planilha</div><div class="qd">Excel com fórmulas</div></div>
        <div class="qc" role="button" tabindex="0" onclick="qa('Cria uma apresentação sobre ')"><div class="qt">Apresentação</div><div class="qd">PowerPoint profissional</div></div>
        <div class="qc" role="button" tabindex="0" onclick="qa('Me faz um app de ')"><div class="qt">Web App</div><div class="qd">App funcional completo</div></div>
        <div class="qc" role="button" tabindex="0" onclick="qa('Gera copy para anúncio de ')"><div class="qt">Copy Ads</div><div class="qd">Textos para anúncios</div></div>
        <div class="qc" role="button" tabindex="0" onclick="qa('/mission Setup digital completo para ')"><div class="qt">Missão</div><div class="qd">Projeto autônomo completo</div></div>
      </div>
    </div>
  </div>
  <div class="drag-overlay" id="dragOv"><div class="do-icon">&#x1F4CE;</div><div class="do-text">Solte o arquivo aqui</div></div>
  <div class="input-area">
    <div id="filePreview"></div>
    <div class="ibox">
      <button class="attach-btn" id="attachBtn" onclick="toggleAttachMenu()" title="Anexar arquivo" aria-label="Anexar arquivo">&#x1F4CE;</button>
      <div class="attach-menu" id="attachMenu" style="display:none" role="menu" aria-label="Tipo de arquivo">
        <button class="am-item" onclick="pickFile('image/*','.jpg,.jpeg,.png,.gif,.webp')">Foto / Imagem</button>
        <button class="am-item" onclick="pickFile('','.pdf,.docx,.doc,.txt,.md')">Documento</button>
        <button class="am-item" onclick="pickFile('','.xlsx,.xls,.csv')">Planilha</button>
        <button class="am-item" onclick="pickFile('','*')">Qualquer arquivo</button>
      </div>
      <label for="inp" class="sr-only">Digite sua mensagem</label>
      <textarea id="inp" rows="1" placeholder="O que você precisa?" autofocus aria-label="Digite sua mensagem"></textarea>
      <div class="ibox-right">
        <button class="mic-btn vis" id="micBtn" onclick="toggleRec()" title="Gravar áudio" aria-label="Gravar áudio"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 1a3 3 0 0 0-3 3v8a3 3 0 0 0 6 0V4a3 3 0 0 0-3-3z"/><path d="M19 10v2a7 7 0 0 1-14 0v-2"/><line x1="12" y1="19" x2="12" y2="23"/><line x1="8" y1="23" x2="16" y2="23"/></svg></button>
        <button class="sbtn" id="sBtn" onclick="sendMessage()" aria-label="Enviar mensagem"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><line x1="5" y1="12" x2="19" y2="12"/><polyline points="12 5 19 12 12 19"/></svg></button>
        <button class="stop-btn" id="stopBtn" onclick="stopGeneration()" aria-label="Parar geração"><svg viewBox="0 0 24 24" fill="currentColor"><rect x="6" y="6" width="12" height="12" rx="2"/></svg></button>
      </div>
    </div>
  </div>
  <div class="lightbox" id="lightbox" onclick="closeLightbox()"><button class="lb-close" onclick="closeLightbox()">&times;</button><img id="lbImg" src=""></div>
</div>
<div class="modal-bg" id="modalBg" onclick="if(event.target===this)clsModal()" role="dialog" aria-modal="true"><div class="modal" id="modalC"></div></div>
<script>
const INF='<svg viewBox="0 0 32 32" style="width:16px;height:16px"><path d="M8 16c0-3 2-6 5-6s5 3 8 6c3 3 5 6 8 6s5-3 5-6-2-6-5-6-5 3-8 6c-3 3-5 6-8 6s-5-3-5-6z" transform="translate(-5,0) scale(.95)" fill="none" stroke="var(--p)" stroke-width="3" stroke-linecap="round"/></svg>';
const T=document.getElementById('term'),I=document.getElementById('inp'),SB=document.getElementById('sBtn');
let ws=null,proc=false,curMsg=null,curBody=null,curTool=null,tStart=0,tTimer=null,rA=0,http=false,hSid='',raw='',me=null,cid='',selMod='haiku';
async function init(){
  try{const r=await fetch('/api/v1/me');me=await r.json();
    document.getElementById('sbAv').textContent=me.email[0].toUpperCase();
    document.getElementById('sbEm').textContent=me.email;
    document.getElementById('sbPl').textContent=me.plan;
    if(me.is_admin)document.getElementById('admSec').style.display='block';
    initMod(me.plan,me.is_admin);
  }catch(e){}
  loadConvs();connectWS();
}
function initMod(plan,adm){const s=document.getElementById('modSel');if(adm){selMod='claude-code';s.style.display='none';const b=document.createElement('span');b.className='mod-pill claude-code';b.style.cssText='pointer-events:none;cursor:default;padding:4px 12px;font-weight:600';b.textContent='Opus';s.parentNode.insertBefore(b,s);return}const can=plan==='pro'||plan==='unlimited';if(!can){s.value='haiku';selMod='haiku';s.disabled=true;const o=s.querySelector('option[value="sonnet"]');if(o)o.disabled=true}s.className='mod-pill '+selMod}
function onMod(){const s=document.getElementById('modSel');selMod=s.value;s.className='mod-pill '+selMod}
function toggleSB(){document.getElementById('sb').classList.toggle('open');document.getElementById('sbOv').classList.toggle('show')}
function togDrop(){document.getElementById('hdrDrop').classList.toggle('show')}
function clsDrop(){document.getElementById('hdrDrop').classList.remove('show')}
document.addEventListener('click',e=>{if(!e.target.closest('.hdr-menu'))clsDrop()});
let pinnedConvs=JSON.parse(localStorage.getItem('clow_pinned')||'[]');
let allConvsCache=[];
let showAllPast=false;
let activeCtxMenu=null;

function closeCtxMenu(){if(activeCtxMenu){activeCtxMenu.remove();activeCtxMenu=null}}
document.addEventListener('click',e=>{if(activeCtxMenu&&!e.target.closest('.conv-ctx-menu')&&!e.target.closest('.ca-btn'))closeCtxMenu()});
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeCtxMenu()});

function getDateGroup(ts){
  const d=new Date(ts*1000);const now=new Date();
  const today=new Date(now.getFullYear(),now.getMonth(),now.getDate());
  const yesterday=new Date(today);yesterday.setDate(today.getDate()-1);
  const week=new Date(today);week.setDate(today.getDate()-7);
  const convDay=new Date(d.getFullYear(),d.getMonth(),d.getDate());
  if(convDay>=today)return'Hoje';
  if(convDay>=yesterday)return'Ontem';
  if(convDay>=week)return'Ultimos 7 dias';
  return'Anteriores';
}

function smartTitle(t){
  if(!t||t==='Nova conversa')return'Nova conversa';
  const generic=['oi','ola','bom dia','boa tarde','boa noite','hey','hello','hi','e ai','fala'];
  const w=t.trim().toLowerCase();
  if(generic.includes(w))return'Nova conversa';
  if(t.length>28)return t.substring(0,28)+'...';
  return t;
}

async function loadConvs(){try{
  const r=await fetch('/api/v1/conversations');const d=await r.json();
  const el=document.getElementById('convList');
  allConvsCache=d.conversations||[];
  const pinned=allConvsCache.filter(c=>pinnedConvs.includes(c.id));
  const unpinned=allConvsCache.filter(c=>!pinnedConvs.includes(c.id));
  // Show search icon if 3+ conversations
  document.getElementById('convSearchWrap').style.display=allConvsCache.length>=3?'block':'none';
  let h='';
  // PINNED
  if(pinned.length){
    h+='<div class="sb-grp-label">FIXADAS</div>';
    pinned.forEach(c=>{h+=convBtn(c,true)});
  }
  // GROUP BY DATE
  const groups={};
  const maxShow=showAllPast?unpinned.length:10;
  unpinned.slice(0,maxShow).forEach(c=>{
    const g=getDateGroup(c.updated_at||c.created_at);
    if(!groups[g])groups[g]=[];
    groups[g].push(c);
  });
  const order=['Hoje','Ontem','Ultimos 7 dias','Anteriores'];
  order.forEach(g=>{
    if(groups[g]&&groups[g].length){
      h+='<div class="sb-grp-label">'+g+'</div>';
      groups[g].forEach(c=>{h+=convBtn(c,false)});
    }
  });
  if(!showAllPast&&unpinned.length>10){
    h+='<button class="sb-conv-more" onclick="showAllPast=true;loadConvs()">Ver anteriores ('+unpinned.length+')</button>';
  }
  el.innerHTML=h||'<div style="padding:12px 8px;color:var(--tm);font-size:12px;text-align:center">Nenhuma conversa</div>';
  bindConvEvents();
}catch(e){}}

function convBtn(c,isPinned){
  const t=smartTitle(c.title);
  const isAct=c.id===cid;
  return '<div class="sb-conv-item'+(isAct?' act':'')+'" data-id="'+c.id+'" data-title="'+esc(c.title)+'">'
    +'<span class="conv-icon">'+(isPinned?'':'&#x1F4AC;')+'</span>'
    +(isPinned?'<span class="conv-pin-static">&#x1F4CC;</span>':'')
    +'<span class="conv-title">'+esc(t)+'</span>'
    +'<span class="conv-actions">'
    +'<button class="ca-btn ca-pin" data-cid="'+c.id+'" data-pinned="'+(isPinned?'1':'0')+'" title="'+(isPinned?'Desafixar':'Fixar')+'" aria-label="'+(isPinned?'Desafixar conversa':'Fixar conversa')+'">&#x1F4CC;</button>'
    +'<button class="ca-btn ca-menu" data-cid="'+c.id+'" data-pinned="'+(isPinned?'1':'0')+'" title="Menu" aria-label="Opções da conversa">\u22EF</button>'
    +'</span></div>';
}
function bindConvEvents(){
  // Bind click on conv items (load conversation)
  document.querySelectorAll('.sb-conv-item').forEach(function(el){
    el.addEventListener('click',function(e){
      if(e.target.closest('.ca-btn'))return; // Don't load if clicking action buttons
      loadConv(el.getAttribute('data-id'));
    });
  });
  // Bind pin buttons
  document.querySelectorAll('.ca-pin').forEach(function(btn){
    btn.addEventListener('click',function(e){
      e.stopPropagation();
      togglePin(btn.getAttribute('data-cid'));
    });
  });
  // Bind menu buttons
  document.querySelectorAll('.ca-menu').forEach(function(btn){
    btn.addEventListener('click',function(e){
      e.stopPropagation();
      showCtxMenu(e,btn.getAttribute('data-cid'),btn.getAttribute('data-pinned')==='1');
    });
  });
}

function showCtxMenu(e,id,isPinned){
  e.preventDefault();e.stopPropagation();
  closeCtxMenu();
  const menu=document.createElement('div');
  menu.className='conv-ctx-menu';
  // Block ALL clicks inside menu from bubbling to document
  menu.addEventListener('click',function(ev){ev.stopPropagation()});
  menu.addEventListener('mousedown',function(ev){ev.stopPropagation()});
  // Build menu items with addEventListener (not inline onclick)
  const items=[
    {icon:'\u270F\uFE0F',label:'Renomear',action:function(){closeCtxMenu();startRename(id)}},
    {icon:'\uD83D\uDCCC',label:isPinned?'Desafixar':'Fixar conversa',action:function(){closeCtxMenu();togglePin(id)}},
    {icon:'\uD83D\uDCCB',label:'Copiar conversa',action:function(){closeCtxMenu();copyConv(id)}},
    {sep:true},
    {icon:'\uD83D\uDDD1\uFE0F',label:'Deletar conversa',danger:true,action:function(ev){confirmDel(id,ev.currentTarget,menu)}}
  ];
  items.forEach(function(item){
    if(item.sep){const s=document.createElement('div');s.className='conv-ctx-sep';menu.appendChild(s);return}
    const btn=document.createElement('button');
    btn.className='ctx-item'+(item.danger?' danger':'');
    btn.innerHTML='<span class="ctx-icon">'+item.icon+'</span>'+item.label;
    btn.addEventListener('click',item.action);
    menu.appendChild(btn);
  });
  document.body.appendChild(menu);
  // Position near click
  const x=Math.min(e.clientX,window.innerWidth-200);
  const y=Math.min(e.clientY,window.innerHeight-menu.offsetHeight-10);
  menu.style.left=x+'px';menu.style.top=y+'px';
  activeCtxMenu=menu;
}

function confirmDel(id,btn,menu){
  const confirm=document.createElement('div');
  confirm.className='conv-del-confirm';
  const span=document.createElement('span');span.textContent='Tem certeza?';
  const yes=document.createElement('button');yes.className='del-yes';yes.textContent='Sim';
  yes.addEventListener('click',function(ev){ev.stopPropagation();delConv(id)});
  const no=document.createElement('button');no.className='del-no';no.textContent='Nao';
  no.addEventListener('click',function(ev){ev.stopPropagation();closeCtxMenu()});
  confirm.appendChild(span);confirm.appendChild(yes);confirm.appendChild(no);
  btn.replaceWith(confirm);
}

async function delConv(id){
  closeCtxMenu();
  const el=document.querySelector('.sb-conv-item[data-id="'+id+'"]');
  if(el){el.classList.add('leaving');await new Promise(r=>setTimeout(r,200))}
  try{await fetch('/api/v1/conversations/'+id,{method:'DELETE'});
    if(id===cid){cid=null;T.innerHTML='';showWelc();document.getElementById('hdrT').textContent='Nova conversa'}
    pinnedConvs=pinnedConvs.filter(x=>x!==id);localStorage.setItem('clow_pinned',JSON.stringify(pinnedConvs));
    loadConvs();
  }catch(e){}
}

function startRename(id){
  closeCtxMenu();
  const el=document.querySelector('.sb-conv-item[data-id="'+id+'"]');
  if(!el)return;
  const titleEl=el.querySelector('.conv-title');
  const oldTitle=el.getAttribute('data-title')||titleEl.textContent;
  const inp=document.createElement('input');
  inp.className='conv-rename-input';inp.value=oldTitle;inp.maxLength=50;
  titleEl.replaceWith(inp);inp.focus();inp.select();
  const save=async()=>{
    const v=inp.value.trim()||oldTitle;
    try{await fetch('/api/v1/conversations/'+id+'/title',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({title:v})});
      if(id===cid)document.getElementById('hdrT').textContent=v;
      loadConvs();
    }catch(e){loadConvs()}
  };
  inp.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();save()}if(e.key==='Escape'){loadConvs()}});
  inp.addEventListener('blur',save);
  inp.addEventListener('click',e=>e.stopPropagation());
}

async function copyConv(id){
  try{const r=await fetch('/api/v1/conversations/'+id+'/messages');const d=await r.json();
    let txt='';d.messages.forEach(m=>{txt+=(m.role==='user'?'Voce':'Clow')+': '+m.content+'\\n\\n'});
    await navigator.clipboard.writeText(txt);
  }catch(e){try{
    const ta=document.createElement('textarea');ta.value='Conversa copiada';document.body.appendChild(ta);ta.select();document.execCommand('copy');ta.remove();
  }catch(e2){}}
}

function togglePin(id){
  if(pinnedConvs.includes(id)){pinnedConvs=pinnedConvs.filter(x=>x!==id)}
  else{if(pinnedConvs.length>=3)return;pinnedConvs.push(id)}
  localStorage.setItem('clow_pinned',JSON.stringify(pinnedConvs));
  loadConvs();
  // Pin bounce animation
  setTimeout(()=>{const el=document.querySelector('.sb-conv-item[data-id="'+id+'"] .conv-pin-static');if(el)el.classList.add('pin-bounce')},50);
}

function filterConvs(q){
  const el=document.getElementById('convList');
  if(!q.trim()){loadConvs();return}
  const ql=q.toLowerCase();
  const filtered=allConvsCache.filter(c=>(c.title||'').toLowerCase().includes(ql));
  let h='';
  if(!filtered.length){h='<div style="padding:12px 8px;color:var(--tm);font-size:12px;text-align:center">Nenhum resultado</div>'}
  else{filtered.forEach(c=>{h+=convBtn(c,pinnedConvs.includes(c.id))})}
  el.innerHTML=h;
  bindConvEvents();
}

function closeConvSearch(){
  document.getElementById('convSearchInp').value='';
  loadConvs();
}

async function showAllConvs(){showAllPast=true;loadConvs()}

async function newConv(){try{const r=await fetch('/api/v1/conversations',{method:'POST'});const d=await r.json();cid=d.id;convMsgCount=0;T.innerHTML='';showWelc();document.getElementById('hdrT').textContent='Nova conversa';loadConvs();if(window.innerWidth<769)toggleSB()}catch(e){}}
async function loadConv(id){cid=id;T.innerHTML='';try{const r=await fetch(`/api/v1/conversations/${id}/messages`);const d=await r.json();d.messages.forEach(m=>{if(m.role==='user')addUser(m.content,false);else{curMsg=null;curBody=null;appendTxt(m.content);finishTxt();curMsg=null;curBody=null}});const cs=await(await fetch('/api/v1/conversations')).json();const c=cs.conversations.find(x=>x.id===id);if(c)document.getElementById('hdrT').textContent=c.title;loadConvs();if(window.innerWidth<769)toggleSB()}catch(e){}}
function showWelc(){const w=document.createElement('div');w.className='welc';w.id='welc';w.innerHTML='<img src="/static/brand/logo.png" alt="Clow - Intelig\u00eancia Infinita" class="welc-logo"><div class="welc-tagline">INTELIG\u00caNCIA INFINITA <span class="welc-dot">\u2022</span> POSSIBILIDADES PREMIUM</div><h1>O que vamos <span>criar</span>?</h1><p>Escolha abaixo ou descreva o que precisa</p>';T.appendChild(w);const wm=document.getElementById('wmark');if(wm)wm.classList.add('empty')}
function connectWS(){const pr=location.protocol==='https:'?'wss:':'ws:';try{ws=new WebSocket(`${pr}//${location.host}/ws`)}catch(e){http=true;setOn('http');return}const to=setTimeout(()=>{if(!ws||ws.readyState!==1){try{ws.close()}catch(e){}http=true;setOn('http')}},4000);ws.onopen=()=>{clearTimeout(to);http=false;setOn('online');rA=0};ws.onmessage=e=>hMsg(JSON.parse(e.data));ws.onclose=()=>{clearTimeout(to);if(rA>=3){http=true;setOn('http');return}setOn('offline');setTimeout(()=>{rA++;connectWS()},Math.min(1000*rA,5000))};ws.onerror=()=>setOn('offline')}
function setOn(s){const b=document.getElementById('onBdg'),l=document.getElementById('onLbl');b.style.color=s==='offline'?'var(--r)':'var(--g)';l.textContent=s}
function hMsg(m){switch(m.type){case'thinking_start':showThink();break;case'thinking_end':hideThink();break;case'text_delta':appendTxt(m.content);break;case'text_done':finishTxt();break;case'tool_call':showTool(m.name,m.args);break;case'tool_result':showToolR(m.name,m.status,m.output);break;case'turn_complete':finishTurn();break;case'error':showErr(m.content);break}}
function sendMessage(){const t=I.value.trim();if(!t||proc)return;if(http){sendHTTP(t);return}if(!ws||ws.readyState!==1)return;addUser(t);ws.send(JSON.stringify({type:'message',content:t,model:selMod,conversation_id:cid}));I.value='';I.style.height='auto';proc=true;SB.disabled=true;document.getElementById('stopBtn').classList.add('vis')}
async function sendHTTP(t){addUser(t);I.value='';I.style.height='auto';proc=true;SB.disabled=true;showThink();const ac=new AbortController();const tm=setTimeout(()=>ac.abort(),60000);try{const r=await fetch('/api/v1/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({content:t,session_id:hSid,conversation_id:cid,model:selMod}),signal:ac.signal});clearTimeout(tm);hideThink();if(!r.ok){const e=await r.json().catch(()=>({error:'Erro'}));showErr(e.error||e.response||'Erro');finishTurn();return}const d=await r.json();hSid=d.session_id||hSid;if(d.tools&&d.tools.length)d.tools.forEach(x=>{showTool(x.name,x.args);showToolR(x.name,x.status,x.output||'')});if(d.response){appendTxt(d.response);finishTxt()}if(d.file)showFile(d.file);if(d.mission)startPoll(d.mission);finishTurn()}catch(e){clearTimeout(tm);hideThink();showErr(e.name==='AbortError'?'Tempo esgotado, tente novamente':'Erro: '+e.message);finishTurn()}}
function sendCmd(c){I.value=c;sendMessage()}
function qa(t){const w=document.getElementById('welc');if(w)w.remove();I.value=t;I.focus();if(window.innerWidth<769)toggleSB()}
function now(){return new Date().toLocaleTimeString('pt-BR',{hour:'2-digit',minute:'2-digit'})}
let convMsgCount=0;
function addUser(t,save=true){const w=document.getElementById('welc');if(w)w.remove();const wm=document.getElementById('wmark');if(wm)wm.classList.remove('empty');const d=document.createElement('div');d.className='ml user';d.innerHTML=`<div class="mh"><span class="mt">${now()}</span><span class="mn">você</span><div class="mav">${me?me.email[0].toUpperCase():'?'}</div></div><div class="mb-wrap"><div class="mb">${esc(t)}</div></div>`;T.appendChild(d);scrl();convMsgCount++;if(!cid&&save){fetch('/api/v1/conversations',{method:'POST'}).then(r=>r.json()).then(d=>{cid=d.id;autoTitle(t);loadConvs()})}else if(convMsgCount===1&&cid){autoTitle(t)}else if(convMsgCount===2&&cid){const hdr=document.getElementById('hdrT');if(!hdr.textContent||hdr.textContent==='Nova conversa')autoTitle(t)}}
function autoTitle(t){
  const generic=['oi','ola','bom dia','boa tarde','boa noite','hey','hello','hi','e ai','fala','oi!','ola!'];
  const w=t.trim().toLowerCase().replace(/[!?.]/g,'');
  if(generic.includes(w))return;
  const words=t.trim().split(/\s+/).slice(0,6).join(' ');
  const title=words.length>28?words.substring(0,28)+'...':words;
  document.getElementById('hdrT').textContent=title;
  fetch(`/api/v1/conversations/${cid}/title`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({title})}).then(()=>loadConvs());
}
function showThink(){hideThink();const d=document.createElement('div');d.className='think';d.id='thinkEl';const st=Date.now();d.innerHTML='<div class="think-in"><div class="think-logo"></div><span class="think-t">Working\u2026</span></div>';T.appendChild(d);d._thinkTimer=setInterval(()=>{const el=d.querySelector('.think-t');if(el){const s=((Date.now()-st)/1000);if(s<60)el.textContent='Working\u2026 ('+s.toFixed(0)+'s)';else el.textContent='Working\u2026 ('+Math.floor(s/60)+'m '+Math.floor(s%60)+'s)'}},1000);scrl()}
function hideThink(){const e=document.getElementById('thinkEl');if(e){if(e._thinkTimer)clearInterval(e._thinkTimer);e.remove()}}
function ensureMsg(){if(!curMsg){hideThink();curMsg=document.createElement('div');curMsg.className='ml assistant';curBody=document.createElement('div');curBody.className='mb';const wrap=document.createElement('div');wrap.className='mb-wrap';wrap.appendChild(curBody);curMsg.appendChild(wrap);T.appendChild(curMsg);raw=''}}
function appendTxt(t){ensureMsg();raw+=t;const c=curBody.querySelector('.scur');if(c)c.remove();curBody.insertAdjacentText('beforeend',t);const s=document.createElement('span');s.className='scur';curBody.appendChild(s);scrl()}
function finishTxt(){if(curBody){const c=curBody.querySelector('.scur');if(c)c.remove();if(raw&&typeof marked!=='undefined'){marked.setOptions({breaks:true,gfm:true});curBody.innerHTML=marked.parse(raw);curBody.querySelectorAll('a').forEach(a=>{a.target='_blank';a.rel='noopener'})}raw=''}}
function showTool(n,a){ensureMsg();const b=document.createElement('div');b.className='tblk act';const as=typeof a==='string'?a:JSON.stringify(a);const short=as.length>80?as.substring(0,80)+'\u2026':as;b.innerHTML=`<div class="thd" onclick="this.parentElement.classList.toggle('open')"><div class="tdot"></div><span class="tlb"><span class="tn">${esc(n)}</span>(<span class="ta">${esc(short)}</span>)</span><span class="tdr">0.0s</span></div><div class="tout"></div>`;curMsg.querySelector('.mb-wrap').appendChild(b);curTool=b;tStart=Date.now();if(tTimer)clearInterval(tTimer);tTimer=setInterval(()=>{if(!curTool){clearInterval(tTimer);return}const d=curTool.querySelector('.tdr');if(d){const s=(Date.now()-tStart)/1000;if(s<60)d.textContent=s.toFixed(1)+'s';else d.textContent=Math.floor(s/60)+'m '+Math.floor(s%60)+'s'}},100);scrl()}
function showToolR(n,s,o){if(tTimer){clearInterval(tTimer);tTimer=null}if(curTool){curTool.classList.remove('act');curTool.classList.add(s==='success'?'done':'err');if(o){const b=curTool.querySelector('.tout');if(b){const lines=o.split('\n').slice(0,4);const rest=o.split('\n').length-4;let h='';lines.forEach(l=>{h+='<div class="tline">'+esc(l)+'</div>'});if(rest>0)h+='<div class="tmore" onclick="this.parentElement.style.maxHeight=\'none\';this.remove()">\u2026 +'+rest+' lines (click to expand)</div>';b.innerHTML=h;b.style.display='block'}}const d=curTool.querySelector('.tdr');if(d){const s2=(Date.now()-tStart)/1000;if(s2<60)d.textContent=s2.toFixed(1)+'s';else d.textContent=Math.floor(s2/60)+'m '+Math.floor(s2%60)+'s'}curTool=null}scrl()}
function showFile(f){ensureMsg();const ic={'landing_page':'\ud83c\udf10','app':'\u26a1','xlsx':'\ud83d\udcca','docx':'\ud83d\udcc4','pptx':'\ud83c\udfac'};const i=ic[f.type]||'\ud83d\udcc1';const wb=f.type==='landing_page'||f.type==='app';const c=document.createElement('div');c.className='fcard';c.innerHTML=`<div class="ficon">${i}</div><div class="finfo"><div class="fname">${esc(f.name)}</div><div class="fmeta">${esc(f.size)}</div></div><div style="display:flex;gap:6px">${wb?`<a href="${esc(f.url)}" target="_blank" class="fbtn pr">Abrir</a>`:''}<a href="${esc(f.url)}" download class="fbtn ${wb?'sc':'pr'}">Download</a></div>`;curMsg.querySelector('.mb-wrap').appendChild(c);scrl()}
function showErr(t){ensureMsg();const e=document.createElement('div');e.className='eline';e.textContent='\u2717 '+t;curMsg.querySelector('.mb-wrap').appendChild(e);scrl()}
function finishTurn(){finishTxt();proc=false;SB.disabled=false;curMsg=null;curBody=null;I.focus();loadConvs();document.getElementById('stopBtn').classList.remove('vis');toggleInputBtns()}
function stopGeneration(){if(ws&&ws.readyState===1){ws.send(JSON.stringify({type:'stop'}))}finishTurn()}
function scrl(){T.scrollTop=T.scrollHeight}
function esc(t){const d=document.createElement('div');d.textContent=t;return d.innerHTML}
let mPolls={};
function startPoll(m){const mid=m.id;let lt=0;const card=document.createElement('div');card.className='mission-card';card.id='mc-'+mid;card.innerHTML=`<div class="mctitle">\ud83d\ude80 ${esc(m.title)}</div><div class="mission-bar"><div class="mission-fill" id="mf-${mid}" style="width:0%"></div></div><div id="ms-${mid}"></div>`;ensureMsg();curMsg.querySelector('.mb-wrap').appendChild(card);scrl();const poll=async()=>{try{const r=await fetch(`/api/v1/missions/${mid}/progress?after=${lt}`);const d=await r.json();d.events.forEach(e=>{lt=Math.max(lt,e.time);const sl=document.getElementById('ms-'+mid);const fl=document.getElementById('mf-'+mid);if(e.type==='step_start'){const p=((e.data.step+1)/e.data.total*100).toFixed(0);if(fl)fl.style.width=p+'%';sl.innerHTML+=`<div class="mission-step running" id="mss-${mid}-${e.data.step}"><span class="ms-icon">\u23f3</span>${esc(e.data.title)}</div>`;scrl()}else if(e.type==='step_done'){const el=document.getElementById(`mss-${mid}-${e.data.step}`);if(el){el.className='mission-step done';el.querySelector('.ms-icon').textContent='\u2705'}if(e.data.file)showFile(e.data.file);scrl()}else if(e.type==='step_retry'){const el=document.getElementById(`mss-${mid}-${e.data.step}`);if(el)el.querySelector('.ms-icon').textContent='\ud83d\udd04'}else if(e.type==='step_failed'){const el=document.getElementById(`mss-${mid}-${e.data.step}`);if(el){el.className='mission-step failed';el.querySelector('.ms-icon').textContent='\u274c'}}else if(e.type==='completed'){const c=document.getElementById('mc-'+mid);if(c)c.className='mission-done';const tt=c?.querySelector('.mctitle');if(tt)tt.innerHTML='\ud83c\udf89 '+esc(e.data.title)+' — Concluida!';if(fl)fl.style.width='100%';if(e.data.summary){appendTxt(e.data.summary);finishTxt()}clearInterval(mPolls[mid]);scrl()}});if(d.status==='completed'||d.status==='failed')clearInterval(mPolls[mid])}catch(e){}};mPolls[mid]=setInterval(poll,2000);setTimeout(poll,500)}
async function showAdmUsers(){const r=await fetch('/api/v1/admin/users');const d=await r.json();let h='<h3>Usuários</h3><table class="adm-tbl"><tr><th>Email</th><th>Plano</th><th>Status</th><th></th></tr>';d.users.forEach(u=>{const st=u.active?'<span style="color:var(--g)">ativo</span>':'<span style="color:var(--r)">inativo</span>';h+=`<tr><td>${u.email}</td><td><select onchange="setPlan('${u.id}',this.value)">${['free','basic','pro','unlimited'].map(p=>`<option ${u.plan===p?'selected':''}>${p}</option>`).join('')}</select></td><td>${st}</td><td><button onclick="togUsr('${u.id}',${u.active?0:1})">${u.active?'Off':'On'}</button></td></tr>`});h+='</table>';document.getElementById('modalC').innerHTML=h;document.getElementById('modalBg').classList.add('show')}
async function showAdmStats(){const r=await fetch('/api/v1/admin/stats');const d=await r.json();let h=`<h3>Consumo</h3><div class="adm-cards"><div class="adm-card"><div class="al">Usuários</div><div class="av">${d.total_users}</div></div><div class="adm-card"><div class="al">Custo Hoje</div><div class="av">$${d.cost_today.toFixed(3)}</div></div><div class="adm-card"><div class="al">Custo Semana</div><div class="av">$${d.cost_week.toFixed(3)}</div></div><div class="adm-card"><div class="al">Tokens Hoje</div><div class="av">${(d.tokens_today/1000).toFixed(0)}k</div></div></div>`;document.getElementById('modalC').innerHTML=h;document.getElementById('modalBg').classList.add('show')}
function showCreateUsr(){document.getElementById('modalC').innerHTML='<h3>Cadastrar Usuario</h3><label>Email</label><input id="nuE" type="email" placeholder="email@exemplo.com"><label>Senha</label><input id="nuP" type="password" placeholder="minimo 6 chars"><label>Nome</label><input id="nuN" placeholder="opcional"><label>Plano</label><select id="nuPl"><option>free</option><option>basic</option><option>pro</option><option>unlimited</option></select><button class="mbtn" onclick="createUsr()">Cadastrar</button><div id="nuM" style="margin-top:8px;font-size:12px"></div>';document.getElementById('modalBg').classList.add('show')}
async function createUsr(){const e=document.getElementById('nuE').value,p=document.getElementById('nuP').value,n=document.getElementById('nuN').value,pl=document.getElementById('nuPl').value;const r=await fetch('/api/v1/admin/create-user',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({email:e,password:p,name:n,plan:pl})});const d=await r.json();document.getElementById('nuM').innerHTML=d.ok?'<span style="color:var(--g)">Criado!</span>':`<span style="color:var(--r)">${d.error}</span>`}
async function setPlan(id,p){await fetch(`/api/v1/admin/users/${id}`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({plan:p})});showAdmUsers()}
async function togUsr(id,a){await fetch(`/api/v1/admin/users/${id}`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({active:a})});showAdmUsers()}
function clsModal(){document.getElementById('modalBg').classList.remove('show')}
I.addEventListener('keydown',e=>{if(e.key==='Enter'&&!e.shiftKey){e.preventDefault();sendMessage()}});
I.addEventListener('input',()=>{I.style.height='auto';I.style.height=Math.min(I.scrollHeight,120)+'px';toggleInputBtns()});
let lte=0;document.addEventListener('touchend',e=>{const n=Date.now();if(n-lte<=300)e.preventDefault();lte=n},false);

// ── FILE & AUDIO SUPPORT ──────────────────────────────────────
const MIC=document.getElementById('micBtn'),FP=document.getElementById('filePreview');
let pendingFile=null,mediaRec=null,audioChunks=[],recStart=0,recTimer=null,audioBlob=null,audioUrl=null;
let speechRec=null,audioTranscript='';
const hasSpeechRec=!!(window.SpeechRecognition||window.webkitSpeechRecognition);
const fileInput=document.createElement('input');fileInput.type='file';fileInput.style.display='none';document.body.appendChild(fileInput);

function toggleInputBtns(){
  const has=I.value.trim().length>0||pendingFile||audioBlob;
  SB.classList.toggle('vis',has);
  MIC.classList.toggle('hid',has);
  MIC.classList.toggle('vis',!has);
}

// ── Attach Menu ──
function toggleAttachMenu(){
  const m=document.getElementById('attachMenu');
  m.style.display=m.style.display==='none'?'block':'none';
}
document.addEventListener('click',e=>{if(!e.target.closest('.attach-btn')&&!e.target.closest('.attach-menu'))document.getElementById('attachMenu').style.display='none'});

function pickFile(accept,exts){
  document.getElementById('attachMenu').style.display='none';
  fileInput.accept=exts==='*'?'':exts;
  fileInput.value='';
  fileInput.onchange=()=>{if(fileInput.files[0])handleFile(fileInput.files[0])};
  fileInput.click();
}

async function handleFile(f){
  if(f.size>20*1024*1024){showToast('Arquivo muito grande. Maximo: 20MB','error');return}
  const blocked=['.exe','.bat','.sh','.cmd','.com','.msi','.scr','.ps1'];
  const ext='.'+f.name.split('.').pop().toLowerCase();
  if(blocked.includes(ext)){showToast('Tipo de arquivo não permitido','error');return}
  pendingFile={file:f,name:f.name,size:f.size,ext:ext};
  showFilePreview(f);
  toggleInputBtns();
}

function showFilePreview(f){
  const ext='.'+f.name.split('.').pop().toLowerCase();
  const isImg=['.jpg','.jpeg','.png','.gif','.webp'].includes(ext);
  let iconHtml='&#x1F4C1;';
  if(isImg)iconHtml='';
  else if(ext==='.pdf')iconHtml='&#x1F4D5;';
  else if(['.xlsx','.xls','.csv'].includes(ext))iconHtml='&#x1F4CA;';
  else if(['.docx','.doc','.txt','.md'].includes(ext))iconHtml='&#x1F4C4;';
  else if(['.py','.js','.ts','.html','.css','.json','.jsx','.tsx'].includes(ext))iconHtml='&#x1F4BB;';
  const sz=f.size>1024*1024?(f.size/1024/1024).toFixed(1)+' MB':(f.size/1024).toFixed(1)+' KB';
  let thumbHtml;
  if(isImg){
    const url=URL.createObjectURL(f);
    thumbHtml='<div class="fp-thumb"><img src="'+url+'" alt="preview"></div>';
  }else{
    thumbHtml='<div class="fp-thumb">'+iconHtml+'</div>';
  }
  FP.innerHTML='<div class="file-preview">'+thumbHtml+'<div class="fp-info"><div class="fp-name">'+esc(f.name)+'</div><div class="fp-meta">'+sz+'</div></div><button class="fp-close" onclick="clearFile()">&times;</button></div>';
}

function clearFile(){pendingFile=null;FP.innerHTML='';toggleInputBtns()}

// ── Audio Recording ──
function toggleRec(){
  if(mediaRec&&mediaRec.state==='recording'){stopRec();return}
  startRec();
}

async function startRec(){
  if(!hasSpeechRec){showToast('Seu navegador não suporta transcrição de áudio. Digite sua mensagem.','error');return}
  try{
    const stream=await navigator.mediaDevices.getUserMedia({audio:true});
    mediaRec=new MediaRecorder(stream,{mimeType:'audio/webm'});
    audioChunks=[];audioTranscript='';
    mediaRec.ondataavailable=e=>{if(e.data.size>0)audioChunks.push(e.data)};
    mediaRec.onstop=()=>{stream.getTracks().forEach(t=>t.stop());audioBlob=new Blob(audioChunks,{type:'audio/webm'});audioUrl=URL.createObjectURL(audioBlob);showAudioPreview()};
    mediaRec.start();
    // Web Speech API — transcricao em tempo real
    const SR=window.SpeechRecognition||window.webkitSpeechRecognition;
    speechRec=new SR();
    speechRec.lang='pt-BR';
    speechRec.continuous=true;
    speechRec.interimResults=false;
    speechRec.onresult=(ev)=>{audioTranscript=Array.from(ev.results).map(r=>r[0].transcript).join(' ')};
    speechRec.onerror=()=>{};
    speechRec.start();
    MIC.classList.add('recording');
    recStart=Date.now();
    recTimer=setInterval(()=>{
      const s=Math.floor((Date.now()-recStart)/1000);
      const mm=Math.floor(s/60),ss=s%60;
      MIC.title=mm+':'+(ss<10?'0':'')+ss;
    },500);
    setTimeout(()=>{if(mediaRec&&mediaRec.state==='recording')stopRec()},300000);
  }catch(e){showToast('Não foi possível acessar o microfone','error')}
}

function stopRec(){
  if(mediaRec&&mediaRec.state==='recording'){
    mediaRec.stop();
    if(speechRec){try{speechRec.stop()}catch(e){}}
    MIC.classList.remove('recording');
    clearInterval(recTimer);
    MIC.title='Gravar audio';
  }
}

function showAudioPreview(){
  const dur=Math.floor((Date.now()-recStart)/1000);
  const mm=Math.floor(dur/60),ss=dur%60;
  const ts=mm+':'+(ss<10?'0':'')+ss;
  let html='<div class="audio-preview"><button class="ap-play" id="apPlay" onclick="playPreviewAudio()">&#x25B6;</button><div class="ap-bar"><div class="ap-fill" id="apFill"></div></div><span class="ap-dur">'+ts+'</span><button class="ap-del" onclick="clearAudio()">&#x1F5D1;</button></div>';
  if(audioTranscript)html+='<div class="ap-transcript">&#x1F3A4; '+esc(audioTranscript)+'</div>';
  FP.innerHTML=html;
  toggleInputBtns();
}

let previewAudio=null;
function playPreviewAudio(){
  if(!audioUrl)return;
  if(previewAudio){previewAudio.pause();previewAudio=null;document.getElementById('apPlay').innerHTML='&#x25B6;';return}
  previewAudio=new Audio(audioUrl);
  document.getElementById('apPlay').innerHTML='&#x23F8;';
  previewAudio.ontimeupdate=()=>{const p=(previewAudio.currentTime/previewAudio.duration*100)||0;document.getElementById('apFill').style.width=p+'%'};
  previewAudio.onended=()=>{document.getElementById('apPlay').innerHTML='&#x25B6;';document.getElementById('apFill').style.width='0%';previewAudio=null};
  previewAudio.play();
}

function clearAudio(){audioBlob=null;audioUrl=null;audioTranscript='';speechRec=null;FP.innerHTML='';if(previewAudio){previewAudio.pause();previewAudio=null}toggleInputBtns()}

// ── Upload & Send ──
async function uploadFile(f){
  const fd=new FormData();
  fd.append('file',f);
  fd.append('message',I.value.trim());
  try{
    const r=await fetch('/api/v1/upload',{method:'POST',body:fd});
    if(!r.ok){const e=await r.json().catch(()=>({error:'Erro no upload'}));showToast(e.error||'Erro','error');return null}
    return await r.json();
  }catch(e){showToast('Erro no upload: '+e.message,'error');return null}
}

// Override sendMessage to handle files/audio
const _origSendMessage=sendMessage;
sendMessage=async function(){
  if(proc)return;
  // Audio pending?
  if(audioBlob){
    const text=I.value.trim();
    const localAudioUrl=audioUrl;
    const trans=audioTranscript;
    // Abordagem 1: Web Speech API transcreveu — envia como texto normal
    if(trans){
      const msgText=text||trans;
      addUserWithAttachment(msgText,'audio','&#x1F3A4;','Audio',null,localAudioUrl,trans);
      I.value='';I.style.height='auto';proc=true;SB.disabled=true;
      clearAudio();
      showThink();
      if(http){
        const ac=new AbortController();const tm=setTimeout(()=>ac.abort(),60000);
        try{
          const body={content:msgText,session_id:hSid,conversation_id:cid,model:selMod};
          const r=await fetch('/api/v1/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body),signal:ac.signal});
          clearTimeout(tm);hideThink();
          if(!r.ok){const e=await r.json().catch(()=>({error:'Erro'}));showErr(e.error||'Erro');finishTurn();return}
          const d=await r.json();
          hSid=d.session_id||hSid;
          if(!cid&&d.conversation_id)cid=d.conversation_id;
          if(d.response){appendTxt(d.response);finishTxt()}
          finishTurn();
        }catch(e){clearTimeout(tm);hideThink();showErr(e.name==='AbortError'?'Tempo esgotado, tente novamente':'Erro: '+e.message);finishTurn()}
      }else if(ws&&ws.readyState===1){
        hideThink();
        ws.send(JSON.stringify({type:'message',content:msgText,model:selMod,conversation_id:cid}));
        proc=true;SB.disabled=true;
      }else{hideThink();finishTurn()}
      return;
    }
    // Fallback: sem transcricao — envia como "[Audio enviado]" com player visual
    const fallbackText=text||'[Audio enviado]';
    addUserWithAttachment(fallbackText,'audio','&#x1F3A4;','Audio',null,localAudioUrl);
    I.value='';I.style.height='auto';proc=true;SB.disabled=true;
    clearAudio();
    showThink();
    if(http){
      const ac=new AbortController();const tm=setTimeout(()=>ac.abort(),60000);
      try{
        const body={content:fallbackText,session_id:hSid,conversation_id:cid,model:selMod};
        const r=await fetch('/api/v1/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body),signal:ac.signal});
        clearTimeout(tm);hideThink();
        if(!r.ok){const e=await r.json().catch(()=>({error:'Erro'}));showErr(e.error||'Erro');finishTurn();return}
        const d=await r.json();
        hSid=d.session_id||hSid;
        if(d.response){appendTxt(d.response);finishTxt()}
        finishTurn();
      }catch(e){clearTimeout(tm);hideThink();showErr(e.name==='AbortError'?'Tempo esgotado, tente novamente':'Erro: '+e.message);finishTurn()}
    }else if(ws&&ws.readyState===1){
      hideThink();
      ws.send(JSON.stringify({type:'message',content:fallbackText,model:selMod,conversation_id:cid}));
      proc=true;SB.disabled=true;
    }else{hideThink();finishTurn()}
    return;
  }
  // File pending?
  if(pendingFile){
    const text=I.value.trim();
    const f=pendingFile;
    const isImg=['.jpg','.jpeg','.png','.gif','.webp'].includes(f.ext);
    const imgUrl=isImg?URL.createObjectURL(f.file):null;
    const icons={'.pdf':'&#x1F4D5;','.xlsx':'&#x1F4CA;','.xls':'&#x1F4CA;','.csv':'&#x1F4CA;','.docx':'&#x1F4C4;','.doc':'&#x1F4C4;','.txt':'&#x1F4C4;','.md':'&#x1F4C4;'};
    const icon=icons[f.ext]||'&#x1F4C1;';
    addUserWithAttachment(text,isImg?'image':'file',icon,f.name,imgUrl);
    I.value='';I.style.height='auto';proc=true;SB.disabled=true;
    const theFile=f.file;
    clearFile();
    showThink();
    const res=await uploadFile(theFile);
    hideThink();
    if(!res||!res.ok){finishTurn();return}
    const fileData={...res};
    if(http){await sendFileHTTP(text,fileData)}
    else if(ws&&ws.readyState===1){ws.send(JSON.stringify({type:'message',content:text,file_data:fileData,model:selMod,conversation_id:cid}));proc=true;SB.disabled=true}
    else{finishTurn()}
    return;
  }
  // Normal text
  _origSendMessage();
};

function addUserWithAttachment(text,type,icon,name,imgUrl,audUrl,transcript){
  const w=document.getElementById('welc');if(w)w.remove();
  const wm=document.getElementById('wmark');if(wm)wm.classList.remove('empty');
  const d=document.createElement('div');d.className='ml user';
  let attachHtml='';
  if(type==='image'&&imgUrl){
    attachHtml='<img class="chat-img" src="'+imgUrl+'" onclick="openLightbox(this.src)" alt="imagem">';
  }else if(type==='audio'){
    attachHtml='<div class="chat-audio"><button class="ca-play" onclick="playChatAudio(this,\''+esc(audUrl||'')+'\')">&#x25B6;</button><div class="ca-bar"><div class="ca-fill"></div></div><span class="ca-dur">0:00</span></div>';
    if(transcript)attachHtml+='<div class="chat-transcription">&#x1F3A4; '+esc(transcript)+'</div>';
  }else{
    const sz=pendingFile?(pendingFile.size>1024*1024?(pendingFile.size/1024/1024).toFixed(1)+' MB':(pendingFile.size/1024).toFixed(1)+' KB'):'';
    attachHtml='<div class="chat-file-card"><span class="cfc-icon">'+icon+'</span><div class="cfc-info"><div class="cfc-name">'+esc(name)+'</div><div class="cfc-meta">'+sz+'</div></div></div>';
  }
  d.innerHTML='<div class="mh"><span class="mt">'+now()+'</span><span class="mn">você</span><div class="mav">'+(me?me.email[0].toUpperCase():'?')+'</div></div><div class="mb-wrap">'+attachHtml+(text?'<div class="mb">'+esc(text)+'</div>':'')+'</div>';
  T.appendChild(d);scrl();
  convMsgCount++;
  if(!cid){fetch('/api/v1/conversations',{method:'POST'}).then(r=>r.json()).then(d=>{cid=d.id;autoTitle(text||name);loadConvs()})}
}

async function sendFileHTTP(text,fileData){
  showThink();
  const ac=new AbortController();const tm=setTimeout(()=>ac.abort(),60000);
  try{
    const content=text||'[arquivo enviado]';
    const body={content:content,session_id:hSid,conversation_id:cid,model:selMod,file_data:fileData};
    const r=await fetch('/api/v1/chat',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body),signal:ac.signal});
    clearTimeout(tm);hideThink();
    if(!r.ok){const e=await r.json().catch(()=>({error:'Erro'}));showErr(e.error||'Erro');finishTurn();return}
    const d=await r.json();
    hSid=d.session_id||hSid;
    if(d.response){appendTxt(d.response);finishTxt()}
    finishTurn();
  }catch(e){clearTimeout(tm);hideThink();showErr(e.name==='AbortError'?'Tempo esgotado, tente novamente':'Erro: '+e.message);finishTurn()}
}

// ── Drag & Drop ──
const mainEl=document.querySelector('.main');
let dragCounter=0;
mainEl.addEventListener('dragenter',e=>{e.preventDefault();dragCounter++;document.getElementById('dragOv').classList.add('show')});
mainEl.addEventListener('dragleave',e=>{e.preventDefault();dragCounter--;if(dragCounter<=0){dragCounter=0;document.getElementById('dragOv').classList.remove('show')}});
mainEl.addEventListener('dragover',e=>e.preventDefault());
mainEl.addEventListener('drop',e=>{e.preventDefault();dragCounter=0;document.getElementById('dragOv').classList.remove('show');if(e.dataTransfer.files.length)handleFile(e.dataTransfer.files[0])});

// ── Ctrl+V Paste ──
document.addEventListener('paste',e=>{
  const items=e.clipboardData&&e.clipboardData.items;
  if(!items)return;
  for(let i=0;i<items.length;i++){
    if(items[i].type.indexOf('image')!==-1){
      const f=items[i].getAsFile();
      if(f){e.preventDefault();handleFile(f)}
      break;
    }
  }
});

// ── Lightbox ──
function openLightbox(src){const lb=document.getElementById('lightbox');document.getElementById('lbImg').src=src;lb.classList.add('show')}
function closeLightbox(){document.getElementById('lightbox').classList.remove('show')}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeLightbox()});

// ── Chat audio player ──
function playChatAudio(btn,url){
  const wrap=btn.closest('.chat-audio');
  const fill=wrap.querySelector('.ca-fill');
  const durEl=wrap.querySelector('.ca-dur');
  if(btn._audio){btn._audio.pause();btn._audio=null;btn.innerHTML='&#x25B6;';fill.style.width='0%';return}
  const a=new Audio(url);btn._audio=a;btn.innerHTML='&#x23F8;';
  a.ontimeupdate=()=>{const p=(a.currentTime/a.duration*100)||0;fill.style.width=p+'%';const s=Math.floor(a.currentTime);durEl.textContent=Math.floor(s/60)+':'+(s%60<10?'0':'')+(s%60)};
  a.onended=()=>{btn.innerHTML='&#x25B6;';fill.style.width='0%';btn._audio=null;durEl.textContent='0:00'};
  a.play();
}

// ── Toast ──
function showToast(msg,type){
  const t=document.createElement('div');t.className='toast'+(type==='error'?' error':'');t.textContent=msg;
  document.body.appendChild(t);setTimeout(()=>t.remove(),3500);
}

toggleInputBtns();
init();
// ── Particles ──
(function(){
  const isMobile=window.innerWidth<769;
  const count=isMobile?8:20;
  const colors=['#9B59FC','#4A9EFF'];
  for(let i=0;i<count;i++){
    const p=document.createElement('div');
    p.className='particle';
    const s=Math.random()*2+2;
    p.style.width=s+'px';p.style.height=s+'px';
    p.style.top=Math.random()*100+'%';
    p.style.left=Math.random()*100+'%';
    p.style.backgroundColor=colors[i%2];
    p.style.animation='floatParticle '+(15+Math.random()*25)+'s ease-in-out '+(Math.random()*10)+'s infinite';
    document.body.appendChild(p);
  }
})();
</script>
</body>
</html>
'''

# ── Feature #24: Dashboard HTML ────────────────────────────────

DASHBOARD_HTML = r'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
<meta name="theme-color" content="#050510">
<link rel="icon" type="image/png" href="/static/brand/favicon.png">
<title>Clow — Dashboard</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@300;400;500;600;700&display=swap');

  :root {
    --bg-deep: #050510;
    --bg-primary: #0A0A1A;
    --bg-secondary: #0F0F24;
    --bg-tertiary: #14142E;
    --bg-surface: #1A1A3A;
    --bg-elevated: #14142E;
    --text-primary: #E8E8F0;
    --text-secondary: #9898B8;
    --text-muted: #585878;
    --purple: #9B59FC;
    --purple-bright: #B77CFF;
    --purple-dim: #8B49EC;
    --purple-deep: #6B39CC;
    --purple-glow: rgba(155,89,252,0.15);
    --blue: #4A9EFF;
    --blue-dim: rgba(74,158,255,0.15);
    --gradient-primary: linear-gradient(135deg, #9B59FC 0%, #4A9EFF 100%);
    --green: #4ADE80;
    --green-dim: rgba(74,222,128,0.15);
    --red: #F87171;
    --red-dim: rgba(248,113,113,0.15);
    --amber: #FBBF24;
    --amber-dim: rgba(251,191,36,0.15);
    --border: rgba(100,100,180,0.12);
    --border-focus: rgba(155,89,252,0.4);
    --font-sans: 'DM Sans', -apple-system, sans-serif;
    --font-mono: "JetBrains Mono", "Fira Code", monospace;
    --safe-top: env(safe-area-inset-top, 0px);
  }

  * { margin: 0; padding: 0; box-sizing: border-box; -webkit-tap-highlight-color: transparent; }
  html { height: 100%; }

  body {
    background: var(--bg-deep);
    color: var(--text-primary);
    font-family: var(--font-sans);
    font-size: 13px;
    line-height: 1.65;
    min-height: 100%;
    -webkit-font-smoothing: antialiased;
    position: relative;
  }
  body::before {
    content: '';
    position: fixed;
    top: 0; left: 0; width: 100%; height: 100%;
    z-index: 0; pointer-events: none;
    background: repeating-linear-gradient(-45deg, transparent, transparent 10px, rgba(100,100,180,.03) 10px, rgba(100,100,180,.03) 11px);
  }

  /* ── Title Bar ── */
  .title-bar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0 20px;
    padding-top: var(--safe-top);
    background: var(--bg-primary);
    border-bottom: 1px solid var(--border);
    height: calc(52px + var(--safe-top));
    position: sticky;
    top: 0;
    z-index: 10;
    backdrop-filter: blur(12px);
  }
  .title-left { display: flex; align-items: center; gap: 10px; }
  .logo-infinity { width: 28px; height: 28px; flex-shrink: 0; }
  .logo-infinity path {
    fill: none; stroke: url(#dashGrad); stroke-width: 3;
    stroke-linecap: round; stroke-linejoin: round;
  }
  .logo-text {
    font-size: 20px; font-weight: 700; letter-spacing: 2px;
    background: var(--gradient-primary);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text; text-transform: uppercase;
  }
  .title-right { display: flex; align-items: center; gap: 10px; }
  .nav-pill {
    background: var(--bg-tertiary); border: 1px solid var(--border);
    color: var(--text-secondary); font-family: var(--font-sans);
    font-size: 10px; font-weight: 500; padding: 5px 12px;
    border-radius: 6px; cursor: pointer; text-decoration: none;
    transition: all 0.2s; letter-spacing: 0.3px;
  }
  .nav-pill:hover, .nav-pill:active { background: var(--bg-surface); color: var(--purple); border-color: var(--border-focus); }
  .nav-pill.active { background: var(--purple-glow); color: var(--purple); border-color: var(--border-focus); }

  /* ── Dashboard Body ── */
  .dash {
    max-width: 1100px;
    margin: 0 auto;
    padding: 24px 20px 40px;
  }

  .dash-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 24px;
  }
  .dash-title {
    font-size: 16px;
    font-weight: 600;
    color: var(--text-primary);
    letter-spacing: 0.5px;
  }
  .refresh-tag {
    font-size: 10px;
    color: var(--text-muted);
    background: var(--bg-tertiary);
    padding: 3px 10px;
    border-radius: 4px;
    border: 1px solid var(--border);
  }

  /* ── Cards ── */
  .cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
    gap: 14px;
    margin-bottom: 32px;
  }
  .card {
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 20px;
    transition: all 0.3s;
  }
  .card:hover { border-color: rgba(155,89,252,.2); box-shadow: 0 0 20px rgba(155,89,252,.05); }
  ::-webkit-scrollbar { width: 5px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: linear-gradient(180deg, var(--purple), var(--blue)); border-radius: 3px; }
  .card-label {
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
    color: var(--text-muted);
    margin-bottom: 10px;
  }
  .card-value {
    font-size: 36px;
    font-weight: 700;
    background: var(--gradient-primary);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    line-height: 1;
  }
  .card-value .unit {
    font-size: 14px;
    -webkit-text-fill-color: var(--text-muted);
  }
  .card-detail {
    margin-top: 8px;
    display: flex;
    flex-wrap: wrap;
    gap: 4px;
  }
  .badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.3px;
  }
  .badge.ok { background: var(--green-dim); color: var(--green); }
  .badge.warn { background: var(--red-dim); color: var(--red); }
  .badge.info { background: var(--blue-dim); color: var(--blue); }
  .badge.purple { background: var(--purple-glow); color: var(--purple); }

  /* ── Section Headers ── */
  .section-title {
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
    background: var(--gradient-primary);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    margin-bottom: 12px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }

  /* ── Tables ── */
  .table-wrap {
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: 12px;
    overflow: hidden;
    margin-bottom: 28px;
  }
  table { width: 100%; border-collapse: collapse; }
  th {
    background: var(--bg-tertiary);
    color: var(--text-muted);
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    text-align: left;
    padding: 10px 14px;
  }
  td {
    padding: 8px 14px;
    border-top: 1px solid var(--border);
    font-size: 12px;
    color: var(--text-primary);
  }
  tr:hover td { background: var(--bg-tertiary); }
  .s-ok { color: var(--green); }
  .s-err { color: var(--red); }
  .s-run { color: var(--blue); }
  .s-muted { color: var(--text-muted); }

  /* ── Responsive ── */
  @media (max-width: 600px) {
    .cards { grid-template-columns: repeat(2, 1fr); gap: 10px; }
    .card { padding: 14px; }
    .card-value { font-size: 28px; }
    .dash { padding: 16px 14px 32px; }
    .table-wrap { overflow-x: auto; }
    th, td { padding: 6px 10px; white-space: nowrap; }
  }
</style>
</head>
<body>

<!-- Title Bar -->
<svg style="position:absolute;width:0;height:0"><defs><linearGradient id="dashGrad" x1="0%" y1="0%" x2="100%" y2="0%"><stop offset="0%" stop-color="#9B59FC"/><stop offset="100%" stop-color="#4A9EFF"/></linearGradient></defs></svg>
<div class="title-bar">
  <div class="title-left">
    <svg class="logo-infinity" viewBox="0 0 32 32">
      <path d="M8 16c0-3 2-6 5-6s5 3 8 6c3 3 5 6 8 6s5-3 5-6-2-6-5-6-5 3-8 6c-3 3-5 6-8 6s-5-3-5-6z" transform="translate(-5,0) scale(0.95)"/>
    </svg>
    <span class="logo-text">Clow</span>
  </div>
  <div class="title-right">
    <a href="/" class="nav-pill">terminal</a>
    <a href="/dashboard" class="nav-pill active">dashboard</a>
    <a href="/logout" class="nav-pill" style="color:var(--red);border-color:rgba(248,113,113,0.2)">sair</a>
  </div>
</div>

<div class="dash">
  <div class="dash-header">
    <span class="dash-title">System Overview</span>
    <span class="refresh-tag">auto-refresh 10s &middot; <span id="lastUpdate">--:--</span></span>
  </div>

  <div class="cards" id="cards">
    <div class="card"><div class="card-label">carregando</div><div class="card-value">...</div></div>
  </div>

  <div class="section-title">Cron Jobs</div>
  <div class="table-wrap">
    <table>
      <thead><tr><th>ID</th><th>Prompt</th><th>Intervalo</th><th>Status</th><th>Runs</th><th>Proxima</th></tr></thead>
      <tbody id="cronBody"><tr><td colspan="6" class="s-muted">Carregando...</td></tr></tbody>
    </table>
  </div>

  <div class="section-title">Acoes Recentes</div>
  <div class="table-wrap">
    <table>
      <thead><tr><th>Hora</th><th>Acao</th><th>Detalhes</th><th>Status</th></tr></thead>
      <tbody id="actionsBody"><tr><td colspan="4" class="s-muted">Carregando...</td></tr></tbody>
    </table>
  </div>
</div>

<script>
async function refresh() {
  try {
    const r = await fetch('/health');
    const d = await r.json();
    document.getElementById('lastUpdate').textContent = new Date().toLocaleTimeString('pt-BR', {hour:'2-digit',minute:'2-digit',second:'2-digit'});

    const c = d.components;
    document.getElementById('cards').innerHTML = `
      <div class="card">
        <div class="card-label">Tasks</div>
        <div class="card-value">${c.tasks.total}</div>
        <div class="card-detail">${Object.entries(c.tasks.by_status).map(([k,v])=>`<span class="badge ${k==='completed'?'ok':k==='failed'?'warn':'info'}">${k}: ${v}</span>`).join('')}</div>
      </div>
      <div class="card">
        <div class="card-label">Cron Jobs</div>
        <div class="card-value">${c.cron.active_jobs}<span class="unit"> / ${c.cron.total_jobs}</span></div>
        <div class="card-detail"><span class="badge purple">ativos</span></div>
      </div>
      <div class="card">
        <div class="card-label">Memoria</div>
        <div class="card-value">${c.memory.total}</div>
        <div class="card-detail">${Object.entries(c.memory.by_type).map(([k,v])=>`<span class="badge purple">${k}: ${v}</span>`).join('')}</div>
      </div>
      <div class="card">
        <div class="card-label">Triggers</div>
        <div class="card-value">${c.triggers.results_count}</div>
        <div class="card-detail">${c.triggers.running?'<span class="badge ok">online :'+c.triggers.port+'</span>':'<span class="badge warn">offline</span>'}</div>
      </div>`;

    const cronBody = document.getElementById('cronBody');
    if (c.cron.jobs.length === 0) {
      cronBody.innerHTML = '<tr><td colspan="6" class="s-muted">Nenhum cron job</td></tr>';
    } else {
      cronBody.innerHTML = c.cron.jobs.map(j => {
        const next = new Date(j.next_run * 1000).toLocaleTimeString('pt-BR',{hour:'2-digit',minute:'2-digit'});
        return `<tr>
          <td style="color:var(--purple)">${j.id}</td>
          <td>${j.prompt}</td>
          <td><span class="badge purple">${j.interval}</span></td>
          <td class="${j.active?'s-ok':'s-err'}">${j.active?'ativo':'pausado'}</td>
          <td>${j.run_count}x</td>
          <td class="s-muted">${next}</td>
        </tr>`;
      }).join('');
    }

    const actionsBody = document.getElementById('actionsBody');
    const actions = d.recent_actions || [];
    if (actions.length === 0) {
      actionsBody.innerHTML = '<tr><td colspan="4" class="s-muted">Nenhuma ação recente</td></tr>';
    } else {
      actionsBody.innerHTML = actions.reverse().map(a => {
        const t = new Date(a.timestamp * 1000).toLocaleTimeString('pt-BR',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
        return `<tr>
          <td class="s-muted">${t}</td>
          <td>${a.action}</td>
          <td style="color:var(--text-secondary)">${a.details}</td>
          <td class="${a.status==='ok'?'s-ok':'s-err'}">${a.status}</td>
        </tr>`;
      }).join('');
    }
  } catch(e) {
    document.getElementById('cards').innerHTML = `<div class="card"><div class="card-label" style="color:var(--red)">Erro</div><div class="card-detail">${e.message}</div></div>`;
  }
}
refresh();
setInterval(refresh, 10000);
</script>
</body>
</html>
'''



LOGIN_HTML = r'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,minimum-scale=1.0,maximum-scale=3.0,user-scalable=yes,viewport-fit=cover">
<meta name="theme-color" content="#050510">
<meta name="description" content="Clow - AI Code Agent | Inteligência Infinita • Possibilidades Premium. Crie landing pages, planilhas, documentos, apps e muito mais.">
<meta property="og:title" content="Clow — Inteligência Infinita">
<meta property="og:description" content="AI Code Agent premium. Crie landing pages, planilhas, apps e muito mais.">
<meta property="og:image" content="/static/brand/logo.png">
<meta property="og:type" content="website">
<link rel="icon" type="image/png" href="/static/brand/favicon.png">
<link rel="apple-touch-icon" href="/static/brand/icon.png">
<title>Clow — Inteligência Infinita</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root{--bg:#050510;--bg2:#0F0F24;--bg3:#14142E;--bd:rgba(100,100,180,.12);--bdf:rgba(124,92,252,.5);--p:#9B59FC;--bl:#4A9EFF;--pg:rgba(155,89,252,.15);--ph:#8B49EC;--gp:linear-gradient(135deg,#9B59FC 0%,#4A9EFF 100%);--r:#F87171;--rd:rgba(248,113,113,.1);--t1:#E8E8F0;--t2:#9898B8;--tm:#585878;--sans:'DM Sans',-apple-system,BlinkMacSystemFont,sans-serif;--mono:'JetBrains Mono',monospace}
*{margin:0;padding:0;box-sizing:border-box}
html,body{height:100%;font-family:var(--sans);background:var(--bg);display:flex;justify-content:center;align-items:center;-webkit-font-smoothing:antialiased;-moz-osx-font-smoothing:grayscale;text-rendering:optimizeLegibility;position:relative}
body::before{content:'';position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none;
  background:repeating-linear-gradient(-45deg,transparent,transparent 10px,rgba(100,100,180,.03) 10px,rgba(100,100,180,.03) 11px),
  url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.02'/%3E%3C/svg%3E")}
body::after{content:'';position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none;
  background:radial-gradient(ellipse at 50% 40%,rgba(155,89,252,.08) 0%,rgba(74,158,255,.04) 30%,transparent 60%)}
.particle{position:fixed;border-radius:50%;pointer-events:none;z-index:1;filter:blur(1px)}
@keyframes floatParticle{0%,100%{transform:translateY(0) translateX(0);opacity:.3}25%{transform:translateY(-30px) translateX(10px);opacity:.6}50%{transform:translateY(-15px) translateX(-10px);opacity:.4}75%{transform:translateY(-40px) translateX(5px);opacity:.5}}
.card{width:100%;max-width:400px;padding:48px 40px;background:var(--bg2);border:1px solid var(--bd);border-radius:20px;box-shadow:0 30px 60px rgba(0,0,0,.5),0 0 60px rgba(155,89,252,.05);animation:fadeIn .4s ease-out;position:relative;z-index:2}
@keyframes fadeIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
.logo{text-align:center;margin-bottom:32px}
.logo img{height:100px;display:block;margin:0 auto 8px;filter:drop-shadow(0 0 24px rgba(155,89,252,.25)) drop-shadow(0 0 24px rgba(74,158,255,.15));animation:float 4s ease-in-out infinite}
@keyframes float{0%,100%{transform:translateY(0)}50%{transform:translateY(-5px)}}
.lens-flare{width:200px;height:1px;margin:12px auto 16px;background:linear-gradient(90deg,transparent,rgba(155,89,252,.5),rgba(74,158,255,.5),transparent);box-shadow:0 0 10px rgba(155,89,252,.3);border-radius:1px}
.tagline{font-family:var(--sans);font-size:10px;color:var(--tm);letter-spacing:3px;text-transform:uppercase}
.tagline .dot{color:var(--p)}
.fg{margin-bottom:16px}
.fg label{display:block;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:var(--tm);margin-bottom:6px}
.fg input{width:100%;padding:12px 14px;background:var(--bg3);border:1px solid var(--bd);border-radius:10px;color:var(--t1);font-family:var(--sans);font-size:14px;outline:none;transition:all .2s}
.fg input:focus{border-color:var(--bdf);box-shadow:0 0 20px rgba(155,89,252,.08)}
.fg input::placeholder{color:var(--tm);font-weight:300}
.lbtn{width:100%;padding:14px;margin-top:8px;height:48px;background:var(--gp);border:none;border-radius:10px;color:#fff;font-family:var(--sans);font-size:15px;font-weight:600;cursor:pointer;transition:all .2s}.lbtn:hover{box-shadow:0 0 30px var(--pg);filter:brightness(1.1)}.lbtn:active{transform:scale(.98)}
.err{margin-top:16px;padding:10px 14px;border-radius:8px;background:var(--rd);border:1px solid rgba(248,113,113,.15);color:var(--r);font-size:12px;text-align:center;display:none}.err.show{display:block}
.sr-only{position:absolute;width:1px;height:1px;padding:0;margin:-1px;overflow:hidden;clip:rect(0,0,0,0);white-space:nowrap;border:0}
::-webkit-scrollbar{width:5px}::-webkit-scrollbar-track{background:transparent}::-webkit-scrollbar-thumb{background:linear-gradient(180deg,var(--p),var(--bl));border-radius:3px}
</style>
</head>
<body>
<div class="card" role="main">
  <div class="logo">
    <img src="/static/brand/logo.png" alt="Clow - Inteligência Infinita">
    <div class="lens-flare"></div>
    <p class="tagline">INTELIG&Ecirc;NCIA INFINITA <span class="dot">&bull;</span> POSSIBILIDADES PREMIUM</p>
  </div>
  <form method="POST" action="/login">
    <div class="fg"><label for="login-email">Email</label><input id="login-email" type="email" name="email" placeholder="seu@email.com" required autofocus autocomplete="email"></div>
    <div class="fg"><label for="login-pass">Senha</label><input id="login-pass" type="password" name="password" placeholder="sua senha" required autocomplete="current-password"></div>
    <button class="lbtn" type="submit" aria-label="Entrar no sistema">Entrar</button>
  </form>
  <div class="err __ERROR_CLASS__">__ERROR_MSG__</div>
</div>
<script>
(function(){
  const isMobile=window.innerWidth<769;
  const count=isMobile?8:20;
  const colors=['#9B59FC','#4A9EFF'];
  for(let i=0;i<count;i++){
    const p=document.createElement('div');
    p.className='particle';
    const s=Math.random()*2+2;
    p.style.width=s+'px';p.style.height=s+'px';
    p.style.top=Math.random()*100+'%';
    p.style.left=Math.random()*100+'%';
    p.style.backgroundColor=colors[i%2];
    p.style.animation='floatParticle '+(15+Math.random()*25)+'s ease-in-out '+(Math.random()*10)+'s infinite';
    document.body.appendChild(p);
  }
})();
</script>
</body>
</html>
'''


ADMIN_HTML = r'''<!DOCTYPE html>
<html lang="pt-BR"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<meta name="theme-color" content="#050510"><link rel="icon" type="image/png" href="/static/brand/favicon.png"><title>Clow Admin</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;600;700&display=swap');
:root{--bg:#050510;--s1:#0A0A1A;--s2:#0F0F24;--s3:#14142E;--t1:#E8E8F0;--t2:#9898B8;--tm:#585878;--p:#9B59FC;--bl:#4A9EFF;--g:#4ADE80;--r:#F87171;--b:rgba(100,100,180,.12);--bf:rgba(155,89,252,.4);--gp:linear-gradient(135deg,#9B59FC 0%,#4A9EFF 100%);--f:'DM Sans',-apple-system,sans-serif;--fm:'JetBrains Mono',monospace}
*{margin:0;padding:0;box-sizing:border-box}body{background:var(--bg);color:var(--t1);font-family:var(--f);font-size:13px;-webkit-font-smoothing:antialiased;position:relative}
body::before{content:'';position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none;background:repeating-linear-gradient(-45deg,transparent,transparent 10px,rgba(100,100,180,.03) 10px,rgba(100,100,180,.03) 11px)}
.bar{display:flex;align-items:center;justify-content:space-between;padding:0 20px;background:var(--s1);border-bottom:1px solid var(--b);height:52px;position:sticky;top:0;z-index:10}
.logo{font-size:20px;font-weight:700;letter-spacing:2px;background:var(--gp);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.nav a{color:var(--t2);font-size:10px;text-decoration:none;padding:5px 10px;border:1px solid var(--b);border-radius:6px;margin-left:8px;transition:all .2s}
.nav a:hover{color:var(--p);border-color:var(--bf)}
.wrap{max-width:900px;margin:0 auto;padding:24px 16px;position:relative;z-index:1}
.row{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:28px}
.stat{background:var(--s2);border:1px solid var(--b);border-radius:12px;padding:16px;transition:all .3s}
.stat:hover{border-color:rgba(155,89,252,.2);box-shadow:0 0 20px rgba(155,89,252,.05)}
.stat-label{font-size:10px;font-weight:600;letter-spacing:.8px;text-transform:uppercase;color:var(--tm);margin-bottom:8px}
.stat-val{font-size:28px;font-weight:700;background:var(--gp);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.stat-sub{font-size:10px;color:var(--tm);margin-top:4px}
.sec{font-size:12px;font-weight:600;letter-spacing:1px;text-transform:uppercase;background:var(--gp);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid var(--b)}
.tbl{width:100%;background:var(--s2);border:1px solid var(--b);border-radius:12px;overflow:hidden;margin-bottom:24px}
table{width:100%;border-collapse:collapse}
th{background:var(--s3);color:var(--tm);font-size:10px;font-weight:600;letter-spacing:.5px;text-transform:uppercase;text-align:left;padding:10px 12px}
td{padding:8px 12px;border-top:1px solid var(--b);font-size:12px}
.badge{padding:2px 8px;border-radius:4px;font-size:10px;font-weight:500}
.badge.g{background:rgba(74,222,128,.15);color:var(--g)}.badge.r{background:rgba(248,113,113,.15);color:var(--r)}
.badge.p{background:rgba(155,89,252,.15);color:var(--p)}
select,button{background:var(--s3);border:1px solid var(--b);color:var(--t1);font-family:var(--fm);font-size:11px;padding:4px 8px;border-radius:4px;cursor:pointer;transition:all .15s}
button:hover{border-color:var(--bf);color:var(--p)}
::-webkit-scrollbar{width:5px}::-webkit-scrollbar-track{background:transparent}::-webkit-scrollbar-thumb{background:linear-gradient(180deg,var(--p),var(--bl));border-radius:3px}
</style></head><body>
<div class="bar"><span class="logo">CLOW ADMIN</span><div class="nav"><a href="/">Terminal</a><a href="/dashboard">Dashboard</a><a href="/logout">Sair</a></div></div>
<div class="wrap">
<div class="row" id="stats"></div>
<div class="sec">Usuários</div>
<div class="tbl"><table><thead><tr><th>Email</th><th>Plano</th><th>Status</th><th>Criado</th><th>Acao</th></tr></thead><tbody id="usersBody"></tbody></table></div>
<div class="sec">Top Consumo (Hoje)</div>
<div class="tbl"><table><thead><tr><th>Email</th><th>Plano</th><th>Tokens</th><th>Custo</th></tr></thead><tbody id="topBody"></tbody></table></div>
</div>
<script>
async function load(){
  const [sr,ur]=await Promise.all([fetch('/api/v1/admin/stats').then(r=>r.json()),fetch('/api/v1/admin/users').then(r=>r.json())]);
  document.getElementById('stats').innerHTML=`
    <div class="stat"><div class="stat-label">Usuários</div><div class="stat-val">${sr.total_users}</div><div class="stat-sub">${sr.active_users} ativos</div></div>
    <div class="stat"><div class="stat-label">Custo Hoje</div><div class="stat-val">$${sr.cost_today.toFixed(3)}</div></div>
    <div class="stat"><div class="stat-label">Custo Semana</div><div class="stat-val">$${sr.cost_week.toFixed(3)}</div></div>
    <div class="stat"><div class="stat-label">Custo Mes</div><div class="stat-val">$${sr.cost_month.toFixed(3)}</div></div>
    <div class="stat"><div class="stat-label">Tokens Hoje</div><div class="stat-val">${(sr.tokens_today/1000).toFixed(0)}k</div></div>`;
  const ub=document.getElementById('usersBody');
  ub.innerHTML=ur.users.map(u=>{
    const dt=new Date(u.created_at*1000).toLocaleDateString('pt-BR');
    const st=u.active?'<span class="badge g">ativo</span>':'<span class="badge r">inativo</span>';
    return `<tr><td>${u.email}</td><td><select onchange="setPlan('${u.id}',this.value)">${['free','basic','pro','unlimited'].map(p=>`<option ${u.plan===p?'selected':''}>${p}</option>`).join('')}</select></td><td>${st}</td><td style="color:var(--tm)">${dt}</td><td><button onclick="toggle('${u.id}',${u.active?0:1})">${u.active?'Desativar':'Ativar'}</button></td></tr>`;
  }).join('');
  const tb=document.getElementById('topBody');
  tb.innerHTML=(sr.top_users_today||[]).map(u=>`<tr><td>${u.email}</td><td><span class="badge p">${u.plan}</span></td><td>${(u.tokens/1000).toFixed(0)}k</td><td>$${u.cost.toFixed(4)}</td></tr>`).join('')||'<tr><td colspan="4" style="color:var(--tm)">Sem dados</td></tr>';
}
async function setPlan(id,plan){await fetch(`/api/v1/admin/users/${id}`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({plan})});load();}
async function toggle(id,active){await fetch(`/api/v1/admin/users/${id}`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({active})});load();}
load();setInterval(load,30000);
</script></body></html>
'''


if HAS_FASTAPI:
    # ── Login routes (sem auth) ──────────────────────────────────
    @app.get("/login", response_class=HTMLResponse)
    async def login_page(request: Request):
        if _get_session_from_request(request):
            return RedirectResponse("/", status_code=302)
        html = LOGIN_HTML.replace("__ERROR_CLASS__", "").replace("__ERROR_MSG__", "")
        return HTMLResponse(html)

    @app.post("/login")
    async def login_submit(request: Request):
        form = await request.form()
        email = form.get("email", "").strip().lower()
        password = form.get("password", "")

        user = authenticate_user(email, password)
        if user:
            token = _create_session(user)
            resp = RedirectResponse("/", status_code=302)
            resp.set_cookie(
                "clow_session", token,
                max_age=_SESSION_TTL, httponly=True,
                samesite="lax", secure=False,
            )
            return resp

        html = LOGIN_HTML.replace("__ERROR_CLASS__", "show").replace("__ERROR_MSG__", "Email ou senha incorretos")
        return HTMLResponse(html, status_code=401)

    @app.get("/logout")
    async def logout(request: Request):
        token = request.cookies.get("clow_session", "")
        if token in _web_sessions:
            del _web_sessions[token]
        resp = RedirectResponse("/login", status_code=302)
        resp.delete_cookie("clow_session")
        return resp

    # ── Protected routes ─────────────────────────────────────────
    @app.get("/", response_class=HTMLResponse)
    async def index(request: Request):
        if not _get_session_from_request(request):
            return RedirectResponse("/login", status_code=302)
        return WEBAPP_HTML

    # ── PWA Routes (System Clow App) ──────────────────────────────
    @app.get("/pwa", response_class=HTMLResponse)
    async def pwa_index():
        """Página principal do PWA."""
        static_dir = Path(__file__).parent.parent / "static"
        if (static_dir / "index.html").exists():
            with open(static_dir / "index.html") as f:
                return f.read()
        return HTMLResponse("<h1>System Clow</h1><p>PWA app</p>")

    @app.get("/static/manifest.json")
    async def manifest():
        """Manifest do PWA."""
        static_dir = Path(__file__).parent.parent / "static"
        manifest_path = static_dir / "manifest.json"
        if manifest_path.exists():
            return FileResponse(manifest_path, media_type="application/manifest+json")
        return JSONResponse({"name": "System Clow", "short_name": "Clow"})

    @app.get("/static/service-worker.js")
    async def service_worker():
        """Service Worker para PWA."""
        static_dir = Path(__file__).parent.parent / "static"
        sw_path = static_dir / "service-worker.js"
        if sw_path.exists():
            return FileResponse(sw_path, media_type="application/javascript")
        return JSONResponse({"error": "Service Worker not found"}, status_code=404)

    @app.get("/static/{file_path:path}")
    async def static_files(file_path: str):
        """Serve arquivos estáticos (CSS, JS, imagens, etc)."""
        static_dir = Path(__file__).parent.parent / "static"
        full_path = (static_dir / file_path).resolve()
        
        # Security: previne path traversal
        if not str(full_path).startswith(str(static_dir)):
            return JSONResponse({"error": "Forbidden"}, status_code=403)
        
        if full_path.exists() and full_path.is_file():
            return FileResponse(full_path)
        return JSONResponse({"error": "Not found"}, status_code=404)

    # Feature #19: Health Check (publico — sem auth)
    @app.get("/health", dependencies=[Depends(_rate_limit_dependency)])
    async def health():
        return JSONResponse(_get_health_data())

    # API endpoints protegidos
    @app.get("/api/v1/status", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_status():
        """Status completo da API com metricas."""
        return JSONResponse({
            "version": __version__,
            "status": "ok",
            **_get_health_data(),
        })

    @app.post("/api/v1/generate-key", dependencies=[Depends(_auth_dependency)])
    async def generate_key():
        """Gera nova API key para autenticacao."""
        key = _generate_api_key()
        return JSONResponse({"api_key": key, "note": "Adicione esta key em settings.json > webapp > api_keys"})

    @app.get("/api/v1/metrics", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_metrics():
        """Retorna metricas coletadas (counters, gauges, histograms)."""
        from .logging import metrics
        return JSONResponse(metrics.snapshot())

    @app.get("/api/v1/sessions", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_sessions():
        """Lista sessoes salvas."""
        from .session import list_sessions
        return JSONResponse({"sessions": list_sessions()})

    @app.get("/api/v1/memory", dependencies=[Depends(_auth_dependency), Depends(_rate_limit_dependency)])
    async def api_memory():
        """Lista memorias persistidas."""
        from .memory import list_memories
        return JSONResponse({"memories": list_memories()})

    @app.get("/api/v1/tools", dependencies=[Depends(_auth_dependency)])
    async def api_tools():
        """Lista todas as ferramentas disponiveis."""
        from .tools.base import create_default_registry
        registry = create_default_registry()
        tools = [{"name": t.name, "description": t.description} for t in registry.all_tools()]
        return JSONResponse({"tools": tools, "count": len(tools)})

    # ── API: Conversations ──────────────────────────────────────────

    @app.get("/api/v1/conversations")
    async def api_conversations(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        convs = list_conversations(sess["user_id"])
        return JSONResponse({"conversations": convs})

    @app.post("/api/v1/conversations")
    async def api_create_conversation(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        cid = create_conversation(sess["user_id"])
        return JSONResponse({"id": cid})

    @app.get("/api/v1/conversations/{conv_id}/messages")
    async def api_get_messages(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        msgs = get_messages(conv_id)
        return JSONResponse({"messages": msgs})

    @app.delete("/api/v1/conversations/{conv_id}")
    async def api_delete_conversation(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        delete_conversation(sess["user_id"], conv_id)
        return JSONResponse({"ok": True})

    @app.post("/api/v1/conversations/{conv_id}/title")
    async def api_update_conv_title(conv_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        body = await request.json()
        title = body.get("title", "")[:50]
        if title:
            update_conversation_title(conv_id, title)
        return JSONResponse({"ok": True})

    # ── API: Upload de Arquivos ─────────────────────────────────────

    _ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    _ALLOWED_DOC_EXT = {".pdf", ".docx", ".doc", ".txt", ".md"}
    _ALLOWED_SHEET_EXT = {".xlsx", ".xls", ".csv"}
    _ALLOWED_CODE_EXT = {".py", ".js", ".html", ".css", ".json", ".ts", ".jsx", ".tsx"}
    _ALLOWED_AUDIO_EXT = {".webm", ".mp3", ".ogg", ".wav", ".m4a"}
    _BLOCKED_EXT = {".exe", ".bat", ".sh", ".cmd", ".com", ".msi", ".scr", ".ps1"}
    _MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
    _UPLOAD_LIMITS = {"free": 5, "basic": 20, "pro": 100, "unlimited": 0}
    _upload_counts: dict[str, dict] = {}  # user_id -> {"date": "YYYY-MM-DD", "count": N}

    def _check_upload_limit(user_id: str, plan: str) -> bool:
        import datetime
        today = datetime.date.today().isoformat()
        rec = _upload_counts.get(user_id, {"date": "", "count": 0})
        if rec["date"] != today:
            rec = {"date": today, "count": 0}
        limit = _UPLOAD_LIMITS.get(plan, 5)
        if limit == 0:
            return True
        return rec["count"] < limit

    def _inc_upload_count(user_id: str):
        import datetime
        today = datetime.date.today().isoformat()
        rec = _upload_counts.get(user_id, {"date": "", "count": 0})
        if rec["date"] != today:
            rec = {"date": today, "count": 0}
        rec["count"] += 1
        _upload_counts[user_id] = rec

    def _sanitize_filename(name: str) -> str:
        name = _re.sub(r'[^\w\-_. ]', '', name)
        return name.strip()[:100] or "arquivo"

    def _format_size(size: int) -> str:
        if size > 1024 * 1024:
            return f"{size / (1024*1024):.1f} MB"
        if size > 1024:
            return f"{size / 1024:.1f} KB"
        return f"{size} bytes"

    def _extract_text_docx(data: bytes) -> str:
        try:
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[Erro ao ler DOCX: {e}]"

    def _extract_text_xlsx(data: bytes) -> tuple[str, int]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            total_rows = 0
            for sheet in wb.worksheets:
                lines.append(f"\n### Aba: {sheet.title}\n")
                rows = list(sheet.iter_rows(values_only=True))
                total_rows += len(rows)
                if rows:
                    # Header
                    header = [str(c) if c is not None else "" for c in rows[0]]
                    lines.append("| " + " | ".join(header) + " |")
                    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                    for row in rows[1:51]:
                        cells = [str(c) if c is not None else "" for c in row]
                        lines.append("| " + " | ".join(cells) + " |")
                    if len(rows) > 51:
                        lines.append(f"\n... e mais {len(rows) - 51} linhas")
            wb.close()
            return "\n".join(lines), total_rows
        except Exception as e:
            return f"[Erro ao ler XLSX: {e}]", 0

    def _extract_text_csv(data: bytes) -> tuple[str, int]:
        import csv
        try:
            text = data.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            lines = []
            if rows:
                header = rows[0]
                lines.append("| " + " | ".join(header) + " |")
                lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                for row in rows[1:51]:
                    lines.append("| " + " | ".join(row) + " |")
                if len(rows) > 51:
                    lines.append(f"\n... e mais {len(rows) - 51} linhas")
            return "\n".join(lines), len(rows)
        except Exception as e:
            return f"[Erro ao ler CSV: {e}]", 0

    def _resize_image(data: bytes, max_px: int = 2000) -> bytes:
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(data))
            if max(img.size) > max_px:
                img.thumbnail((max_px, max_px), Image.LANCZOS)
                buf = io.BytesIO()
                fmt = img.format or "JPEG"
                if fmt.upper() == "WEBP":
                    img.save(buf, format="WEBP", quality=85)
                else:
                    img.save(buf, format="JPEG", quality=85)
                return buf.getvalue()
        except Exception:
            pass
        return data

    async def _transcribe_audio(file_path: str) -> str:
        """Transcricao de audio — agora feita no frontend via Web Speech API.
        Backend nao usa mais OpenAI Whisper. Retorna vazio para fallback."""
        return ""

    @app.post("/api/v1/upload")
    async def api_upload(request: Request, file: UploadFile = File(...), message: str = Form("")):
        """Upload de arquivo com processamento automatico."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        user_id = sess["user_id"]
        user_plan = sess.get("plan", "free")

        if not _check_upload_limit(user_id, user_plan):
            limit = _UPLOAD_LIMITS.get(user_plan, 5)
            return JSONResponse({"error": f"Limite de {limit} uploads/dia atingido. Faca upgrade do plano."}, status_code=429)

        # Lê arquivo
        data = await file.read()
        if len(data) > _MAX_FILE_SIZE:
            return JSONResponse({"error": "Arquivo muito grande. Maximo: 20MB"}, status_code=413)

        original_name = file.filename or "arquivo"
        safe_name = _sanitize_filename(original_name)
        ext = Path(original_name).suffix.lower()

        if ext in _BLOCKED_EXT:
            return JSONResponse({"error": f"Tipo de arquivo não permitido: {ext}"}, status_code=400)

        # Salva arquivo
        upload_dir = Path(__file__).parent.parent / "static" / "uploads" / user_id
        upload_dir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        saved_name = f"{ts}_{safe_name}"
        saved_path = upload_dir / saved_name
        saved_path.write_bytes(data)

        _inc_upload_count(user_id)
        file_size = len(data)
        file_url = f"/static/uploads/{user_id}/{saved_name}"

        result: dict[str, Any] = {
            "ok": True,
            "file_name": safe_name,
            "file_url": file_url,
            "file_size": _format_size(file_size),
            "file_ext": ext,
            "type": "unknown",
        }

        # Processa por tipo
        if ext in _ALLOWED_IMAGE_EXT:
            resized = _resize_image(data)
            # Detecta media_type real pelos magic bytes (apos resize)
            if resized[:3] == b'\xff\xd8\xff':
                media_type = 'image/jpeg'
            elif resized[:4] == b'\x89PNG':
                media_type = 'image/png'
            elif resized[:4] == b'GIF8':
                media_type = 'image/gif'
            elif resized[:4] == b'RIFF' and len(resized) > 11 and resized[8:12] == b'WEBP':
                media_type = 'image/webp'
            else:
                media_type = 'image/jpeg'
            b64 = base64.b64encode(resized).decode("ascii")
            result["type"] = "image"
            result["media_type"] = media_type
            result["base64"] = b64

        elif ext == ".pdf":
            b64 = base64.b64encode(data).decode("ascii")
            # Conta paginas
            pages = 0
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(data))
                pages = len(reader.pages)
            except Exception:
                pass
            if pages > 50:
                result["warning"] = f"PDF com {pages} paginas. Apenas as primeiras 50 serao analisadas."
            result["type"] = "pdf"
            result["media_type"] = "application/pdf"
            result["base64"] = b64
            result["pages"] = pages

        elif ext in _ALLOWED_SHEET_EXT:
            if ext == ".csv":
                text, rows = _extract_text_csv(data)
            else:
                text, rows = _extract_text_xlsx(data)
            result["type"] = "spreadsheet"
            result["extracted_text"] = text
            result["rows"] = rows

        elif ext in _ALLOWED_DOC_EXT:
            if ext == ".docx":
                text = _extract_text_docx(data)
            else:
                text = data.decode("utf-8", errors="replace")
            words = len(text.split())
            if words > 5000:
                text = " ".join(text.split()[:5000]) + f"\n\n... (truncado, {words} palavras total)"
            result["type"] = "document"
            result["extracted_text"] = text
            result["words"] = words

        elif ext in _ALLOWED_CODE_EXT:
            text = data.decode("utf-8", errors="replace")
            lang_map = {".py": "python", ".js": "javascript", ".ts": "typescript",
                        ".html": "html", ".css": "css", ".json": "json", ".jsx": "jsx", ".tsx": "tsx"}
            result["type"] = "code"
            result["extracted_text"] = text
            result["language"] = lang_map.get(ext, "text")

        elif ext in _ALLOWED_AUDIO_EXT:
            # Transcreve audio
            loop = asyncio.get_event_loop()
            transcription = await _transcribe_audio(str(saved_path))
            result["type"] = "audio"
            result["transcription"] = transcription
            result["has_transcription"] = bool(transcription and not transcription.startswith("[Erro"))

        else:
            result["type"] = "file"

        track_action("file_upload", f"{ext}: {safe_name}", "ok")
        return JSONResponse(result)

    # ── API: Usage ───────────────────────────────────────────────────

    @app.get("/api/v1/usage")
    async def api_usage(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        usage = get_user_usage_today(sess["user_id"])
        plan = PLANS.get(sess.get("plan", "free"), PLANS["free"])
        return JSONResponse({
            "usage": usage,
            "plan": sess.get("plan", "free"),
            "plan_label": plan["label"],
            "daily_limit": plan["daily_tokens"],
        })

    @app.get("/api/v1/me")
    async def api_me(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        plan = sess.get("plan", "free")
        is_admin = sess.get("is_admin", False)
        if is_admin:
            models = ["claude-code"]
        else:
            models = ["haiku"]
            if plan in ("pro", "unlimited"):
                models.append("sonnet")
        return JSONResponse({
            "email": sess["email"],
            "user_id": sess["user_id"],
            "is_admin": is_admin,
            "plan": plan,
            "available_models": models,
        })

    # ── API: Admin ───────────────────────────────────────────────────

    @app.get("/api/v1/admin/stats")
    async def api_admin_stats(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        return JSONResponse(get_admin_stats())

    @app.get("/api/v1/admin/users")
    async def api_admin_users(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        return JSONResponse({"users": list_users()})

    @app.post("/api/v1/admin/users/{user_id}")
    async def api_admin_update_user(user_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        body = await request.json()
        update_user(user_id, **body)
        return JSONResponse({"ok": True})

    @app.post("/api/v1/admin/create-user")
    async def api_admin_create_user(request: Request):
        sess = _get_user_session(request)
        if not sess or not sess.get("is_admin"):
            return JSONResponse({"error": "Acesso negado"}, status_code=403)
        body = await request.json()
        email = body.get("email", "").strip().lower()
        password = body.get("password", "")
        name = body.get("name", "")
        plan = body.get("plan", "free")
        if not email or len(password) < 6:
            return JSONResponse({"error": "Email e senha (min 6 chars) obrigatorios"}, status_code=400)
        user = create_user(email, password, name)
        if not user:
            return JSONResponse({"error": "Email ja cadastrado"}, status_code=400)
        if plan != "free":
            update_user(user["id"], plan=plan)
        return JSONResponse({"ok": True, "user": user})

    # ── API: Missions ────────────────────────────────────────────────
    _mission_progress: dict[str, list] = {}  # mission_id -> [events]

    @app.post("/api/v1/missions/plan")
    async def api_mission_plan(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        body = await request.json()
        description = body.get("description", "").strip()
        if not description:
            return JSONResponse({"error": "Descricao vazia"}, status_code=400)

        from .agents.mission_engine import plan_mission
        loop = asyncio.get_event_loop()
        try:
            plan = await loop.run_in_executor(None, plan_mission, description)
            return JSONResponse({"plan": plan})
        except Exception as e:
            return JSONResponse({"error": str(e)[:300]}, status_code=500)

    @app.post("/api/v1/missions/start")
    async def api_mission_start(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        body = await request.json()
        description = body.get("description", "")
        plan_data = body.get("plan", {})
        title = plan_data.get("title", description[:60])
        steps = plan_data.get("steps", [])

        if not steps:
            return JSONResponse({"error": "Plano sem etapas"}, status_code=400)

        from .database import create_mission
        mission_id = create_mission(sess["user_id"], title, description, steps)
        _mission_progress[mission_id] = []

        async def on_progress(mid, event_type, data):
            _mission_progress.setdefault(mid, []).append({
                "type": event_type, "data": data, "time": time.time()
            })

        from .agents.mission_engine import execute_mission
        asyncio.create_task(execute_mission(mission_id, sess["user_id"], on_progress))

        return JSONResponse({"mission_id": mission_id, "title": title, "total_steps": len(steps)})

    @app.get("/api/v1/missions/{mission_id}/progress")
    async def api_mission_progress(mission_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        after = float(request.query_params.get("after", "0"))
        events = _mission_progress.get(mission_id, [])
        new_events = [e for e in events if e["time"] > after]

        from .database import get_mission
        mission = get_mission(mission_id)
        status = mission["status"] if mission else "unknown"

        return JSONResponse({"status": status, "events": new_events})

    @app.get("/api/v1/missions")
    async def api_list_missions(request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        from .database import list_missions
        return JSONResponse({"missions": list_missions(sess["user_id"])})

    @app.get("/api/v1/missions/{mission_id}")
    async def api_get_mission(mission_id: str, request: Request):
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)
        from .database import get_mission
        m = get_mission(mission_id)
        if not m:
            return JSONResponse({"error": "Missão não encontrada"}, status_code=404)
        return JSONResponse({"mission": m})

    # ── HTTP Chat Fallback (para mobile sem WebSocket) ────────────
    _http_sessions: dict[str, Any] = {}

    @app.post("/api/v1/chat", dependencies=[Depends(_rate_limit_dependency)])
    async def api_chat(request: Request):
        """Chat HTTP com deteccao automatica de geracao de arquivos."""
        sess = _get_user_session(request)
        if not sess:
            return JSONResponse({"error": "Nao autenticado"}, status_code=401)

        # Checa limite de uso
        allowed, pct = check_limit(sess["user_id"])
        if not allowed:
            return JSONResponse({
                "session_id": "",
                "response": "Voce atingiu seu limite diario de tokens. Volte amanha ou faca upgrade do seu plano.\n\nUse `/plan` para ver seu plano atual.",
                "tools": [], "file": None,
            })

        from .agent import Agent
        import uuid

        body = await request.json()
        content = body.get("content", "").strip()
        conv_id = body.get("conversation_id", "")
        session_id = body.get("session_id", "")
        chosen_model = body.get("model", "haiku")  # haiku ou sonnet
        file_data = body.get("file_data")

        if not content and not file_data:
            return JSONResponse({"error": "content vazio"}, status_code=400)

        user_email = sess["email"]
        user_id = sess["user_id"]
        user_plan = sess.get("plan", "free")
        is_admin = sess.get("is_admin", False)

        # Admin SEMPRE usa Claude Code CLI (conta Max, gratis)
        if is_admin:
            from .claude_code_bridge import ask_claude_code, log_claude_code_usage
            track_action("user_message_claude_code", content[:60])

            if conv_id:
                save_message(conv_id, "user", content)

            loop = asyncio.get_event_loop()
            response_text, elapsed = await loop.run_in_executor(None, lambda: ask_claude_code(content, "/root/clow/workspace", conv_id))

            log_claude_code_usage(user_id, content, elapsed)
            track_action("claude_code_response", response_text[:60] if response_text else "")

            if conv_id:
                save_message(conv_id, "assistant", response_text)

            return JSONResponse({
                "session_id": session_id or str(uuid.uuid4())[:8],
                "response": response_text,
                "tools": [], "file": None,
            })

        # Nao-admin: valida modelo pelo plano
        from .generators.base import MODELS as AI_MODELS
        allowed_models = ["haiku"]
        if user_plan in ("pro", "unlimited"):
            allowed_models.append("sonnet")
        if chosen_model not in allowed_models:
            chosen_model = "haiku"
        model_id = AI_MODELS.get(chosen_model, AI_MODELS["haiku"])
        track_action("user_message_http", content[:60])

        # Salva mensagem do usuario no historico
        if conv_id:
            save_message(conv_id, "user", content)

        # ── Comandos internos ──
        if content.startswith("/"):
            cmd_lower = content.lower().strip()
            cmd_resp = None

            if cmd_lower.startswith("/skills"):
                from .skills.loader import format_skills_list
                cat = content[7:].strip()
                cmd_resp = format_skills_list(cat)
            elif cmd_lower == "/memories":
                from .memory_web import format_memories_list
                cmd_resp = format_memories_list(user_id)
            elif cmd_lower.startswith("/forget"):
                from .memory_web import forget_memory
                kw = content[7:].strip()
                cmd_resp = forget_memory(user_id, kw) if kw else "Use: `/forget palavra-chave`"
            elif cmd_lower.startswith("/mission"):
                mission_desc = content[8:].strip() if len(content) > 8 else ""
                if not mission_desc:
                    cmd_resp = (
                        "## Missões Autônomas\n\n"
                        "Descreva uma missao complexa e o Clow executa sozinho:\n\n"
                        "**Exemplos:**\n"
                        "- `/mission Cria um site completo para uma pizzaria com cardapio e contato`\n"
                        "- `/mission Campanha de tráfego para seguro de vida com landing page e copies`\n"
                        "- `/mission Setup digital completo para barbearia`\n\n"
                        "O Clow vai planejar, mostrar as etapas, e executar tudo automaticamente."
                    )
                else:
                    # Planeja e inicia missao
                    from .agents.mission_engine import plan_mission
                    loop = asyncio.get_event_loop()
                    try:
                        plan_data = await loop.run_in_executor(None, plan_mission, mission_desc)
                        steps = plan_data.get("steps", [])
                        title = plan_data.get("title", mission_desc[:60])
                        est = plan_data.get("estimated_minutes", 5)

                        # Cria e inicia missao
                        from .database import create_mission as db_create_mission
                        mid = db_create_mission(user_id, title, mission_desc, steps)
                        _mission_progress[mid] = []

                        async def on_progress(m_id, evt, data):
                            _mission_progress.setdefault(m_id, []).append({"type": evt, "data": data, "time": time.time()})

                        from .agents.mission_engine import execute_mission
                        asyncio.create_task(execute_mission(mid, user_id, on_progress))

                        # Mostra plano e inicia
                        steps_text = "\n".join(f"{i+1}. {s.get('title', '?')}" for i, s in enumerate(steps))
                        cmd_resp = (
                            f"## Missão Iniciada\n\n"
                            f"**{title}**\n\n"
                            f"### Plano ({len(steps)} etapas, ~{est} min):\n{steps_text}\n\n"
                            f"Executando em background... Acompanhe o progresso abaixo."
                        )

                        # Retorna com mission_id pra frontend fazer polling
                        if conv_id:
                            save_message(conv_id, "assistant", cmd_resp)
                        return JSONResponse({
                            "session_id": session_id or str(uuid.uuid4())[:8],
                            "response": cmd_resp,
                            "tools": [], "file": None,
                            "mission": {"id": mid, "title": title, "total_steps": len(steps)},
                        })
                    except Exception as e:
                        cmd_resp = f"Erro ao planejar missao: {str(e)[:200]}"

            elif cmd_lower == "/help":
                cmd_resp = (
                    "## Comandos Disponiveis\n\n"
                    "| Comando | Descricao |\n|---------|----------|\n"
                    "| `/mission X` | Iniciar missao autonoma |\n"
                    "| `/skills` | Listar skills disponiveis |\n"
                    "| `/memories` | Ver memorias salvas |\n"
                    "| `/forget X` | Esquecer memoria |\n"
                    "| `/connect` | Conectar servico externo |\n"
                    "| `/connections` | Ver conexões ativas |\n"
                    "| `/disconnect X` | Desconectar servico |\n"
                    "| `/usage` | Ver consumo de tokens hoje |\n"
                    "| `/plan` | Ver plano atual e limites |\n"
                    "| `/help` | Esta lista de comandos |\n\n"
                    "**Missões:** `/mission cria um site completo para pizzaria`\n\n"
                    "**Geracao de arquivos:** peca naturalmente (ex: 'cria uma planilha de vendas')\n\n"
                    "**Integracoes:** pergunte direto (ex: 'mostra minhas campanhas meta ads')"
                )
            elif cmd_lower == "/usage":
                usage = get_user_usage_today(user_id)
                plan_info = PLANS.get(sess.get("plan", "free"), PLANS["free"])
                limit = plan_info["daily_tokens"]
                used = usage["total_tokens"]
                pct_str = f"{(used/limit*100):.0f}%" if limit > 0 else "ilimitado"
                cmd_resp = (
                    f"## Seu Consumo Hoje\n\n"
                    f"- Tokens usados: **{used:,}**\n"
                    f"- Limite diario: **{limit:,}** ({pct_str})\n"
                    f"- Requests: **{usage['requests']}**\n"
                    f"- Custo estimado: **${usage['total_cost']:.4f}**"
                )
            elif cmd_lower == "/plan":
                plan_info = PLANS.get(sess.get("plan", "free"), PLANS["free"])
                cmd_resp = (
                    f"## Seu Plano: {plan_info['label']}\n\n"
                    f"- Limite diario: **{plan_info['daily_tokens']:,} tokens**\n\n"
                    "**Planos disponiveis:**\n"
                    "| Plano | Tokens/dia |\n|-------|------------|\n"
                    "| Free | 50.000 |\n| Basic | 200.000 |\n| Pro | 1.000.000 |\n| Unlimited | Sem limite |"
                )

            if cmd_resp:
                if conv_id:
                    save_message(conv_id, "assistant", cmd_resp)
                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": cmd_resp, "tools": [], "file": None,
                })

            # /connect, /disconnect, /connections
            from .integrations.command_handler import handle_command
            cmd_result = handle_command(content, user_email)
            if cmd_result:
                if conv_id:
                    save_message(conv_id, "assistant", cmd_result["response"])
                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": cmd_result["response"],
                    "tools": [], "file": None,
                })

        # ── Detecta pedidos de integracao (meta ads, supabase, etc) ──
        from .integrations.command_handler import detect_integration_request
        int_result = detect_integration_request(content, user_email)
        if int_result:
            return JSONResponse({
                "session_id": session_id or str(uuid.uuid4())[:8],
                "response": int_result["response"],
                "tools": [],
                "file": None,
            })

        # ── Detecta geracao de arquivo ──
        from .generators.dispatcher import detect, run_generator
        gen_module, gen_type = detect(content)

        if gen_module:
            loop = asyncio.get_event_loop()
            try:
                result = await loop.run_in_executor(None, run_generator, gen_module, content, model_id, user_id)
                track_action("file_generated", f"{gen_type}: {result.get('name', '')}", "ok")

                if result.get("type") == "text":
                    return JSONResponse({
                        "session_id": session_id or str(uuid.uuid4())[:8],
                        "response": result["content"],
                        "tools": [],
                        "file": None,
                    })

                # Formata tamanho
                size_raw = result.get("size", 0)
                if isinstance(size_raw, str):
                    size_str = size_raw
                elif isinstance(size_raw, (int, float)) and size_raw > 1024 * 1024:
                    size_str = f"{size_raw / (1024*1024):.1f} MB"
                elif isinstance(size_raw, (int, float)) and size_raw > 1024:
                    size_str = f"{size_raw / 1024:.1f} KB"
                else:
                    size_str = f"{size_raw} bytes"

                type_labels = {
                    "landing_page": "Landing Page",
                    "app": "Web App",
                    "xlsx": "Planilha Excel",
                    "docx": "Documento Word",
                    "pptx": "Apresentação PowerPoint",
                }
                type_label = type_labels.get(result["type"], result["type"])
                msg = f"Pronto! Aqui esta seu arquivo:\n\n**{type_label}** — {result['name']} ({size_str})"

                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": msg,
                    "tools": [],
                    "file": {
                        "type": result["type"],
                        "name": result["name"],
                        "url": result["url"],
                        "size": size_str,
                    },
                })
            except Exception as e:
                track_action("file_gen_error", str(e)[:60], "error")
                return JSONResponse({
                    "session_id": session_id or str(uuid.uuid4())[:8],
                    "response": f"Erro ao gerar arquivo: {str(e)}",
                    "tools": [],
                    "file": None,
                }, status_code=500)

        # ── Injetar skills no prompt ──
        from .skills.loader import build_skill_prompt
        skill_context = build_skill_prompt(content)
        if skill_context:
            content = f"[CONTEXTO DE SKILLS ATIVAS - siga estas instrucoes]\n{skill_context}\n[FIM DO CONTEXTO]\n\nPedido do usuario: {content}"

        # ── Chat normal via Agent ──
        session_key = f"{session_id}_{chosen_model}"
        if session_id and session_key in _http_sessions:
            agent = _http_sessions[session_key]["agent"]
        else:
            session_id = str(uuid.uuid4())[:8]
            session_key = f"{session_id}_{chosen_model}"
            agent = Agent(cwd=os.getcwd(), model=model_id, auto_approve=True)
            _http_sessions[session_key] = {"agent": agent, "last_used": time.time()}

        _http_sessions[session_key]["last_used"] = time.time()

        now = time.time()
        stale = [k for k, v in _http_sessions.items() if now - v["last_used"] > 1800]
        for k in stale:
            del _http_sessions[k]

        loop = asyncio.get_event_loop()
        collected_text: list[str] = []
        tools_used: list[dict] = []

        def on_text_delta(delta: str):
            collected_text.append(delta)

        def on_tool_call(name: str, args: dict):
            tools_used.append({"name": name, "args": args, "status": "running", "output": ""})
            track_action("tool_call", name, "running")

        def on_tool_result(name: str, status: str, output: str):
            for t in tools_used:
                if t["name"] == name and t["status"] == "running":
                    t["status"] = status
                    t["output"] = output[:500]
                    break
            track_action("tool_result", f"{name}: {status}", status)

        agent.on_text_delta = on_text_delta
        agent.on_text_done = lambda t: None
        agent.on_tool_call = on_tool_call
        agent.on_tool_result = on_tool_result

        # Monta mensagem multimodal se tem arquivo
        if file_data:
            user_msg = _build_multimodal_message(content, file_data)
        else:
            user_msg = content

        try:
            result = await loop.run_in_executor(None, agent.run_turn, user_msg)
        except Exception as e:
            return JSONResponse({
                "session_id": session_id,
                "error": str(e),
            }, status_code=500)

        response_text = "".join(collected_text) if collected_text else (result or "")
        track_action("agent_response_http", response_text[:60] if response_text else "")

        return JSONResponse({
            "session_id": session_id,
            "response": response_text,
            "tools": tools_used,
            "file": None,
        })

    # ── Helpers para geração de imagens ────────────────────────────

    def _should_generate_image(content: str) -> bool:
        """Detecta se o pedido é para gerar imagem."""
        keywords = [
            "gera imagem", "cria imagem", "faz uma imagem", "gerar imagem",
            "criar imagem", "desenha", "ilustração", "arte", "visual",
            "criativo visual", "banner", "thumbnail", "foto", "picture", "image",
            "criativo pra", "criativo de", "criativo para"
        ]
        content_lower = content.lower()
        return any(kw in content_lower for kw in keywords)

    async def _process_image_request(content: str, agent) -> tuple[str | None, str | None, str]:
        """
        Processa pedido de imagem:
        1. Otimiza prompt em inglês via Claude
        2. Gera imagem via Pollinations
        3. Retorna (filepath, filename, resposta_formatada)
        """
        from .generators.image_gen import optimize_prompt_for_image, generate_image
        
        # Step 1: Otimiza prompt
        try:
            optimized_prompt = optimize_prompt_for_image(content, agent._anthropic)
            if not optimized_prompt:
                optimized_prompt = content
        except Exception as e:
            optimized_prompt = content
            logging.error(f"Erro ao otimizar prompt: {e}")

        # Step 2: Gera imagem
        try:
            filepath, filename = await asyncio.get_event_loop().run_in_executor(
                None, lambda: generate_image(optimized_prompt, 1024, 1024)
            )
            
            if filepath and filename:
                html = f'''<div style="margin: 12px 0;">
  <img src="/static/files/{filename}" style="max-width:400px;border-radius:12px;cursor:pointer;border:1px solid #ddd;" onclick="window.open(this.src)">
  <div style="margin-top:12px;display:flex;gap:8px;align-items:center;">
    <a href="/static/files/{filename}" download style="padding:8px 16px;background:#5b5fc7;color:#fff;border-radius:6px;text-decoration:none;font-size:12px;cursor:pointer;">⬇️ Baixar</a>
    <span style="font-size:11px;color:#999;">Prompt: {optimized_prompt[:60]}...</span>
  </div>
</div>'''
                response = f"✨ Pronto! Aqui está sua imagem.\n\n{html}"
                return filepath, filename, response
            else:
                return None, None, "⏳ A geração de imagem demorou mais que o esperado. Tente novamente ou peça um briefing visual que eu monto pra você usar no Canva."
        except Exception as e:
            logging.error(f"Erro ao gerar imagem: {e}")
            return None, None, f"❌ Erro ao gerar imagem: {str(e)}"

    @app.websocket("/ws")
    async def websocket_endpoint(websocket: WebSocket):
        # Verificacao de sessao via cookie para WebSocket
        ws_cookie = websocket.cookies.get("clow_session", "")
        ws_sess = _validate_session(ws_cookie)
        if not ws_sess:
            # Fallback: API key via query param
            api_key = websocket.query_params.get("api_key", "")
            keys = _get_api_keys()
            if keys and not _verify_api_key(api_key):
                await websocket.close(code=4001, reason="Nao autenticado")
                return
            elif not keys and not ws_cookie:
                await websocket.close(code=4001, reason="Nao autenticado")
                return

        ws_is_admin = ws_sess.get("is_admin", False) if ws_sess else False
        ws_user_id = ws_sess.get("user_id", "") if ws_sess else ""

        # Rate limit para WebSocket
        client_ip = websocket.client.host if websocket.client else "unknown"
        if not _ws_rate_limiter.is_allowed(client_ip):
            await websocket.close(code=4029, reason="Rate limit excedido")
            return

        await websocket.accept()

        from .agent import Agent

        loop = asyncio.get_event_loop()

        # Cria agente com callbacks que enviam via WebSocket
        send_queue: asyncio.Queue = asyncio.Queue()

        def on_text_delta(delta: str):
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "text_delta", "content": delta}),
                loop,
            )

        def on_text_done(text: str):
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "text_done"}),
                loop,
            )

        def on_tool_call(name: str, args: dict):
            track_action("tool_call", f"{name}", "running")
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "tool_call", "name": name, "args": args}),
                loop,
            )

        def on_tool_result(name: str, status: str, output: str):
            track_action("tool_result", f"{name}: {status}", status)
            asyncio.run_coroutine_threadsafe(
                send_queue.put({"type": "tool_result", "name": name, "status": status, "output": output[:2000]}),
                loop,
            )

        agent = Agent(
            cwd=os.getcwd(),
            on_text_delta=on_text_delta,
            on_text_done=on_text_done,
            on_tool_call=on_tool_call,
            on_tool_result=on_tool_result,
            auto_approve=True,
        )

        # Task para enviar mensagens da fila para o WebSocket
        async def send_loop():
            try:
                while True:
                    msg = await send_queue.get()
                    await websocket.send_json(msg)
            except Exception:
                pass

        sender_task = asyncio.create_task(send_loop())

        try:
            while True:
                data = await websocket.receive_json()
                if data.get("type") == "message":
                    content = data.get("content", "")
                    file_data = data.get("file_data")
                    ws_model = data.get("model", "haiku")

                    if not content and not file_data:
                        continue

                    track_action("user_message", content[:60])

                    # Envia thinking
                    await websocket.send_json({"type": "thinking_start"})

                    # ── Admin SEMPRE usa Claude Code CLI com streaming (Opus via Max) ──
                    if ws_is_admin:
                        await websocket.send_json({"type": "thinking_end"})

                        from .claude_code_bridge import ask_claude_code_stream, log_claude_code_usage
                        track_action("user_message_claude_code", content[:60])

                        def _on_delta(delta: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "text_delta", "content": delta}),
                                loop,
                            )

                        def _on_done(full: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "text_done"}),
                                loop,
                            )

                        def _on_error(err: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "error", "content": err}),
                                loop,
                            )

                        def _on_tool_call(name: str, args: dict):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "tool_call", "name": name, "args": args}),
                                loop,
                            )

                        def _on_tool_result(name: str, status: str, output: str):
                            asyncio.run_coroutine_threadsafe(
                                send_queue.put({"type": "tool_result", "name": name, "status": status, "output": output}),
                                loop,
                            )

                        ws_conv_id = data.get("conversation_id", "")
                        try:
                            elapsed = await loop.run_in_executor(
                                None, lambda: ask_claude_code_stream(
                                    content, _on_delta, _on_done, _on_error,
                                    "/root/clow/workspace", _on_tool_call, _on_tool_result,
                                    ws_conv_id
                                )
                            )
                            log_claude_code_usage(ws_user_id, content, elapsed)
                            track_action("claude_code_response_stream", f"{elapsed:.1f}s")
                        except Exception as e:
                            await websocket.send_json({"type": "error", "content": str(e)})
                            track_action("claude_code_error", str(e)[:60], "error")

                        await websocket.send_json({"type": "turn_complete"})
                        continue

                    # ── Detecta e processa pedido de imagem ──
                    if _should_generate_image(content) and not file_data:
                        await websocket.send_json({"type": "thinking_end"})
                        
                        # Gera imagem
                        try:
                            filepath, filename, response_html = await _process_image_request(content, agent)
                            await websocket.send_json({"type": "text_delta", "content": response_html})
                            await websocket.send_json({"type": "text_done"})
                            track_action("image_generated", filename or "failed")
                        except Exception as e:
                            await websocket.send_json({"type": "error", "content": f"Erro ao gerar imagem: {str(e)}"})
                            track_action("image_error", str(e)[:60], "error")
                        
                        # Finaliza turno
                        await websocket.send_json({"type": "turn_complete"})
                        continue

                    # Monta mensagem multimodal se tem arquivo
                    if file_data:
                        user_msg = _build_multimodal_message(content, file_data)
                    else:
                        user_msg = content

                    # Executa agente em thread separada (chat normal)
                    try:
                        result = await loop.run_in_executor(
                            None, agent.run_turn, user_msg
                        )
                        track_action("agent_response", result[:60] if result else "")
                    except Exception as e:
                        await websocket.send_json({"type": "thinking_end"})
                        await websocket.send_json({"type": "error", "content": str(e)})
                        track_action("agent_error", str(e)[:60], "error")

                    # Finaliza turno
                    await websocket.send_json({"type": "turn_complete"})

        except WebSocketDisconnect:
            pass
        except Exception:
            pass
        finally:
            sender_task.cancel()


# ── Mount de arquivos estáticos ──────────────────────────────────
if app and HAS_FASTAPI:
    try:
        # Garante que diretório existe
        os.makedirs("/root/clow/static/files", exist_ok=True)
        # Mount da pasta /static
        app.mount("/static", StaticFiles(directory="/root/clow/static"), name="static")
    except Exception as e:
        logging.error(f"Erro ao montar /static: {e}")
